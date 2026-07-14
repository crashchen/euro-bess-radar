"""Contract pins for the contracted-floor CF-B cockpit increment.

Drives the real panel with Streamlit AppTest and pins the locked contract's
gating, source provenance, state invalidation, non-additive chart isolation,
and self-contained Excel export.
"""

from __future__ import annotations

import ast
import inspect
import json
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

import src.pages.simulation_cockpit as cockpit
from src.contracted_floor import (
    MAX_FLOOR_TRAJECTORY_YEARS,
    compute_decaying_contracted_floor_overlay,
)
from src.export import cockpit_tables_to_excel
from src.pages.simulation_cockpit import (
    _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION,
    _CONTRACTED_FLOOR_HARD_CAPTION,
    _CONTRACTED_FLOOR_LIQUIDITY_SOURCE_LABEL,
    _CONTRACTED_FLOOR_MINIMUM_SHARE_HELP,
    _CONTRACTED_FLOOR_NEGATIVE_YEAR_CAPTION,
    _CONTRACTED_FLOOR_SOURCE_LABEL,
    _CONTRACTED_FLOOR_TRAJECTORY_CAPTION,
    _append_contracted_floor_assumptions,
    _contracted_floor_best_row,
    _contracted_floor_export_table,
    _contracted_floor_fingerprint,
    _contracted_floor_source_label,
)

_EXPANDER_TITLE = "Contracted floor versus merchant cash flow"


def test_contract_locked_copy_is_verbatim() -> None:
    """Anchor the section-3 source label and section-5 hard caption literally.

    The AppTest assertions compare rendered captions against the module
    constants; without this anchor a reworded constant would keep every
    test green while drifting from the locked contract copy.
    """
    assert _CONTRACTED_FLOOR_SOURCE_LABEL == (
        "DA-only merchant net after linear wear - cycle-frontier best cap"
    )
    assert _CONTRACTED_FLOOR_LIQUIDITY_SOURCE_LABEL == (
        "DA-only merchant net after linear wear (liquidity-capped) - "
        "cycle-frontier best cap"
    )
    assert _CONTRACTED_FLOOR_HARD_CAPTION == (
        "Screening floor overlay, not a binding contract model; DA only; "
        "linear wear proxy; no credit, performance, tax, financing, or "
        "post-term merchant assumption."
    )
    assert _CONTRACTED_FLOOR_TRAJECTORY_CAPTION == (
        "Merchant trajectory: projected by this panel's own decay and "
        "floor-escalation inputs (flat when both are inactive); the Risk "
        "Analysis revenue-decay assumption is never used in this calculation."
    )
    assert _CONTRACTED_FLOOR_NEGATIVE_YEAR_CAPTION == (
        "Negative merchant years: under the wear-net settlement basis the "
        "annual top-up exceeds the floor itself in those years; a revenue-"
        "settled contract would cap the top-up at the floor (settlement basis "
        "is deferred, section 9 of this contract)."
    )
    assert _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION == (
        "Decaying-merchant floor composition: screening projection of the "
        "frontier's year-1 merchant net using this panel's decay and escalation "
        "inputs; wear stays flat inside the trajectory, so late merchant years "
        "can be negative rather than idled. No re-dispatch, credit, tax, "
        "financing, or post-term value is modelled. PVs cover the contract "
        "window only."
    )
    assert (
        "Merchant baseline: flat annual revenue across the tenor"
        not in inspect.getsource(cockpit)
    )


def _context(
    *, merchant_per_mw: float = 40000.0, avg_efc: float = 1.2,
    valid_days: int = 30, frontier_token: str = "frontier-v1",
    liquidity: bool = False, gross_per_mw: float = 50000.0,
) -> dict:
    frontier = pd.DataFrame({
        "cycle_cap": [1.2, float("nan")],
        "label": ["1.2 EFC/day", "uncapped"],
        "gross_eur": [3500.0, 3600.0],
        "avg_efc_per_day": [avg_efc, 1.8],
        "wear_eur": [500.0, 900.0],
        "net_eur": [3000.0, 2700.0],
        "gross_eur_per_mw_yr": [gross_per_mw, 48000.0],
        "net_eur_per_mw_yr": [merchant_per_mw, 36000.0],
    })
    context = {
        "fingerprint": (frontier_token, "DE_LU", valid_days),
        "frontier": frontier,
        "summary": {
            "best_cap_label": "1.2 EFC/day",
            "cycle_life": 6000.0,
            "capex_eur_kwh": 150.0,
            "valid_days": valid_days,
            "frontier_flat": False,
            "n_tiebreak_fallback_days": 0,
        },
        "sweep_dates": ("2026-03-01", "2026-03-30"),
        "primary_zone": "DE_LU",
        "power_mw": 1.0,
        "duration_hours": 2.0,
        "cycle_life": 6000.0,
        "capex_eur_kwh": 150.0,
    }
    if liquidity:
        context["liquidity"] = {
            "power_mw": 1.0,
            "zone_da_volume_mw": 5.0,
            "max_participation_share": 0.1,
            "executable_power_mw": 0.5,
            "participation_at_full_power": 0.2,
            "binding": True,
        }
    return context


def _result() -> dict[str, float]:
    return {
        "merchant_net_eur": 40000.0,
        "merchant_net_eur_per_mw_yr": 40000.0,
        "quoted_floor_eur": 50000.0,
        "effective_floor_eur": 40000.0,
        "effective_floor_eur_per_mw_yr": 40000.0,
        "floor_protected_cashflow_eur": 40000.0,
        "annual_top_up_eur": 0.0,
        "merchant_pv_eur": 268400.0,
        "floor_protected_pv_eur": 268400.0,
        "floor_pv_uplift_eur": 0.0,
        "floor_tenor_years": 10.0,
        "discount_rate": 0.08,
        "contract_availability": 0.8,
    }


def _active_result(
    *, decay: float = 0.5, decay_floor: float = 0.0,
    escalation: float = 0.0, tenor: float = 3.0,
) -> dict[str, object]:
    return compute_decaying_contracted_floor_overlay(
        merchant_net_eur_per_mw_yr=40000.0,
        merchant_gross_eur_per_mw_yr=50000.0,
        power_mw=1.0,
        quoted_floor_eur_per_mw_yr=50000.0,
        floor_tenor_years=tenor,
        contract_availability=0.8,
        discount_rate=0.08,
        annual_decay_rate=decay,
        decay_floor_share=decay_floor,
        floor_escalation_rate=escalation,
    )


class TestContractedFloorHelpers:
    def test_best_row_is_selected_by_frontier_summary(self) -> None:
        best = _contracted_floor_best_row(_context())
        assert best["label"] == "1.2 EFC/day"
        assert best["net_eur_per_mw_yr"] == 40000.0

    def test_missing_or_duplicate_best_row_raises(self) -> None:
        missing = _context()
        missing["summary"]["best_cap_label"] = "missing"
        with pytest.raises(ValueError, match="unique best-cap"):
            _contracted_floor_best_row(missing)

        duplicate = _context()
        duplicate["frontier"] = pd.concat(
            [duplicate["frontier"], duplicate["frontier"].iloc[[0]]],
            ignore_index=True,
        )
        with pytest.raises(ValueError, match="unique best-cap"):
            _contracted_floor_best_row(duplicate)

    def test_source_label_changes_only_when_liquidity_cap_binds(self) -> None:
        assert _contracted_floor_source_label(_context()) == (
            _CONTRACTED_FLOOR_SOURCE_LABEL
        )
        assert _contracted_floor_source_label(_context(liquidity=True)) == (
            _CONTRACTED_FLOOR_LIQUIDITY_SOURCE_LABEL
        )
        nonbinding = _context(liquidity=True)
        nonbinding["liquidity"]["binding"] = False
        nonbinding["liquidity"]["executable_power_mw"] = 1.0
        assert _contracted_floor_source_label(nonbinding) == (
            _CONTRACTED_FLOOR_SOURCE_LABEL
        )

    def test_identical_fingerprints_match(self) -> None:
        kwargs = dict(
            quoted_floor_eur_per_mw_yr=50000.0,
            contract_availability=0.8,
            floor_tenor_years=10.0,
            discount_rate=0.08,
            annual_decay_rate=0.0,
            decay_floor_share=0.0,
            floor_escalation_rate=0.0,
        )
        assert _contracted_floor_fingerprint(
            frontier_context=_context(), **kwargs,
        ) == _contracted_floor_fingerprint(
            frontier_context=_context(), **kwargs,
        )

    @pytest.mark.parametrize(
        ("field", "value"),
        [
            ("quoted_floor_eur_per_mw_yr", 60000.0),
            ("contract_availability", 0.75),
            ("floor_tenor_years", 12.0),
            ("discount_rate", 0.1),
            ("annual_decay_rate", 0.1),
            ("decay_floor_share", 0.4),
            ("floor_escalation_rate", 0.02),
        ],
    )
    def test_any_contract_input_invalidates(self, field: str, value: float) -> None:
        kwargs = dict(
            quoted_floor_eur_per_mw_yr=50000.0,
            contract_availability=0.8,
            floor_tenor_years=10.0,
            discount_rate=0.08,
            annual_decay_rate=0.0,
            decay_floor_share=0.0,
            floor_escalation_rate=0.0,
        )
        base = _contracted_floor_fingerprint(
            frontier_context=_context(), **kwargs,
        )
        kwargs[field] = value
        changed = _contracted_floor_fingerprint(
            frontier_context=_context(), **kwargs,
        )
        assert changed != base

    @pytest.mark.parametrize(
        "context",
        [
            _context(frontier_token="new-frontier"),
            _context(merchant_per_mw=41000.0),
            _context(gross_per_mw=51000.0),
            _context(avg_efc=1.3),
            _context(valid_days=29),
        ],
    )
    def test_frontier_input_or_recomputed_result_invalidates(
        self, context: dict,
    ) -> None:
        kwargs = dict(
            quoted_floor_eur_per_mw_yr=50000.0,
            contract_availability=0.8,
            floor_tenor_years=10.0,
            discount_rate=0.08,
            annual_decay_rate=0.0,
            decay_floor_share=0.0,
            floor_escalation_rate=0.0,
        )
        base = _contracted_floor_fingerprint(
            frontier_context=_context(), **kwargs,
        )
        changed = _contracted_floor_fingerprint(
            frontier_context=context, **kwargs,
        )
        assert changed != base


class TestContractedFloorExport:
    def test_assumptions_are_self_contained_without_global_table(self) -> None:
        rows = _append_contracted_floor_assumptions(
            None, frontier_context=_context(), result=_result(),
        )
        parameters = set(rows["parameter"])
        expected = {
            "Contract merchant baseline",
            "Contract power basis",
            "Quoted contracted floor",
            "Contract availability",
            "Effective contracted floor",
            "Floor term",
            "Floor discount rate",
            "Frontier best cap",
            "Frontier realised cycling",
            "Frontier cycle life",
            "Frontier BESS duration",
            "Frontier capex basis",
            "Frontier valid days",
            "Frontier annualisation",
            "Frontier window",
            "Frontier zone",
            "Frontier flat",
            "Frontier tie-break fallback days",
        }
        assert parameters == expected
        baseline = rows[rows["parameter"] == "Contract merchant baseline"].iloc[0]
        assert baseline["source"] == _CONTRACTED_FLOOR_SOURCE_LABEL
        assert baseline["value"] == "40000.00"
        availability = rows[rows["parameter"] == "Contract availability"].iloc[0]
        assert availability["value"] == "80.00%"

    def test_appends_to_existing_assumptions_without_mutating_them(self) -> None:
        original = pd.DataFrame([{
            "parameter": "Power", "value": "1", "unit": "MW",
            "source": "Sidebar", "affects": "Everything",
        }])
        before = original.copy(deep=True)
        rows = _append_contracted_floor_assumptions(
            original, frontier_context=_context(), result=_result(),
        )
        pd.testing.assert_frame_equal(original, before)
        assert len(rows) == len(original) + 18

    def test_liquidity_capped_baseline_inherits_full_provenance(self) -> None:
        context = _context(liquidity=True)
        rows = _append_contracted_floor_assumptions(
            None, frontier_context=context, result=_result(),
        )
        by_parameter = rows.set_index("parameter")
        assert by_parameter.loc["Contract merchant baseline", "source"] == (
            _CONTRACTED_FLOOR_LIQUIDITY_SOURCE_LABEL
        )
        assert by_parameter.loc["Contract power basis", "value"] == "1"
        assert by_parameter.loc[
            "Inherited user-entered zone DA volume", "value"
        ] == "5"
        assert by_parameter.loc[
            "Inherited maximum DA participation share", "value"
        ] == "10.00%"
        assert by_parameter.loc[
            "Inherited liquidity executable power", "value"
        ] == "0.5"
        assert by_parameter.loc[
            "Inherited liquidity cap binding", "value"
        ] == "True"
        export = _contracted_floor_export_table(
            frontier_context=context, result=_result(),
        )
        assert export.loc[0, "merchant_baseline_source"] == (
            _CONTRACTED_FLOOR_LIQUIDITY_SOURCE_LABEL
        )

    def test_excel_contains_contract_table_and_complete_provenance(self) -> None:
        context = _context()
        result = _result()
        table = _contracted_floor_export_table(
            frontier_context=context, result=result,
        )
        assumptions = _append_contracted_floor_assumptions(
            None, frontier_context=context, result=result,
        )
        data = cockpit_tables_to_excel(
            {"Contracted floor": table}, assumptions=assumptions,
        )
        workbook = load_workbook(BytesIO(data))
        assert workbook.sheetnames == ["Contracted floor", "Assumptions"]
        output_headers = [cell.value for cell in workbook["Contracted floor"][1]]
        assert "merchant_baseline_source" in output_headers
        assert "floor_protected_pv_eur" in output_headers
        assumption_values = {
            row[0].value: row[1].value
            for row in workbook["Assumptions"].iter_rows(min_row=2)
        }
        assert assumption_values["Contract merchant baseline"] == "40000.00"
        assert assumption_values["Frontier best cap"] == "1.2 EFC/day"

    def test_active_excel_contains_per_year_reconciliation_and_inputs(self) -> None:
        context = _context()
        result = _active_result()
        table = _contracted_floor_export_table(
            frontier_context=context, result=result,
        )
        assumptions = _append_contracted_floor_assumptions(
            None, frontier_context=context, result=result,
        )
        data = cockpit_tables_to_excel(
            {"Contracted floor": table}, assumptions=assumptions,
        )
        workbook = load_workbook(BytesIO(data), data_only=True)
        sheet = workbook["Contracted floor"]
        headers = [cell.value for cell in sheet[1]]
        assert sheet.max_row == 4
        assert "discount_factor" in headers
        assert "year_fraction" in headers
        assert "top_up_eur" in headers
        assert "per_year" not in headers
        values = {
            row[0].value: row[1].value
            for row in workbook["Assumptions"].iter_rows(min_row=2)
        }
        assert values["Annual merchant revenue decay"] == "50.00%"
        assert values["Minimum merchant share"] == "0.00%"
        assert values["Floor escalation"] == "0.00%"
        assert values["Floor crossover year"] == "2"
        assert values["Floor binding evaluated years"] == "2 of 3"


def test_render_places_floor_immediately_after_frontier() -> None:
    tree = ast.parse(inspect.getsource(cockpit))
    render_fn = next(
        node for node in tree.body
        if isinstance(node, ast.FunctionDef) and node.name == "render"
    )

    def _call_name(statement: ast.stmt) -> str | None:
        value = None
        if isinstance(statement, (ast.Expr, ast.Assign)):
            value = statement.value
        if isinstance(value, ast.Call) and isinstance(value.func, ast.Name):
            return value.func.id
        return None

    calls = [name for statement in render_fn.body if (name := _call_name(statement))]
    order = (
        "_render_cycle_frontier_section",
        "_render_contracted_floor_section",
        "_render_forecast_policy_section",
    )
    indices = [calls.index(name) for name in order]
    assert indices[1] == indices[0] + 1
    assert indices[2] == indices[1] + 1


def _gated_floor_app() -> None:
    from src.pages.simulation_cockpit import _render_contracted_floor_section

    _render_contracted_floor_section(
        frontier_context=None, chart_template="plotly_dark",
    )


def _floor_app() -> None:
    import pandas as pd
    import streamlit as st

    from src.pages.simulation_cockpit import (
        _render_contracted_floor_section,
        _render_strategy_comparison,
    )

    merchant = st.number_input(
        "Harness frontier merchant", value=40000.0,
        key="harness_frontier_merchant",
    )
    gross = st.number_input(
        "Harness frontier gross", value=50000.0,
        key="harness_frontier_gross",
    )
    comparison = pd.DataFrame({
        "strategy": ["DA-only", "DA + IDA1 (forecast-driven)"],
        "window_revenue_eur": [100.0, 120.0],
        "annualized_eur_per_mw": [36525.0, 43830.0],
        "uplift_vs_da_pct": [0.0, 20.0],
    })
    _render_strategy_comparison(comparison, "plotly_dark")
    frontier = pd.DataFrame({
        "cycle_cap": [1.2],
        "label": ["1.2 EFC/day"],
        "gross_eur": [3500.0],
        "avg_efc_per_day": [1.2],
        "wear_eur": [500.0],
        "net_eur": [3000.0],
        "gross_eur_per_mw_yr": [float(gross)],
        "net_eur_per_mw_yr": [float(merchant)],
    })
    context = {
        "fingerprint": ("frontier-v1", "DE_LU", 30),
        "frontier": frontier,
        "summary": {
            "best_cap_label": "1.2 EFC/day",
            "cycle_life": 6000.0,
            "capex_eur_kwh": 150.0,
            "valid_days": 30,
            "frontier_flat": False,
            "n_tiebreak_fallback_days": 0,
        },
        "sweep_dates": ("2026-03-01", "2026-03-30"),
        "primary_zone": "DE_LU",
        "power_mw": 1.0,
        "duration_hours": 2.0,
        "cycle_life": 6000.0,
        "capex_eur_kwh": 150.0,
    }
    assumptions = pd.DataFrame([{
        "parameter": "Power", "value": "1", "unit": "MW",
        "source": "Sidebar", "affects": "Everything",
    }])
    _render_contracted_floor_section(
        frontier_context=context,
        chart_template="plotly_dark",
        assumptions=assumptions,
    )


def _find_elements(node, type_name: str) -> list:
    found: list = []
    children = getattr(node, "children", None)
    if isinstance(children, dict):
        for child in children.values():
            if getattr(child, "type", "") == type_name:
                found.append(child)
            found.extend(_find_elements(child, type_name))
    return found


class TestContractedFloorAppTest:
    def test_frontier_result_gates_panel(self) -> None:
        app = AppTest.from_function(_gated_floor_app).run(timeout=30)
        assert not app.exception
        assert app.expander[0].label == _EXPANDER_TITLE
        assert any("Run a valid cycle-cap frontier" in info.value for info in app.info)
        assert len(app.number_input) == 0

    @pytest.fixture()
    def app(self) -> AppTest:
        return AppTest.from_function(_floor_app).run(timeout=30)

    def _run_floor(self, app: AppTest) -> AppTest:
        app.number_input(key="contracted_floor_availability_pct").set_value(
            80.0
        ).run(timeout=30)
        return app.button(key="contracted_floor_run").click().run(timeout=30)

    def _run_composition(
        self, app: AppTest, *, decay_pct: float = 50.0,
        minimum_share_pct: float = 0.0, escalation_pct: float = 0.0,
    ) -> AppTest:
        app.number_input(key="contracted_floor_availability_pct").set_value(
            80.0
        ).run(timeout=30)
        app.number_input(key="contracted_floor_decay_pct").set_value(
            decay_pct
        ).run(timeout=30)
        app.number_input(key="contracted_floor_minimum_share_pct").set_value(
            minimum_share_pct
        ).run(timeout=30)
        app.number_input(key="contracted_floor_escalation_pct").set_value(
            escalation_pct
        ).run(timeout=30)
        return app.button(key="contracted_floor_run").click().run(timeout=30)

    def test_run_renders_outputs_and_preserves_percent_boundaries(
        self, app: AppTest,
    ) -> None:
        self._run_floor(app)
        assert not app.exception
        metrics = {metric.label: metric.value for metric in app.metric}
        assert metrics["Annual merchant net"] == "EUR 40,000"
        assert metrics["Effective contracted floor"] == "EUR 40,000"
        assert metrics["Floor-protected annual cash flow"] == "EUR 40,000"
        assert metrics["Annual top-up"] == "EUR 0"

        captions = [caption.value for caption in app.caption]
        assert any(_CONTRACTED_FLOOR_SOURCE_LABEL in text for text in captions)
        assert _CONTRACTED_FLOOR_HARD_CAPTION in captions
        assert _CONTRACTED_FLOOR_TRAJECTORY_CAPTION in captions
        assert _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION not in captions
        assert _CONTRACTED_FLOOR_NEGATIVE_YEAR_CAPTION not in captions

        charts = _find_elements(app.main, "plotly_chart")
        assert len(charts) == 2
        specs = [json.loads(chart.proto.spec) for chart in charts]
        by_title = {spec["layout"]["title"]["text"]: spec for spec in specs}
        strategy = by_title["Annualised revenue by strategy"]
        assert strategy["data"][0]["x"] == [
            "DA-only", "DA + IDA1 (forecast-driven)",
        ]
        assert all("floor" not in str(label).lower() for label in strategy["data"][0]["x"])
        floor = by_title["Merchant versus contracted floor"]
        assert [trace["name"] for trace in floor["data"]] == [
            "Merchant net", "Effective floor", "Floor protected",
        ]

        downloads = _find_elements(app.main, "download_button")
        assert len(downloads) == 1
        assert "contracted-floor" in downloads[0].proto.label

        assert (
            app.number_input(key="contracted_floor_minimum_share_pct").label
            == "Minimum merchant share (% of year-1 merchant)"
        )
        assert "minimum may not be reached" in _CONTRACTED_FLOOR_MINIMUM_SHARE_HELP
        assert app.number_input(key="contracted_floor_tenor").max == (
            MAX_FLOOR_TRAJECTORY_YEARS
        )

    def test_active_decay_renders_trajectory_and_locked_disclosures(
        self, app: AppTest,
    ) -> None:
        self._run_composition(app)
        assert not app.exception
        captions = [caption.value for caption in app.caption]
        assert any("floor first binds in year 2" in text for text in captions)
        assert any("binds in 9 of 10 evaluated years" in text for text in captions)
        assert _CONTRACTED_FLOOR_NEGATIVE_YEAR_CAPTION in captions
        assert _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION in captions
        charts = _find_elements(app.main, "plotly_chart")
        specs = [json.loads(chart.proto.spec) for chart in charts]
        by_title = {spec["layout"]["title"]["text"]: spec for spec in specs}
        assert set(by_title) == {
            "Annualised revenue by strategy",
            "Merchant versus contracted floor",
            "Merchant and contracted-floor trajectory",
        }
        trajectory = by_title["Merchant and contracted-floor trajectory"]
        assert [trace["name"] for trace in trajectory["data"]] == [
            "Top-up T_t", "Merchant M_t", "Floor F_t",
        ]
        metrics = {metric.label: metric.value for metric in app.metric}
        assert metrics["Annual merchant net"] == "EUR 40,000"
        assert metrics["Effective contracted floor"] == "EUR 40,000"
        assert metrics["Annual top-up"] == "EUR 0"
        assert metrics["Merchant contract-window PV"] != "EUR 268,403"

    def test_active_without_negative_year_omits_negative_disclosure(
        self, app: AppTest,
    ) -> None:
        self._run_composition(app, decay_pct=10.0, minimum_share_pct=80.0)
        assert not app.exception
        captions = [caption.value for caption in app.caption]
        assert _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION in captions
        assert _CONTRACTED_FLOOR_NEGATIVE_YEAR_CAPTION not in captions

    def test_decay_floor_at_100_percent_keeps_v1_surface(
        self, app: AppTest,
    ) -> None:
        self._run_composition(app, decay_pct=20.0, minimum_share_pct=100.0)
        assert not app.exception
        captions = [caption.value for caption in app.caption]
        assert _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION not in captions
        assert _CONTRACTED_FLOOR_NEGATIVE_YEAR_CAPTION not in captions
        assert len(_find_elements(app.main, "plotly_chart")) == 2
        metrics = {metric.label: metric.value for metric in app.metric}
        assert metrics["Merchant contract-window PV"] == "EUR 268,403"

    def test_escalation_alone_activates_trajectory(self, app: AppTest) -> None:
        self._run_composition(app, decay_pct=0.0, escalation_pct=10.0)
        assert not app.exception
        assert _CONTRACTED_FLOOR_COMPOSITION_HARD_CAPTION in [
            caption.value for caption in app.caption
        ]
        assert len(_find_elements(app.main, "plotly_chart")) == 3

    @pytest.mark.parametrize(
        "key",
        [
            "contracted_floor_decay_pct",
            "contracted_floor_minimum_share_pct",
            "contracted_floor_escalation_pct",
        ],
    )
    def test_cleared_composition_input_prompts_instead_of_crashing(
        self, key: str,
    ) -> None:
        app = AppTest.from_function(_floor_app)
        app.session_state[key] = None
        app.run(timeout=30)
        assert not app.exception
        assert any("Enter all contract" in info.value for info in app.info)
        assert len(_find_elements(app.main, "download_button")) == 0

    def test_percent_inputs_are_divided_before_calculation(
        self, monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        observed: list[tuple[float, float, float]] = []
        original = cockpit.compute_decaying_contracted_floor_overlay

        def spy(*args, **kwargs):
            observed.append((
                kwargs["annual_decay_rate"],
                kwargs["decay_floor_share"],
                kwargs["floor_escalation_rate"],
            ))
            return original(*args, **kwargs)

        monkeypatch.setattr(
            cockpit, "compute_decaying_contracted_floor_overlay", spy,
        )
        app = AppTest.from_function(_floor_app).run(timeout=30)
        self._run_composition(
            app, decay_pct=10.0, minimum_share_pct=40.0,
            escalation_pct=2.0,
        )
        assert not app.exception
        assert observed[-1] == pytest.approx((0.10, 0.40, 0.02))

    def test_divergence_hint_is_display_only_and_gated(self, app: AppTest) -> None:
        app.session_state["merchant_revenue_decay_pct"] = 12.0
        app.session_state["merchant_revenue_decay_floor_pct"] = 30.0
        app.run(timeout=30)
        assert any(
            "Risk Analysis currently asserts 12%/yr, floor 30%; this panel "
            "uses its own inputs." in info.value
            for info in app.info
        )
        self._run_floor(app)
        metrics = {metric.label: metric.value for metric in app.metric}
        assert metrics["Merchant contract-window PV"] == "EUR 268,403"

    def test_equal_or_absent_risk_values_render_no_divergence_hint(
        self, app: AppTest,
    ) -> None:
        assert not any("Risk Analysis currently" in i.value for i in app.info)
        app.session_state["merchant_revenue_decay_pct"] = 0.0
        app.session_state["merchant_revenue_decay_floor_pct"] = 0.0
        app.run(timeout=30)
        assert not any("Risk Analysis currently" in i.value for i in app.info)

    def test_results_survive_unrelated_rerun(self, app: AppTest) -> None:
        self._run_floor(app)
        app.run(timeout=30)
        assert not app.exception
        assert "Annual merchant net" in [metric.label for metric in app.metric]
        assert len(_find_elements(app.main, "download_button")) == 1

    def test_contract_input_change_invalidates_cached_result(
        self, app: AppTest,
    ) -> None:
        self._run_floor(app)
        app.number_input(key="contracted_floor_quote").set_value(
            60000.0
        ).run(timeout=30)
        assert not app.exception
        assert any("Contract or frontier inputs changed" in i.value for i in app.info)
        assert "Annual merchant net" not in [metric.label for metric in app.metric]
        assert len(_find_elements(app.main, "download_button")) == 0

    def test_recomputed_frontier_value_invalidates_cached_result(
        self, app: AppTest,
    ) -> None:
        self._run_floor(app)
        app.number_input(key="harness_frontier_merchant").set_value(
            45000.0
        ).run(timeout=30)
        assert not app.exception
        assert any("Contract or frontier inputs changed" in i.value for i in app.info)
        assert "Annual merchant net" not in [metric.label for metric in app.metric]

    @pytest.mark.parametrize(
        ("key", "value"),
        [
            ("contracted_floor_decay_pct", 10.0),
            ("contracted_floor_minimum_share_pct", 20.0),
            ("contracted_floor_escalation_pct", 3.0),
            ("harness_frontier_gross", 51000.0),
        ],
    )
    def test_composition_or_gross_change_invalidates_cached_result(
        self, app: AppTest, key: str, value: float,
    ) -> None:
        self._run_floor(app)
        app.number_input(key=key).set_value(value).run(timeout=30)
        assert not app.exception
        assert any("Contract or frontier inputs changed" in i.value for i in app.info)
        assert "Annual merchant net" not in [metric.label for metric in app.metric]
        assert len(_find_elements(app.main, "download_button")) == 0
