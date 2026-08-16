"""PC-C Excel presentation tests for typed Project Case RunResult outputs."""

from __future__ import annotations

import base64
import dataclasses as dc
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.export import export_to_bytes, project_case_to_excel
from src.project_case import (
    AugmentationEvent,
    CapacityMaintenanceBasis,
    LifecycleCase,
    RunResult,
    compute_project_case,
    fingerprint_hex,
)
from src.project_case.schema import _issue_strategy_run_result
from tests import pc_case_fixtures as fx

_UNCHANGED = object()


def _known_result():
    return compute_project_case(fx.project_case())


def _unknown_result():
    case = fx.project_case()
    lifecycle = LifecycleCase(
        project_life_years=case.lifecycle_case.project_life_years,
        capacity_maintenance_basis=CapacityMaintenanceBasis.UNKNOWN,
        capacity_maintenance_source=None,
        capacity_maintenance_as_of=None,
        augmentation_events=(),
        eol_residual_value_eur=case.lifecycle_case.eol_residual_value_eur,
        decommissioning_cost_eur=case.lifecycle_case.decommissioning_cost_eur,
    )
    return compute_project_case(dc.replace(case, lifecycle_case=lifecycle))


def _contract_result():
    return compute_project_case(
        fx.project_case(
            contract=fx.contract_case(rates=(10_000.0, 20_000.0))
        )
    )


def _rows_by_label(ws) -> dict[str, tuple[object, object]]:
    return {
        ws.cell(row=row, column=1).value: (
            ws.cell(row=row, column=2).value,
            ws.cell(row=row, column=3).value,
        )
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=1).value is not None
    }


def _audit_values(ws) -> dict[str, object]:
    return {
        ws.cell(row=row, column=1).value: ws.cell(row=row, column=2).value
        for row in range(2, ws.max_row + 1)
    }


def _result_with_provenance(
    result: RunResult,
    provenance: dict,
    *,
    input_fingerprint: str | None = None,
    screening_outcome: object = _UNCHANGED,
    lifecycle_outcome: object = _UNCHANGED,
    screening_table: object = _UNCHANGED,
    lifecycle_table: object = _UNCHANGED,
) -> RunResult:
    """Forge an invalid envelope to mutation-test the independent export gate."""
    unsafe = object.__new__(RunResult)
    for name, value in (
        ("input_fingerprint", input_fingerprint or result.input_fingerprint),
        (
            "no_lifecycle_cost_screening_npv",
            result.no_lifecycle_cost_screening_npv
            if screening_outcome is _UNCHANGED
            else screening_outcome,
        ),
        (
            "lifecycle_cash_npv",
            result.lifecycle_cash_npv
            if lifecycle_outcome is _UNCHANGED
            else lifecycle_outcome,
        ),
        ("provenance", provenance),
        ("schema_version", result.schema_version),
        (
            "screening_cashflow_table",
            result.screening_cashflow_table
            if screening_table is _UNCHANGED
            else screening_table,
        ),
        (
            "lifecycle_cashflow_table",
            result.lifecycle_cashflow_table
            if lifecycle_table is _UNCHANGED
            else lifecycle_table,
        ),
    ):
        object.__setattr__(unsafe, name, value)
    return unsafe


def _unsafe_replace(value, **changes):
    """Bypass a frozen dataclass constructor for an adversarial export probe."""
    unsafe = object.__new__(type(value))
    for field in dc.fields(value):
        object.__setattr__(
            unsafe,
            field.name,
            changes.get(field.name, getattr(value, field.name)),
        )
    return unsafe


def _unsafe_table_with_row(table, row_index: int, **changes):
    rows = list(table.rows)
    rows[row_index] = _unsafe_replace(rows[row_index], **changes)
    return _unsafe_replace(table, rows=tuple(rows))


def _reissue_strategy_run_result(result, **changes):
    """Build one valid producer fixture variant through the guarded issuer."""
    fields = {
        field.name: changes.get(field.name, getattr(result, field.name))
        for field in dc.fields(result)
    }
    with _issue_strategy_run_result():
        return type(result)(**fields)


def test_project_case_to_excel_writes_stable_full_workbook() -> None:
    result = _known_result()
    workbook = load_workbook(BytesIO(project_case_to_excel(result)))

    assert workbook.sheetnames == [
        "Project Case NPVs",
        "Screening Cash Flow",
        "Lifecycle Cash Flow",
        "Assumptions & Provenance",
    ]

    npv_rows = _rows_by_label(workbook["Project Case NPVs"])
    assert npv_rows["Available"] == (True, True)
    assert npv_rows["Status"] == ("ok", "ok")
    assert npv_rows["P10 (Downside, EUR)"] == pytest.approx(
        (
            result.no_lifecycle_cost_screening_npv.distribution.p10,
            result.lifecycle_cash_npv.distribution.p10,
        )
    )
    assert npv_rows["P50 (Median, EUR)"] == pytest.approx(
        (
            result.no_lifecycle_cost_screening_npv.distribution.p50,
            result.lifecycle_cash_npv.distribution.p50,
        )
    )
    assert npv_rows["P90 (Upside, EUR)"] == pytest.approx(
        (
            result.no_lifecycle_cost_screening_npv.distribution.p90,
            result.lifecycle_cash_npv.distribution.p90,
        )
    )
    assert npv_rows["Strategy Kind"] == ("DA_ONLY", None)
    assert npv_rows["Producer Adapter"] == ("PC_ADP_DA_ONLY", None)
    assert npv_rows["Observed Dates"] == (2, None)
    assert npv_rows["Valid Dates"] == (2, None)
    assert npv_rows["Missing Dates"] == (0, None)
    assert npv_rows["Solver-failed Dates"] == (0, None)

    expected_headers = [
        "Project Year",
        "Merchant Strategy Cash (EUR)",
        "Effective Contract Floor (EUR)",
        "Contract Top-Up (EUR)",
        "Settled Strategy Cash (EUR)",
        "Fixed O&M (EUR)",
        "Augmentation Net Cost (EUR)",
        "Terminal Cash (EUR)",
        "Net Cash Flow (EUR)",
        "Discount Factor",
        "Discounted Net Cash Flow (EUR)",
    ]
    for sheet_name, table in (
        ("Screening Cash Flow", result.screening_cashflow_table),
        ("Lifecycle Cash Flow", result.lifecycle_cashflow_table),
    ):
        sheet = workbook[sheet_name]
        assert [cell.value for cell in sheet[1]][:2] == [
            "Representative P50 reconciliation",
            "Value (EUR)",
        ]
        reconciled = {
            sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value
            for row in range(2, 7)
        }
        assert reconciled["Installed CapEx at year 0"] == 5_000_000
        expected_p50 = (
            result.no_lifecycle_cost_screening_npv.distribution.p50
            if sheet_name == "Screening Cash Flow"
            else result.lifecycle_cash_npv.distribution.p50
        )
        assert reconciled["Reconciled P50 NPV"] == pytest.approx(expected_p50)
        assert reconciled["Reported P50 NPV"] == pytest.approx(expected_p50)
        assert reconciled["Reconciliation difference"] == pytest.approx(
            0.0, abs=1e-6
        )
        assert [
            sheet.cell(row=8, column=col).value
            for col in range(1, len(expected_headers) + 1)
        ] == expected_headers
        assert sheet.max_row == len(table.rows) + 8
        first = table.rows[0]
        exported_first = [
            sheet.cell(row=9, column=col).value
            for col in range(1, len(expected_headers) + 1)
        ]
        assert exported_first[2] is None
        assert [*exported_first[:2], *exported_first[3:]] == pytest.approx(
            [
                first.year,
                first.merchant_revenue_eur,
                first.contract_top_up_eur,
                first.revenue_eur,
                first.opex_eur,
                first.augmentation_eur,
                first.terminal_eur,
                first.net_eur,
                first.discount_factor,
                first.discounted_net_eur,
            ]
        )

    audit = _audit_values(workbook["Assumptions & Provenance"])
    assert audit["run_result.input_fingerprint"] == result.input_fingerprint
    assert (
        audit[
            "provenance.red_line_assertions.wear_net_floor_comparator_included"
        ]
        == "false"
    )
    assert (
        audit["provenance.red_line_assertions.contract_settlement_included"]
        == "false"
    )
    assert audit["provenance.red_line_assertions.pre_tax_unlevered"] == "true"
    assert audit["provenance.project_case.asset_case.installed_capex_eur"] == "5000000.0"
    assert audit["provenance.project_case.lifecycle_case.augmentation_events[0].year"] == "8"


def test_contract_case_export_discloses_settlement_basis_and_full_cash_waterfall() -> None:
    result = _contract_result()
    workbook = load_workbook(BytesIO(project_case_to_excel(result)))

    assert workbook.sheetnames == [
        "Project Case NPVs",
        "Screening Cash Flow",
        "Lifecycle Cash Flow",
        "Contract Settlement",
        "Assumptions & Provenance",
    ]
    npv_rows = _rows_by_label(workbook["Project Case NPVs"])
    assert npv_rows["Typed Contract Settlement Included"] == (True, None)
    assert npv_rows["Contract Settlement Basis"] == (
        "ANNUAL_PRE_LIFECYCLE_STRATEGY_CASH_FLOOR_V1",
        None,
    )
    assert npv_rows["Wear-net Floor Comparator Included"] == (False, None)
    assert "not MACSE" in npv_rows["Contract Product Boundary"][0]

    contract_sheet = workbook["Contract Settlement"]
    contract_rows = _rows_by_label(contract_sheet)
    assert contract_rows["Modelled Whole-project Power (MW)"] == (10, None)
    assert contract_rows["Asset Scope"] == ("WHOLE_PROJECT_MODELED_MW", None)
    assert contract_rows["Contract Start Project Year"] == (2, None)
    assert contract_rows["Contract Tenor (years)"] == (2, None)
    assert contract_rows["Real-EUR Base Year"] == (2026, None)
    assert contract_rows["Source"] == ("PC-D test scenario", None)
    assert "not a complete legal-contract model" in contract_rows["Product Boundary"][0]

    header_row = next(
        row
        for row in range(1, contract_sheet.max_row + 1)
        if contract_sheet.cell(row=row, column=1).value == "Contract Year"
    )
    assert [
        contract_sheet.cell(row=header_row, column=column).value
        for column in range(1, 6)
    ] == [
        "Contract Year",
        "Project Year",
        "Real Floor Rate (EUR/modelled MW-year)",
        "Floor Entitlement Factor",
        "Effective Whole-project Floor (EUR)",
    ]
    assert [
        contract_sheet.cell(row=header_row + 1, column=column).value
        for column in range(1, 6)
    ] == pytest.approx([1, 2, 10_000, 0.5, 50_000])
    assert [
        contract_sheet.cell(row=header_row + 2, column=column).value
        for column in range(1, 6)
    ] == pytest.approx([2, 3, 20_000, 1.0, 200_000])

    cash_sheet = workbook["Screening Cash Flow"]
    assert [cash_sheet.cell(row=8, column=column).value for column in range(1, 6)] == [
        "Project Year",
        "Merchant Strategy Cash (EUR)",
        "Effective Contract Floor (EUR)",
        "Contract Top-Up (EUR)",
        "Settled Strategy Cash (EUR)",
    ]
    row_year_2 = 10
    assert cash_sheet.cell(row=row_year_2, column=1).value == 2
    assert cash_sheet.cell(row=row_year_2, column=3).value == pytest.approx(50_000)
    assert cash_sheet.cell(row=row_year_2, column=5).value == pytest.approx(
        max(
            cash_sheet.cell(row=row_year_2, column=2).value,
            cash_sheet.cell(row=row_year_2, column=3).value,
        )
    )

    audit = _audit_values(workbook["Assumptions & Provenance"])
    assert audit["provenance.red_line_assertions.contract_settlement_included"] == "true"
    assert (
        audit["provenance.red_line_assertions.wear_net_floor_comparator_included"]
        == "false"
    )
    assert audit["provenance.contract_settlement.resolved_floor_by_project_year[0].year"] == "2"


def test_unknown_lifecycle_is_blank_and_has_no_cashflow_sheet() -> None:
    result = _unknown_result()
    workbook = load_workbook(BytesIO(project_case_to_excel(result)))

    assert workbook.sheetnames == [
        "Project Case NPVs",
        "Screening Cash Flow",
        "Assumptions & Provenance",
    ]
    rows = _rows_by_label(workbook["Project Case NPVs"])
    assert rows["Available"] == (True, False)
    assert rows["Status"][1] == "capacity_maintenance_unknown"
    assert rows["Message"][1] == "Engineering capacity-maintenance basis is unknown."
    for label in (
        "P10 (Downside, EUR)",
        "P50 (Median, EUR)",
        "P90 (Upside, EUR)",
        "P(NPV > 0)",
    ):
        assert rows[label][0] is not None
        assert rows[label][1] is None


def test_project_case_export_is_formula_safe_and_preserves_large_seed_as_text() -> None:
    case = fx.project_case()
    lifecycle = dc.replace(
        case.lifecycle_case,
        capacity_maintenance_source="=HYPERLINK(\"https://invalid\",\"click\")",
    )
    bootstrap = dc.replace(case.bootstrap_case, seed=2**64 - 1)
    result = compute_project_case(
        dc.replace(case, lifecycle_case=lifecycle, bootstrap_case=bootstrap)
    )
    sheet = load_workbook(BytesIO(project_case_to_excel(result)))[
        "Assumptions & Provenance"
    ]
    audit = _audit_values(sheet)
    source_path = "provenance.project_case.lifecycle_case.capacity_maintenance_source"
    assert audit[source_path] == '=HYPERLINK("https://invalid","click")'
    assert audit["provenance.bootstrap.seed"] == str(2**64 - 1)
    for row in range(2, sheet.max_row + 1):
        assert sheet.cell(row=row, column=2).data_type != "f"
        assert sheet.cell(row=row, column=2).quotePrefix is True


def test_project_case_export_rejects_wear_net_comparator_or_wrong_type() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    provenance["red_line_assertions"]["wear_net_floor_comparator_included"] = True
    unsafe = _result_with_provenance(result, provenance)
    with pytest.raises(ValueError, match="wear_net_floor_comparator_included"):
        project_case_to_excel(unsafe)
    with pytest.raises(TypeError, match="RunResult"):
        project_case_to_excel(object())


def test_project_case_export_rejects_mismatched_audit_fingerprint() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    provenance["project_case_input_fingerprint"] = "00" * 32
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="fingerprint does not match"):
        project_case_to_excel(unsafe)


def test_project_case_export_recomputes_strategy_fingerprint_from_payload() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    provenance["strategy_run_result"]["daily_realised_cash_series"][0][1] += 99.0
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="StrategyRunResult fingerprint"):
        project_case_to_excel(unsafe)


def test_project_case_export_binds_project_case_to_strategy_payload() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    strategy = provenance["strategy_run_result"]
    strategy["daily_realised_cash_series"][0][1] += 99.0
    provenance["strategy_run_fingerprint"] = fingerprint_hex(
        "StrategyRunResult", strategy
    )
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="not the canonical typed"):
        project_case_to_excel(unsafe)


def test_project_case_export_recomputes_project_case_fingerprint_from_payload() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    provenance["project_case"]["asset_case"]["installed_capex_eur"] += 1.0
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="ProjectCase fingerprint"):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_contract_payload_alias_after_valid_fingerprint() -> None:
    result = _contract_result()
    provenance = result.to_payload()["provenance"]
    provenance["project_case"]["contract_case"]["settlement_terms"][
        "floor_protected_cashflow_eur"
    ] = 123.0
    fingerprint = fingerprint_hex("ProjectCase", provenance["project_case"])
    provenance["project_case_input_fingerprint"] = fingerprint
    unsafe = _result_with_provenance(
        result,
        provenance,
        input_fingerprint=fingerprint,
    )

    with pytest.raises(ValueError, match=r"settlement_terms keys invalid"):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_noncanonical_contract_as_of_date() -> None:
    result = _contract_result()
    provenance = result.to_payload()["provenance"]
    provenance["project_case"]["contract_case"]["settlement_terms"][
        "source_as_of_date"
    ] = "2026-02-30"
    fingerprint = fingerprint_hex("ProjectCase", provenance["project_case"])
    provenance["project_case_input_fingerprint"] = fingerprint
    unsafe = _result_with_provenance(
        result,
        provenance,
        input_fingerprint=fingerprint,
    )

    with pytest.raises(ValueError, match="source_as_of_date is not a valid date"):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_integer_alias_for_fingerprinted_real() -> None:
    result = _contract_result()
    provenance = result.to_payload()["provenance"]
    terms = provenance["project_case"]["contract_case"]["settlement_terms"]
    terms["floor_rate_real_eur_per_modeled_mw_year_by_contract_year"][0] = 10_000
    fingerprint = fingerprint_hex("ProjectCase", provenance["project_case"])
    provenance["project_case_input_fingerprint"] = fingerprint
    unsafe = _result_with_provenance(
        result,
        provenance,
        input_fingerprint=fingerprint,
    )

    with pytest.raises(ValueError, match=r"floor rate.*canonical float"):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_integer_alias_in_strategy_wire() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    strategy = provenance["strategy_run_result"]
    strategy["power_mw"] = 10
    strategy_fingerprint = fingerprint_hex("StrategyRunResult", strategy)
    provenance["strategy_run_fingerprint"] = strategy_fingerprint
    provenance["project_case"]["market_case"][
        "strategy_run_fingerprint"
    ] = strategy_fingerprint
    project_fingerprint = fingerprint_hex("ProjectCase", provenance["project_case"])
    provenance["project_case_input_fingerprint"] = project_fingerprint
    unsafe = _result_with_provenance(
        result,
        provenance,
        input_fingerprint=project_fingerprint,
    )

    with pytest.raises(ValueError, match=r"power_mw.*canonical float"):
        project_case_to_excel(unsafe)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("power", r"power_mw != AssetCase.power_mw"),
        ("currency", "target_base_year must equal ValuationCase.base_year"),
    ],
)
def test_project_case_export_binds_strategy_engineering_and_currency(
    mutation: str,
    message: str,
) -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    strategy = provenance["strategy_run_result"]
    if mutation == "power":
        strategy["power_mw"] = 11.0
    else:
        strategy["currency_basis"]["target_base_year"] = 2027
    strategy_fingerprint = fingerprint_hex("StrategyRunResult", strategy)
    provenance["strategy_run_fingerprint"] = strategy_fingerprint
    provenance["project_case"]["market_case"][
        "strategy_run_fingerprint"
    ] = strategy_fingerprint
    project_fingerprint = fingerprint_hex("ProjectCase", provenance["project_case"])
    provenance["project_case_input_fingerprint"] = project_fingerprint
    unsafe = _result_with_provenance(
        result,
        provenance,
        input_fingerprint=project_fingerprint,
    )

    with pytest.raises(ValueError, match=message):
        project_case_to_excel(unsafe)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("lower_sorted_rank", 0, "ranks do not match linear P50"),
        ("interpolation_weight", 0.25, "weight does not match linear P50"),
    ],
)
def test_project_case_export_rejects_noncanonical_p50_interpolation(
    field: str,
    value: int | float,
    message: str,
) -> None:
    result = _contract_result()
    provenance = result.to_payload()["provenance"]
    provenance["contract_settlement"]["representative_interpolation"][field] = value
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match=message):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_wrong_resolved_floor_even_with_valid_contract() -> None:
    result = _contract_result()
    provenance = result.to_payload()["provenance"]
    provenance["contract_settlement"]["resolved_floor_by_project_year"][0][
        "effective_floor_eur"
    ] += 1.0
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="rate x MW x entitlement"):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_comparator_smuggling_under_alias() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    # The old semantic denylist only searched for the token "floor" and let this
    # equally dangerous contract-cash alias through.
    provenance["external_contract_comparator"] = {
        "protected_cashflow_eur": 123.0
    }
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="exact current PC-D2 schema"):
        project_case_to_excel(unsafe)


def test_project_case_export_rejects_extra_red_line_assertion() -> None:
    result = _known_result()
    provenance = result.to_payload()["provenance"]
    provenance["red_line_assertions"]["contract_comparator_included"] = False
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match=r"red_line_assertions.*exact current PC-D2"):
        project_case_to_excel(unsafe)


def test_project_case_export_accepts_locked_floor_fields_only_in_their_schema_slots() -> None:
    # The fixture's ProjectCase projection carries the legitimate
    # decay_floor_share slot, and red lines exclude the wear-net comparator.
    project_case_to_excel(_known_result())


def test_assumption_export_preserves_full_float_precision() -> None:
    case = fx.project_case()
    valuation = dc.replace(case.valuation_case, discount_rate=0.081234567890123)
    result = compute_project_case(dc.replace(case, valuation_case=valuation))
    sheet = load_workbook(BytesIO(project_case_to_excel(result)))[
        "Assumptions & Provenance"
    ]
    audit = _audit_values(sheet)
    assert audit["provenance.project_case.valuation_case.discount_rate"] == (
        "0.081234567890123"
    )


def test_assumption_export_chunks_long_raw_text_losslessly_and_formula_safely() -> None:
    case = fx.project_case()
    raw_source = "=" + "x" * 70_000
    lifecycle = dc.replace(case.lifecycle_case, capacity_maintenance_source=raw_source)
    result = compute_project_case(dc.replace(case, lifecycle_case=lifecycle))
    sheet = load_workbook(BytesIO(project_case_to_excel(result)))[
        "Assumptions & Provenance"
    ]
    path = "provenance.project_case.lifecycle_case.capacity_maintenance_source"
    chunks = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value != path:
            continue
        chunks.append(
            (
                sheet.cell(row=row, column=4).value,
                sheet.cell(row=row, column=5).value,
                sheet.cell(row=row, column=2),
            )
        )

    assert [index for index, _count, _cell in chunks] == [1, 2, 3]
    assert {count for _index, count, _cell in chunks} == {3}
    assert "".join(cell.value for _index, _count, cell in chunks) == raw_source
    assert all(len(cell.value) <= 32_767 for _index, _count, cell in chunks)
    assert all(cell.data_type == "s" for _index, _count, cell in chunks)
    assert all(cell.quotePrefix is True for _index, _count, cell in chunks)


def test_assumption_export_reversibly_encodes_xml_illegal_control_characters() -> None:
    case = fx.project_case()
    raw_source = "\x00" + "x" * 70_000 + "\x01"
    lifecycle = dc.replace(
        case.lifecycle_case,
        capacity_maintenance_source=raw_source,
    )
    result = compute_project_case(dc.replace(case, lifecycle_case=lifecycle))
    sheet = load_workbook(BytesIO(project_case_to_excel(result)))[
        "Assumptions & Provenance"
    ]
    path = "provenance.project_case.lifecycle_case.capacity_maintenance_source"
    rows = [
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value == path
    ]

    assert [sheet.cell(row=row, column=4).value for row in rows] == [1, 2, 3]
    assert {sheet.cell(row=row, column=5).value for row in rows} == {3}
    assert {sheet.cell(row=row, column=6).value for row in rows} == {
        "base64-utf8-surrogatepass"
    }
    assert {sheet.cell(row=row, column=7).value for row in rows} == {len(raw_source)}
    expected_bytes = raw_source.encode("utf-8", errors="surrogatepass")
    assert {sheet.cell(row=row, column=8).value for row in rows} == {
        len(expected_bytes)
    }
    encoded = "".join(sheet.cell(row=row, column=2).value for row in rows)
    reconstructed = base64.b64decode(encoded).decode(
        "utf-8", errors="surrogatepass"
    )
    assert reconstructed == raw_source
    assert all(len(sheet.cell(row=row, column=2).value) <= 32_767 for row in rows)


def test_export_replays_contract_floor_instead_of_trusting_table_row() -> None:
    result = _contract_result()
    table = result.screening_cashflow_table
    assert table is not None
    row = table.rows[1]
    assert row.effective_contract_floor_eur is not None
    forged = _unsafe_table_with_row(
        table,
        1,
        effective_contract_floor_eur=row.effective_contract_floor_eur + 1.0,
    )
    unsafe = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_table=forged,
    )

    with pytest.raises(ValueError, match=r"effective floor.*does not match fingerprinted"):
        project_case_to_excel(unsafe)


def test_export_replays_merchant_and_top_up_even_when_settled_p50_is_unchanged() -> None:
    result = _contract_result()
    table = result.screening_cashflow_table
    assert table is not None
    row = table.rows[1]
    forged = _unsafe_table_with_row(
        table,
        1,
        merchant_revenue_eur=row.merchant_revenue_eur - 999.0,
        contract_top_up_eur=row.contract_top_up_eur + 999.0,
    )
    unsafe = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_table=forged,
    )

    with pytest.raises(ValueError, match=r"merchant cash.*does not match fingerprinted"):
        project_case_to_excel(unsafe)


def test_export_replays_null_contract_decomposition_and_year_universe() -> None:
    result = _known_result()
    table = result.screening_cashflow_table
    assert table is not None
    row = table.rows[0]
    forged_row = _unsafe_table_with_row(
        table,
        0,
        merchant_revenue_eur=row.merchant_revenue_eur - 1.0,
        contract_top_up_eur=1.0,
    )
    unsafe_row = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_table=forged_row,
    )
    with pytest.raises(ValueError, match=r"merchant cash.*does not match fingerprinted"):
        project_case_to_excel(unsafe_row)

    shortened = _unsafe_replace(table, rows=table.rows[1:])
    unsafe_years = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_table=shortened,
    )
    with pytest.raises(ValueError, match="cover exactly years"):
        project_case_to_excel(unsafe_years)


def test_export_replays_discount_and_lifecycle_cash_components() -> None:
    result = _known_result()
    screening = result.screening_cashflow_table
    lifecycle = result.lifecycle_cashflow_table
    assert screening is not None and lifecycle is not None
    screening_row = screening.rows[0]
    new_discount = screening_row.discount_factor / 2.0
    forged_screening = _unsafe_table_with_row(
        screening,
        0,
        discount_factor=new_discount,
        discounted_net_eur=screening_row.net_eur * new_discount,
    )
    unsafe_discount = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_table=forged_screening,
    )
    with pytest.raises(ValueError, match=r"discount factor.*does not match fingerprinted"):
        project_case_to_excel(unsafe_discount)

    lifecycle_row = lifecycle.rows[0]
    forged_lifecycle = _unsafe_table_with_row(
        lifecycle,
        0,
        opex_eur=lifecycle_row.opex_eur + 100.0,
        terminal_eur=lifecycle_row.terminal_eur + 100.0,
    )
    unsafe_lifecycle = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        lifecycle_table=forged_lifecycle,
    )
    with pytest.raises(ValueError, match=r"fixed O&M.*does not match fingerprinted"):
        project_case_to_excel(unsafe_lifecycle)


def test_export_unknown_basis_skips_overflowing_lifecycle_combinations() -> None:
    case = fx.project_case()
    strategy = _reissue_strategy_run_result(
        case.market_case.strategy_run_result,
        power_mw=1e308,
    )
    asset = dc.replace(
        case.asset_case,
        power_mw=1e308,
        fixed_om_eur_per_mw_yr=1e308,
    )
    lifecycle = LifecycleCase(
        project_life_years=case.lifecycle_case.project_life_years,
        capacity_maintenance_basis=CapacityMaintenanceBasis.UNKNOWN,
        capacity_maintenance_source=None,
        capacity_maintenance_as_of=None,
        augmentation_events=case.lifecycle_case.augmentation_events,
        eol_residual_value_eur=1e308,
        decommissioning_cost_eur=1e308,
    )
    result = compute_project_case(
        dc.replace(
            case,
            asset_case=asset,
            lifecycle_case=lifecycle,
            market_case=dc.replace(
                case.market_case,
                strategy_run_result=strategy,
            ),
        )
    )

    workbook = load_workbook(BytesIO(project_case_to_excel(result)))
    assert "Lifecycle Cash Flow" not in workbook.sheetnames
    assert workbook["Project Case NPVs"].cell(row=3, column=3).value == (
        "capacity_maintenance_unknown"
    )


def test_export_replay_uses_fsum_for_same_year_augmentation_events() -> None:
    case = fx.project_case()
    lifecycle = dc.replace(
        case.lifecycle_case,
        augmentation_events=(
            AugmentationEvent(2, 1e16, 0.1, 0.0),
            AugmentationEvent(2, 1.0, 0.0, 0.0),
            AugmentationEvent(2, 0.0, 0.0, 1e16),
        ),
    )
    result = compute_project_case(dc.replace(case, lifecycle_case=lifecycle))
    assert result.lifecycle_cashflow_table is not None
    assert result.lifecycle_cashflow_table.rows[1].augmentation_eur == 1.0

    project_case_to_excel(result)


def test_export_binds_active_maintenance_basis_to_available_lifecycle_state() -> None:
    active = _known_result()
    unknown = _unknown_result()
    unsafe = _result_with_provenance(
        active,
        active.to_payload()["provenance"],
        lifecycle_outcome=unknown.lifecycle_cash_npv,
        lifecycle_table=None,
    )

    with pytest.raises(ValueError, match=r"lifecycle NPV outcome.*available state"):
        project_case_to_excel(unsafe)


def test_export_binds_unknown_maintenance_basis_to_unavailable_lifecycle_state() -> None:
    unknown = _unknown_result()
    active = _known_result()
    unsafe = _result_with_provenance(
        unknown,
        unknown.to_payload()["provenance"],
        lifecycle_outcome=active.lifecycle_cash_npv,
        lifecycle_table=active.lifecycle_cashflow_table,
    )

    with pytest.raises(ValueError, match="UNKNOWN capacity-maintenance basis"):
        project_case_to_excel(unsafe)


@pytest.mark.parametrize(
    ("field", "delta"),
    [
        ("p10", 1.0),
        ("p90", 1.0),
        ("prob_positive", 0.1),
    ],
)
def test_export_replays_every_reported_npv_distribution_statistic(
    field: str,
    delta: float,
) -> None:
    result = _known_result()
    distribution = result.no_lifecycle_cost_screening_npv.distribution
    assert distribution is not None
    forged_distribution = dc.replace(
        distribution,
        **{field: getattr(distribution, field) + delta},
    )
    forged_outcome = dc.replace(
        result.no_lifecycle_cost_screening_npv,
        distribution=forged_distribution,
    )
    unsafe = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_outcome=forged_outcome,
    )

    with pytest.raises(ValueError, match=rf"screening NPV {field}.*does not match"):
        project_case_to_excel(unsafe)


def test_export_rejects_large_cash_row_change_hidden_by_aggregate_tolerance() -> None:
    case = fx.project_case()
    original = case.market_case.strategy_run_result
    strategy = _reissue_strategy_run_result(
        original,
        daily_realised_cash_series=tuple(
            (date, 60_000_000_000.0)
            for date, _cash in original.daily_realised_cash_series
        ),
    )
    result = compute_project_case(
        dc.replace(
            case,
            market_case=dc.replace(
                case.market_case,
                strategy_run_result=strategy,
            ),
        )
    )
    table = result.screening_cashflow_table
    assert table is not None
    row = table.rows[0]
    forged = _unsafe_table_with_row(
        table,
        0,
        merchant_revenue_eur=row.merchant_revenue_eur + 100.0,
        revenue_eur=row.revenue_eur + 100.0,
        net_eur=row.net_eur + 100.0,
        discounted_net_eur=(row.net_eur + 100.0) * row.discount_factor,
    )
    unsafe = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_table=forged,
    )

    with pytest.raises(ValueError, match=r"merchant cash.*does not match fingerprinted"):
        project_case_to_excel(unsafe)


def test_export_rejects_large_resolved_floor_change_hidden_by_relative_tolerance() -> None:
    result = compute_project_case(
        fx.project_case(
            contract=fx.contract_case(
                rates=(1e15,),
                entitlement_factors=(1.0,),
            )
        )
    )
    provenance = result.to_payload()["provenance"]
    provenance["contract_settlement"]["resolved_floor_by_project_year"][0][
        "effective_floor_eur"
    ] += 100.0
    unsafe = _result_with_provenance(result, provenance)

    with pytest.raises(ValueError, match="Resolved contract floor does not match"):
        project_case_to_excel(unsafe)


def test_export_rejects_unknown_lifecycle_with_bogus_source_and_as_of() -> None:
    result = _unknown_result()
    provenance = result.to_payload()["provenance"]
    lifecycle = provenance["project_case"]["lifecycle_case"]
    lifecycle["capacity_maintenance_source"] = "bogus"
    lifecycle["capacity_maintenance_as_of"] = "2026-08-16"
    fingerprint = fingerprint_hex("ProjectCase", provenance["project_case"])
    provenance["project_case_input_fingerprint"] = fingerprint
    unsafe = _result_with_provenance(
        result,
        provenance,
        input_fingerprint=fingerprint,
    )

    with pytest.raises(ValueError, match="UNKNOWN maintenance basis must have null"):
        project_case_to_excel(unsafe)


def test_cashflow_export_fails_closed_when_reported_p50_does_not_reconcile() -> None:
    result = _known_result()
    distribution = dc.replace(
        result.no_lifecycle_cost_screening_npv.distribution,
        p50=result.no_lifecycle_cost_screening_npv.distribution.p50 + 1.0,
    )
    screening = dc.replace(
        result.no_lifecycle_cost_screening_npv,
        distribution=distribution,
    )
    inconsistent = _result_with_provenance(
        result,
        result.to_payload()["provenance"],
        screening_outcome=screening,
    )

    with pytest.raises(ValueError, match=r"screening NPV p50.*does not match"):
        project_case_to_excel(inconsistent)


def test_existing_report_export_optionally_appends_project_case_sheets() -> None:
    index = pd.date_range("2026-01-01", periods=24, freq="h", tz="UTC")
    prices = pd.DataFrame({"price_eur_mwh": range(24)}, index=index)
    prices.index.name = "timestamp"
    daily = pd.DataFrame({"date": ["2026-01-01"], "spread": [23.0]})
    monthly = pd.DataFrame({"month": ["2026-01"], "spread": [23.0]})
    result = _known_result()

    data = export_to_bytes(
        "GB",
        prices,
        daily,
        monthly,
        {"p50": 23.0, "p75": 23.0, "p90": 23.0, "mean": 23.0},
        {
            "annual_revenue_eur_per_mw": 1.0,
            "cycles_per_day_assumption": 1.0,
            "capture_rate_assumption": 1.0,
        },
        {"negative_hours": 0, "negative_intervals": 0, "pct_negative": 0.0},
        tz="UTC",
        project_case_result=result,
    )
    workbook = load_workbook(BytesIO(data))
    assert workbook.sheetnames[-4:] == [
        "Project Case NPVs",
        "Screening Cash Flow",
        "Lifecycle Cash Flow",
        "Assumptions & Provenance",
    ]
