"""Tests for export module."""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.analytics import (
    calculate_daily_spreads,
    calculate_monthly_spreads,
    calculate_negative_price_hours,
    calculate_spread_percentiles,
    estimate_annual_arbitrage_revenue,
)
from src.data_ingestion import clean_prices
from src.export import (
    _render_figure_to_image,
    export_to_bytes,
    export_to_excel,
    export_to_pdf_bytes,
)


def _render_or_skip(fig) -> bytes:
    """Render a Plotly figure only in explicit Kaleido integration runs.

    Kaleido 1.x launches a local Chrome process for static image rendering. On
    some macOS desktop sessions that can produce OS-level Chrome crash dialogs,
    so the default unit suite keeps these checks opt-in.
    """
    if os.environ.get("BESS_PULSE_RUN_KALEIDO_TESTS") != "1":
        pytest.skip("Set BESS_PULSE_RUN_KALEIDO_TESTS=1 to run Kaleido/Chrome rendering tests.")
    try:
        return _render_figure_to_image(fig)
    except Exception as exc:
        pytest.skip(f"Kaleido image rendering unavailable in this environment: {exc}")


@pytest.fixture
def sample_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict, dict, dict]:
    """Build minimal analytics data for export tests."""
    idx = pd.date_range("2025-01-01", periods=48, freq="h", tz="UTC")
    price_df = pd.DataFrame(
        {"price_eur_mwh": [50.0 + (i % 24) * 2.0 for i in range(48)]},
        index=idx,
    )
    price_df.index.name = "timestamp"

    daily = calculate_daily_spreads(price_df)
    monthly = calculate_monthly_spreads(price_df)
    pctls = calculate_spread_percentiles(daily)
    rev = estimate_annual_arbitrage_revenue(daily)
    neg = calculate_negative_price_hours(price_df)

    return price_df, daily, monthly, pctls, rev, neg


class TestExportToExcel:
    def test_creates_valid_xlsx(self, tmp_path: Path, sample_data) -> None:
        price_df, daily, monthly, pctls, rev, neg = sample_data
        out = tmp_path / "test_report.xlsx"

        result_path = export_to_excel(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            output_path=out,
        )

        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

    def test_contains_expected_sheets(self, tmp_path: Path, sample_data) -> None:
        price_df, daily, monthly, pctls, rev, neg = sample_data
        out = tmp_path / "test_report.xlsx"

        export_to_excel(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            output_path=out,
        )

        wb = load_workbook(out)
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Daily Spreads" in sheet_names
        assert "Monthly Summary" in sheet_names
        assert "Hourly Prices" in sheet_names
        assert "Price Heatmap" in sheet_names

    def test_summary_has_zone(self, tmp_path: Path, sample_data) -> None:
        price_df, daily, monthly, pctls, rev, neg = sample_data
        out = tmp_path / "test_report.xlsx"

        export_to_excel(
            zone="FR",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            output_path=out,
        )

        wb = load_workbook(out)
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 20)]
        assert "FR" in values


class TestExportToBytes:
    def test_returns_bytes(self, sample_data) -> None:
        price_df, daily, monthly, pctls, rev, neg = sample_data
        result = export_to_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
        )
        assert isinstance(result, bytes)
        assert len(result) > 0
        # xlsx magic bytes: PK (zip)
        assert result[:2] == b"PK"

    def test_negative_stats_zeroed_when_all_days_excluded(self) -> None:
        """If every local day is incomplete (and therefore excluded from the
        spread / heatmap), the export summary's neg-price stats must zero
        out — not leak the caller's raw stats.
        """
        idx = pd.date_range("2025-01-01", periods=24, freq="h", tz="UTC")
        # Single day fully negative, but with a 5-hour NaN gap → long gap →
        # day stays NaN → entire complete-day subset is empty.
        prices = [-30.0] * 24
        for h in range(3, 8):
            prices[h] = float("nan")
        price_df = pd.DataFrame({"price_eur_mwh": prices}, index=idx)
        price_df.index.name = "timestamp"
        cleaned = clean_prices(price_df)
        daily = calculate_daily_spreads(cleaned)
        monthly = calculate_monthly_spreads(cleaned)
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        raw_neg = calculate_negative_price_hours(cleaned)
        # Pre-condition: caller's raw stats DO show negatives.
        assert raw_neg["negative_hours"] > 0

        out = export_to_bytes(
            zone="DE_LU",
            price_df=cleaned,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=raw_neg,
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Summary"]
        neg_in_sheet = None
        for r in range(1, 50):
            if ws.cell(row=r, column=1).value == "Negative Price Hours":
                neg_in_sheet = ws.cell(row=r, column=2).value
                break
        assert neg_in_sheet == 0

    def test_negative_stats_recomputed_for_consistency_with_heatmap(self) -> None:
        """When the caller passes raw negative_stats from the unfiltered
        price_df but the heatmap is filtered to complete days, the Excel
        summary used to display incompatible numbers. Export now recomputes
        neg-stats from the same filtered subset.
        """
        idx = pd.date_range("2025-01-01", periods=48, freq="h", tz="UTC")
        prices = [-30.0] * 24 + [40.0] * 24
        # Corrupt 5 contiguous hours of day 1 — long enough to exceed
        # MAX_SHORT_GAP_HOURS so the gap stays NaN and day 1 is excluded.
        for h in range(3, 8):
            prices[h] = float("nan")
        price_df = pd.DataFrame({"price_eur_mwh": prices}, index=idx)
        price_df.index.name = "timestamp"
        # daily / monthly / percentiles / revenue derived from valid window
        cleaned = clean_prices(price_df)
        daily = calculate_daily_spreads(cleaned)
        monthly = calculate_monthly_spreads(cleaned)
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        # Caller passes RAW (unfiltered) neg-stats: includes day-1 negatives
        raw_neg = calculate_negative_price_hours(cleaned)
        assert raw_neg["negative_hours"] > 0  # day 1 has 24 negatives

        out = export_to_bytes(
            zone="DE_LU",
            price_df=cleaned,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=raw_neg,
        )
        wb = load_workbook(BytesIO(out))
        ws = wb["Summary"]
        # Find "Negative Price Hours" row
        neg_hours_in_sheet = None
        for r in range(1, 50):
            if ws.cell(row=r, column=1).value == "Negative Price Hours":
                neg_hours_in_sheet = ws.cell(row=r, column=2).value
                break
        # Day 1 was excluded (NaN→long gap→excluded), so summary should
        # report 0 negatives, not the 24 hours from the unfiltered input.
        assert neg_hours_in_sheet == 0

    def test_with_timezone(self, sample_data) -> None:
        """Export with tz parameter should produce valid xlsx with timezone in summary."""
        price_df, daily, monthly, pctls, rev, neg = sample_data
        result = export_to_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            tz="Europe/Berlin",
        )
        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 20)]
        assert "Europe/Berlin" in values

    def test_summary_dates_follow_local_timezone(self) -> None:
        """Summary dates should reflect local calendar dates, not raw UTC dates."""
        idx = pd.date_range("2025-01-01", periods=48, freq="h", tz="UTC")
        price_df = pd.DataFrame({"price_eur_mwh": [50.0] * len(idx)}, index=idx)
        price_df.index.name = "timestamp"

        daily = calculate_daily_spreads(price_df, tz="Europe/Berlin")
        monthly = calculate_monthly_spreads(price_df, tz="Europe/Berlin")
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        neg = calculate_negative_price_hours(price_df)

        result = export_to_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            tz="Europe/Berlin",
        )

        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]
        summary = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, 20)}

        assert summary["Timezone"] == "Europe/Berlin"
        assert summary["Date Range Start"] == "2025-01-01"
        assert summary["Date Range End"] == "2025-01-03"
        assert summary["Total Days"] == 3

    def test_heatmap_respects_local_hour_shift(self) -> None:
        """Price Heatmap sheet should reflect the requested local timezone."""
        idx = pd.date_range("2025-01-01", periods=48, freq="h", tz="UTC")
        prices = [50.0] * 48
        prices[23] = 200.0  # UTC 23:00 Jan 1 = Europe/Berlin 00:00 Jan 2
        price_df = pd.DataFrame({"price_eur_mwh": prices}, index=idx)
        price_df.index.name = "timestamp"

        daily = calculate_daily_spreads(price_df, tz="Europe/Berlin")
        monthly = calculate_monthly_spreads(price_df, tz="Europe/Berlin")
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        neg = calculate_negative_price_hours(price_df)

        result = export_to_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            tz="Europe/Berlin",
        )

        wb = load_workbook(BytesIO(result), data_only=True)
        ws = wb["Price Heatmap"]

        january_col = None
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "2025-01":
                january_col = col
                break

        assert january_col is not None

        local_midnight_value = None
        utc_23_value = None
        for row in range(2, ws.max_row + 1):
            hour = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=january_col).value
            if hour == 0:
                local_midnight_value = value
            if hour == 23:
                utc_23_value = value

        assert local_midnight_value is not None
        assert utc_23_value is not None
        assert local_midnight_value > utc_23_value
        assert local_midnight_value > 100

    def test_negative_summary_tracks_hours_and_intervals(self) -> None:
        """Summary should distinguish sub-hourly negative intervals from hours."""
        idx = pd.date_range("2025-03-01", periods=48, freq="30min", tz="UTC")
        price_df = pd.DataFrame(
            {"price_eur_mwh": [-10.0] * 24 + [40.0] * 24},
            index=idx,
        )
        price_df.index.name = "timestamp"

        daily = calculate_daily_spreads(price_df)
        monthly = calculate_monthly_spreads(price_df)
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        neg = calculate_negative_price_hours(price_df)

        result = export_to_bytes(
            zone="GB",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]
        summary = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, 25)
        }

        assert summary["Negative Price Hours"] == 12
        assert summary["Negative Price Intervals"] == 24
        assert summary["Negative Price % of Intervals"] == 0.5

    def test_summary_uses_full_revenue_stack_when_present(self, sample_data) -> None:
        """Export should reflect DA+ancillary totals, not just DA-only revenue."""
        price_df, daily, monthly, pctls, rev, neg = sample_data
        stacked_rev = {
            **rev,
            "annual_revenue_eur": 120000.0,
            "annual_revenue_eur_per_mw": 120000.0,
            "total_eur": 120000.0,
            "total_per_mw": 120000.0,
            "gross_additive_total_eur": 145000.0,
            "headline_total_mode": "conservative_da_primary",
            "capacity_stack_warning": "Capacity reserve is not co-optimized.",
            "joint_cooptimized_total_eur": 132000.0,
            "joint_cooptimized_avg_reserve_fraction": 0.42,
            "annual_degradation_cost_eur": 18000.0,
            "net_revenue_eur": 102000.0,
            "degradation_pct": 15.0,
            "effective_life_years": 16.4,
            "lifetime_limiting_factor": "cycling",
            "annual_throughput_mwh": 2922.0,
            "lcos_eur_mwh": 54.2,
            "net_payback_years": 7.5,
            "source_revenues": {
                "DA Arbitrage": 80000.0,
                "FCR-N": 25000.0,
                "aFRR Up": 15000.0,
            },
        }

        result = export_to_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=stacked_rev,
            negative_stats=neg,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]
        summary = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, 45)
        }

        assert summary["Est. Annual Revenue (EUR/MW)"] == 120000
        assert summary["Headline Annual Revenue (EUR)"] == 120000
        assert summary["Gross Additive Total (Reference, EUR)"] == 145000
        assert summary["Headline Total Mode"] == "conservative_da_primary"
        assert summary["Capacity Stack Warning"] == "Capacity reserve is not co-optimized."
        assert summary["Joint MILP Co-optimized Total (EUR)"] == 132000
        assert summary["Joint MILP Avg Reserve Commitment"] == 0.42
        assert summary["DA Arbitrage Revenue (EUR)"] == 80000
        assert summary["FCR-N Revenue (EUR)"] == 25000
        assert summary["Annual Degradation Cost (EUR)"] == 18000
        assert summary["Net Revenue after Degradation (EUR)"] == 102000
        assert summary["Degradation % of Gross Revenue"] == 0.15
        assert summary["Effective Battery Lifetime (years)"] == 16.4
        assert summary["Lifetime Limiting Factor"] == "cycling"
        assert summary["Annual Throughput (MWh)"] == 2922
        assert summary["LCOS (EUR/MWh)"] == 54.2
        assert summary["Net Payback (years)"] == 7.5

    def test_summary_includes_data_quality_when_gaps_exist(self) -> None:
        idx = pd.date_range("2025-01-01", periods=8, freq="h", tz="UTC")
        raw = pd.DataFrame(
            {"price_eur_mwh": [50.0, None, 52.0, None, None, None, None, 57.0]},
            index=idx,
        )
        raw.index.name = "timestamp"
        price_df = clean_prices(raw)
        daily = calculate_daily_spreads(price_df)
        monthly = calculate_monthly_spreads(price_df)
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        neg = calculate_negative_price_hours(price_df)

        result = export_to_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
        )

        wb = load_workbook(BytesIO(result))
        ws = wb["Summary"]
        summary = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, 25)
        }

        assert summary["Source Gap Intervals"] == 5
        assert summary["Short-Gap Imputed Intervals"] == 1
        assert summary["Unresolved Missing Intervals"] == 4


# ── PDF export tests ──────────────────────────────────────────────────────────

class TestPdfExport:
    @pytest.fixture()
    def sample_data(self):
        idx = pd.date_range("2025-01-01", periods=72, freq="h", tz="UTC")
        price_df = pd.DataFrame({"price_eur_mwh": range(72)}, index=idx)
        price_df.index.name = "timestamp"

        daily = calculate_daily_spreads(price_df)
        monthly = calculate_monthly_spreads(price_df)
        pctls = calculate_spread_percentiles(daily)
        rev = estimate_annual_arbitrage_revenue(daily)
        neg = calculate_negative_price_hours(price_df)
        return price_df, daily, monthly, pctls, rev, neg

    def test_pdf_export_returns_bytes(self, sample_data) -> None:
        price_df, daily, monthly, pctls, rev, neg = sample_data
        result = export_to_pdf_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
        )
        assert isinstance(result, bytes)
        assert len(result) > 100
        assert result[:5] == b"%PDF-"

    def test_pdf_export_with_timezone(self, sample_data) -> None:
        price_df, daily, monthly, pctls, rev, neg = sample_data
        result = export_to_pdf_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            tz="Europe/Berlin",
        )
        assert result[:5] == b"%PDF-"

    def test_pdf_export_with_figures(self, sample_data) -> None:
        import plotly.graph_objects as go

        price_df, daily, monthly, pctls, rev, neg = sample_data
        fig = go.Figure(go.Scatter(x=[1, 2, 3], y=[4, 5, 6]))
        _render_or_skip(fig)
        result = export_to_pdf_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
            figures={"price_ts": fig},
        )
        assert result[:5] == b"%PDF-"
        # PDF with chart should be larger than text-only
        text_only = export_to_pdf_bytes(
            zone="DE_LU",
            price_df=price_df,
            daily_spreads=daily,
            monthly_spreads=monthly,
            percentiles=pctls,
            revenue_estimate=rev,
            negative_stats=neg,
        )
        assert len(result) > len(text_only)

    def test_render_figure_to_image(self) -> None:
        import plotly.graph_objects as go

        fig = go.Figure(go.Bar(x=["A", "B"], y=[1, 2]))
        img = _render_or_skip(fig)
        assert isinstance(img, bytes)
        assert img[:8] == b"\x89PNG\r\n\x1a\n"
