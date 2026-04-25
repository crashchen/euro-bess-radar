"""Tests for zone comparison export."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.export import export_comparison_to_bytes


@pytest.fixture
def comparison_df() -> pd.DataFrame:
    return pd.DataFrame([
        {
            "zone": "DE_LU",
            "avg_price": 55.0,
            "std_price": 20.0,
            "avg_spread": 30.0,
            "p50_spread": 28.0,
            "p90_spread": 45.0,
            "negative_pct": 5.2,
            "estimated_annual_revenue_per_mw": 95000,
            "dispatch_method": "lp",
            "avg_cycles_per_day": 1.5,
            "net_revenue_per_mw": 78000,
            "lcos_eur_mwh": 52.3,
            "payback_years": 7.2,
            "effective_life_years": 16.4,
            "limiting_factor": "cycling",
        },
        {
            "zone": "FR",
            "avg_price": 48.0,
            "std_price": 15.0,
            "avg_spread": 22.0,
            "p50_spread": 20.0,
            "p90_spread": 35.0,
            "negative_pct": 2.1,
            "estimated_annual_revenue_per_mw": 72000,
            "dispatch_method": "lp",
            "avg_cycles_per_day": 1.2,
            "net_revenue_per_mw": 60000,
            "lcos_eur_mwh": 61.0,
            "payback_years": 9.5,
            "effective_life_years": 18.2,
            "limiting_factor": "calendar",
        },
    ])


class TestComparisonExport:
    def test_returns_valid_excel(self, comparison_df: pd.DataFrame) -> None:
        data = export_comparison_to_bytes(comparison_df)
        assert isinstance(data, bytes)
        wb = load_workbook(BytesIO(data))
        assert "Zone Comparison" in wb.sheetnames

    def test_columns_match(self, comparison_df: pd.DataFrame) -> None:
        data = export_comparison_to_bytes(comparison_df)
        wb = load_workbook(BytesIO(data))
        ws = wb["Zone Comparison"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Zone" in headers
        assert "Revenue (EUR/MW/yr)" in headers
        assert "LCOS (EUR/MWh)" in headers
        assert "Payback (years)" in headers

    def test_data_rows(self, comparison_df: pd.DataFrame) -> None:
        data = export_comparison_to_bytes(comparison_df)
        wb = load_workbook(BytesIO(data))
        ws = wb["Zone Comparison"]
        assert ws.max_row == 3  # header + 2 data rows
        assert ws.cell(row=2, column=1).value == "DE_LU"
        assert ws.cell(row=3, column=1).value == "FR"

    def test_negative_percentage_is_excel_ratio(self, comparison_df: pd.DataFrame) -> None:
        data = export_comparison_to_bytes(comparison_df)
        wb = load_workbook(BytesIO(data))
        ws = wb["Zone Comparison"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        neg_col = headers.index("Neg Price %") + 1

        assert ws.cell(row=2, column=neg_col).value == pytest.approx(0.052)
        assert ws.cell(row=2, column=neg_col).number_format == "0.0%"

    def test_empty_dataframe(self) -> None:
        empty = pd.DataFrame(columns=["zone", "avg_price", "estimated_annual_revenue_per_mw"])
        data = export_comparison_to_bytes(empty)
        wb = load_workbook(BytesIO(data))
        ws = wb["Zone Comparison"]
        assert ws.max_row == 1  # header only
