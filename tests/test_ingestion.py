"""Tests for data_ingestion module. All API calls are mocked."""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
import requests

from src.config import (
    ELEXON_ZONES,
    ENTSOE_ZONES,
    FINGRID_BASE_URL,
    GBP_EUR_YEARLY,
    PRICE_CACHE_TTL_HOURS,
    is_elexon_zone,
)
from src.data_ingestion import (
    DataSourceAuthError,
    DataSourceNetworkError,
    DataSourceParseError,
    build_zone_query_window,
    clean_prices,
    fetch_elexon_generation,
    fetch_elexon_prices,
    fetch_elexon_system_prices,
    fetch_entsoe_imbalance_prices,
    fetch_entsoe_prices,
    fetch_esios_ancillary_prices,
    fetch_esios_indicator,
    fetch_fingrid_afrr_prices,
    fetch_fingrid_data,
    fetch_fingrid_fcr_prices,
    fetch_generation_data,
    fetch_intraday_prices,
    fetch_prices,
    fetch_regelleistung_results,
    generate_intraday_template_csv,
    parse_intraday_csv,
    persist_intraday_frame,
    read_cache,
    read_intraday_cache,
    summarize_price_data_quality,
    write_cache,
)

# ── Test 1: ENTSO-E schema ──────────────────────────────────────────────────

class TestFetchEntsoePrices:
    @patch("src.data_ingestion._call_entsoe_api")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_returns_correct_schema(
        self, _mock_key: MagicMock, mock_api: MagicMock, mock_entsoe_series: pd.Series,
    ) -> None:
        """Returned DataFrame has correct columns and UTC timestamps."""
        mock_api.return_value = mock_entsoe_series

        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-02", tz="UTC")
        df = fetch_entsoe_prices("DE_LU", start, end)

        assert "price_eur_mwh" in df.columns
        assert df.index.name == "timestamp"
        assert df.index.tz is not None  # timezone-aware
        assert str(df.index.tz) == "UTC"
        assert len(df) == 24

    @patch("src.data_ingestion.get_api_key", side_effect=OSError("missing key"))
    def test_missing_api_key_raises_auth_error(self, _mock_key: MagicMock) -> None:
        with pytest.raises(DataSourceAuthError, match="API key is missing or invalid"):
            fetch_entsoe_prices(
                "DE_LU",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )


class TestRetryAuthBypass:
    """The shared retry decorator must not retry on authentication errors.

    Before this fix, invalid-token responses from entsoe-py were caught by
    the bare ``@retry(exceptions=(Exception,))`` wrapper and re-tried 3
    times with exponential backoff before reaching the auth-classifier.
    """

    def test_auth_error_raises_on_first_attempt(self) -> None:
        from src.data_ingestion import _looks_like_auth_error, retry

        call_count = {"n": 0}

        @retry(
            max_retries=3, backoff=0.0,
            auth_check=lambda exc: _looks_like_auth_error(exc),
        )
        def fake_call() -> None:
            call_count["n"] += 1
            raise RuntimeError("Invalid token: 401 Unauthorized")

        with pytest.raises(RuntimeError, match="Invalid token"):
            fake_call()
        assert call_count["n"] == 1, (
            "auth errors must not be retried — saw "
            f"{call_count['n']} attempts"
        )

    def test_network_error_still_retries(self) -> None:
        from src.data_ingestion import _looks_like_auth_error, retry

        call_count = {"n": 0}

        @retry(
            max_retries=3, backoff=0.0,
            auth_check=lambda exc: _looks_like_auth_error(exc),
        )
        def fake_call() -> None:
            call_count["n"] += 1
            raise requests.ConnectionError("connection reset by peer")

        with pytest.raises(requests.ConnectionError):
            fake_call()
        # 1 initial attempt + 3 retries = 4 total
        assert call_count["n"] == 4

    def test_entsoe_invalid_token_no_retry(
        self, caplog: pytest.LogCaptureFixture,
    ) -> None:
        """End-to-end: an entsoe-py-flavoured auth error reaches the user as
        DataSourceAuthError, with only one underlying attempt and no
        ``securityToken=...`` in the captured logs.
        """
        with (
            patch("src.data_ingestion.get_api_key", return_value="fake-key"),
            patch(
                "entsoe.EntsoePandasClient.query_day_ahead_prices",
                side_effect=Exception(
                    "401 Client Error: Unauthorized for url: "
                    "https://web-api.tp.entsoe.eu/api?securityToken=LEAKED&"
                    "documentType=A44"
                ),
            ) as mock_call,
            caplog.at_level(logging.ERROR, logger="src.data_ingestion"),
            pytest.raises(DataSourceAuthError),
        ):
            fetch_entsoe_prices(
                "DE_LU",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        # No retry on auth -> exactly one upstream call.
        assert mock_call.call_count == 1
        # No leaked credentials anywhere in captured logs.
        assert "LEAKED" not in caplog.text
        assert "securityToken=LEAKED" not in caplog.text


class TestScrubSecrets:
    def test_redacts_security_token(self) -> None:
        from src.data_ingestion import _scrub_secrets
        url = (
            "HTTPSConnectionPool(host='web-api.tp.entsoe.eu', port=443): "
            "Max retries exceeded with url: /api?securityToken=abcdef123456"
            "&documentType=A44 (Caused by ...)"
        )
        out = _scrub_secrets(url)
        assert "abcdef123456" not in out
        assert "securityToken=***" in out
        # Keep the rest of the URL intact for debuggability.
        assert "documentType=A44" in out

    def test_redacts_multiple_secret_keys(self) -> None:
        from src.data_ingestion import _scrub_secrets
        text = "api_key=AAA&token=BBB&apiKey=CCC&other=keep"
        out = _scrub_secrets(text)
        for leaked in ("AAA", "BBB", "CCC"):
            assert leaked not in out
        assert "other=keep" in out


# ── Test 2: Elexon schema ───────────────────────────────────────────────────

class TestFetchElexonPrices:
    @patch("src.data_ingestion._call_elexon_api")
    def test_returns_correct_schema(
        self, mock_api: MagicMock, mock_elexon_json: list[dict],
    ) -> None:
        """Returned DataFrame has correct columns, EUR conversion, 30-min res."""
        mock_api.return_value = mock_elexon_json

        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-01", tz="UTC")
        df = fetch_elexon_prices(start, end)

        assert "price_eur_mwh" in df.columns
        assert df.index.name == "timestamp"
        assert df.index.tz is not None
        assert len(df) == 48

        # Verify GBP->EUR conversion
        original_gbp = mock_elexon_json[0]["price"]
        expected_eur = original_gbp * GBP_EUR_YEARLY[2025]
        assert abs(df["price_eur_mwh"].iloc[0] - expected_eur) < 0.01

    @patch("src.data_ingestion._call_elexon_api")
    def test_uses_year_specific_fx_rate(self, mock_api: MagicMock) -> None:
        mock_api.return_value = [
            {"startTime": "2023-01-01T00:00:00Z", "price": 100.0},
            {"startTime": "2023-01-01T00:30:00Z", "price": 100.0},
        ]

        df = fetch_elexon_prices(
            pd.Timestamp("2023-01-01", tz="UTC"),
            pd.Timestamp("2023-01-01", tz="UTC"),
        )

        assert df["price_eur_mwh"].iloc[0] == 100.0 * GBP_EUR_YEARLY[2023]

    @patch("src.data_ingestion._call_elexon_api")
    def test_preserves_negative_prices_while_dropping_zero_provider_rows(
        self, mock_api: MagicMock,
    ) -> None:
        mock_api.return_value = [
            {"startTime": "2025-01-01T00:00:00Z", "price": -20.0, "volume": 2.0},
            {"startTime": "2025-01-01T00:00:00Z", "price": 0.0, "volume": 0.0},
            {"startTime": "2025-01-01T00:30:00Z", "price": 50.0, "volume": 3.0},
            {"startTime": "2025-01-01T00:30:00Z", "price": 0.0, "volume": 0.0},
        ]

        df = fetch_elexon_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-01", tz="UTC"),
        )

        assert len(df) == 2
        assert df["price_eur_mwh"].iloc[0] == pytest.approx(-20.0 * GBP_EUR_YEARLY[2025])
        assert df["price_eur_mwh"].iloc[1] == pytest.approx(50.0 * GBP_EUR_YEARLY[2025])

    @patch("src.data_ingestion._call_elexon_api")
    def test_preserves_legitimate_zero_prices(self, mock_api: MagicMock) -> None:
        """A real £0/MWh market price should not be filtered away."""
        mock_api.return_value = [
            {"startTime": "2025-01-01T00:00:00Z", "price": 0.0, "volume": 2.0},
            {"startTime": "2025-01-01T00:30:00Z", "price": -5.0, "volume": 2.0},
        ]

        df = fetch_elexon_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-01", tz="UTC"),
        )

        assert len(df) == 2
        assert df["price_eur_mwh"].iloc[0] == 0.0
        assert df["price_eur_mwh"].iloc[1] == pytest.approx(-5.0 * GBP_EUR_YEARLY[2025])

    @patch("src.data_ingestion._call_elexon_api")
    def test_unknown_year_uses_nearest_fx_rate(
        self, mock_api: MagicMock, caplog: pytest.LogCaptureFixture,
    ) -> None:
        mock_api.return_value = [
            {"startTime": "2027-01-01T00:00:00Z", "price": 100.0},
            {"startTime": "2027-01-01T00:30:00Z", "price": 100.0},
        ]

        with caplog.at_level(logging.WARNING):
            df = fetch_elexon_prices(
                pd.Timestamp("2027-01-01", tz="UTC"),
                pd.Timestamp("2027-01-01", tz="UTC"),
            )

        assert df["price_eur_mwh"].iloc[0] == 100.0 * GBP_EUR_YEARLY[2026]
        assert "No GBP/EUR rate configured for 2027" in caplog.text

    @patch("src.data_ingestion._call_elexon_api")
    def test_filters_to_requested_window(self, mock_api: MagicMock) -> None:
        """Partial-day windows should be trimmed back to [start, end)."""
        day_1 = [
            {"startTime": "2025-01-01T00:00:00Z", "price": 40.0},
            {"startTime": "2025-01-01T00:30:00Z", "price": 41.0},
        ]
        day_2 = [
            {"startTime": "2025-01-02T00:00:00Z", "price": 42.0},
            {"startTime": "2025-01-02T00:30:00Z", "price": 43.0},
        ]
        mock_api.return_value = day_1 + day_2

        df = fetch_elexon_prices(
            start=pd.Timestamp("2025-01-01T00:30:00Z"),
            end=pd.Timestamp("2025-01-02T00:30:00Z"),
        )

        mock_api.assert_called_once_with("2025-01-01", "2025-01-03")
        assert list(df.index.astype(str)) == [
            "2025-01-01 00:30:00+00:00",
            "2025-01-02 00:00:00+00:00",
        ]

    @patch("src.data_ingestion._call_elexon_api")
    def test_chunks_long_date_ranges_to_seven_day_windows(self, mock_api: MagicMock) -> None:
        """Long GB requests should use short chunks that Elexon accepts."""
        mock_api.return_value = [
            {"startTime": "2025-01-15T00:00:00Z", "price": 40.0},
        ]

        fetch_elexon_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-04-01", tz="UTC"),
        )

        assert mock_api.call_count >= 13
        for call in mock_api.call_args_list:
            from_date, to_date = call.args
            assert pd.Timestamp(to_date) - pd.Timestamp(from_date) <= pd.Timedelta(days=7)

    @patch("src.data_ingestion._call_elexon_api")
    def test_partial_chunk_failure_raises_instead_of_returning_partial_data(
        self, mock_api: MagicMock,
    ) -> None:
        """A failed later chunk must not be forward-filled into fake flat prices."""
        mock_api.side_effect = [
            [{"startTime": "2025-01-01T00:00:00Z", "price": 40.0}],
            requests.RequestException("boom"),
        ]

        with pytest.raises(DataSourceNetworkError, match="Elexon request failed"):
            fetch_elexon_prices(
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-10", tz="UTC"),
            )

    @patch("src.data_ingestion._call_elexon_api", side_effect=requests.RequestException("boom"))
    def test_request_failure_raises_network_error(self, _mock_api: MagicMock) -> None:
        with pytest.raises(DataSourceNetworkError, match="Elexon request failed"):
            fetch_elexon_prices(
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-01", tz="UTC"),
            )

    @patch("src.data_ingestion._call_elexon_api")
    def test_malformed_payload_raises_parse_error(self, mock_api: MagicMock) -> None:
        mock_api.return_value = {"unexpected": "shape"}

        with pytest.raises(DataSourceParseError, match="could not be parsed"):
            fetch_elexon_prices(
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-01", tz="UTC"),
            )


# ── Test 3: GB routes to Elexon ─────────────────────────────────────────────

class TestFetchPricesRouting:
    @patch("src.data_ingestion.fetch_elexon_prices")
    @patch("src.data_ingestion.fetch_entsoe_prices")
    def test_gb_routes_to_elexon(
        self, mock_entsoe: MagicMock, mock_elexon: MagicMock,
        mock_price_df: pd.DataFrame,
    ) -> None:
        """fetch_prices(zone='GB') should call Elexon, not ENTSO-E."""
        mock_elexon.return_value = mock_price_df

        df = fetch_prices(
            zone="GB",
            start=pd.Timestamp("2025-01-01", tz="UTC"),
            end=pd.Timestamp("2025-01-02", tz="UTC"),
            use_cache=False,
        )
        mock_elexon.assert_called_once()
        mock_entsoe.assert_not_called()
        assert len(df) > 0

    # ── Test 4: DE_LU routes to ENTSO-E ──────────────────────────────────────

    @patch("src.data_ingestion.fetch_elexon_prices")
    @patch("src.data_ingestion.fetch_entsoe_prices")
    def test_de_lu_routes_to_entsoe(
        self, mock_entsoe: MagicMock, mock_elexon: MagicMock,
        mock_price_df: pd.DataFrame,
    ) -> None:
        """fetch_prices(zone='DE_LU') should call ENTSO-E, not Elexon."""
        mock_entsoe.return_value = mock_price_df

        df = fetch_prices(
            zone="DE_LU",
            start=pd.Timestamp("2025-01-01", tz="UTC"),
            end=pd.Timestamp("2025-01-02", tz="UTC"),
            use_cache=False,
        )
        mock_entsoe.assert_called_once()
        mock_elexon.assert_not_called()
        assert len(df) > 0

    @patch("src.data_ingestion.read_cache")
    @patch("src.data_ingestion.fetch_entsoe_prices")
    def test_use_cache_false_bypasses_cache_lookup(
        self,
        mock_entsoe: MagicMock,
        mock_read_cache: MagicMock,
        mock_price_df: pd.DataFrame,
    ) -> None:
        mock_read_cache.return_value = mock_price_df
        mock_entsoe.return_value = mock_price_df

        fetch_prices(
            zone="DE_LU",
            start=pd.Timestamp("2025-01-01", tz="UTC"),
            end=pd.Timestamp("2025-01-02", tz="UTC"),
            use_cache=False,
        )

        mock_read_cache.assert_not_called()
        mock_entsoe.assert_called_once()


# ── Test 5: clean_prices fills gaps ──────────────────────────────────────────

class TestCleanPrices:
    def test_fills_gaps(self) -> None:
        """Short internal NaN gaps should be interpolated and flagged."""
        idx = pd.date_range("2025-01-01", periods=24, freq="h", tz="UTC")
        df = pd.DataFrame({"price_eur_mwh": [50.0 + i for i in range(24)]}, index=idx)
        df.index.name = "timestamp"

        df.loc[df.index[5:7], "price_eur_mwh"] = None

        result = clean_prices(df)
        assert result["price_eur_mwh"].isna().sum() == 0
        assert len(result) == 24
        assert result.loc[df.index[5], "filled"]
        assert result.loc[df.index[5], "imputed"]
        assert result.loc[df.index[5], "price_eur_mwh"] == pytest.approx(55.0)
        assert result.loc[df.index[6], "price_eur_mwh"] == pytest.approx(56.0)

    def test_transition_boundary_row_detected_when_missing(self) -> None:
        """A row missing exactly at the transition boundary itself
        (e.g. the first 15-min sample at 2025-10-01 00:00) must surface as
        filled/imputed instead of silently disappearing — and the same for
        a row missing right before the boundary.
        """
        # Case A: missing 00:00 (first post-transition sample)
        idx = pd.DatetimeIndex([
            "2025-09-30T22:00:00Z", "2025-09-30T23:00:00Z",
            "2025-10-01T00:15:00Z", "2025-10-01T00:30:00Z",
            "2025-10-01T00:45:00Z",
        ], name="timestamp")
        df = pd.DataFrame({"price_eur_mwh": [10.0, 11.0, 13.0, 14.0, 15.0]}, index=idx)
        result = clean_prices(df, zone="DE_LU")
        boundary_ts = pd.Timestamp("2025-10-01T00:00:00Z")
        assert boundary_ts in result.index
        assert bool(result.loc[boundary_ts, "filled"]) is True

        # Case B: missing 23:00 (last pre-transition sample)
        idx_b = pd.DatetimeIndex([
            "2025-09-30T22:00:00Z",
            "2025-10-01T00:00:00Z", "2025-10-01T00:15:00Z", "2025-10-01T00:30:00Z",
        ], name="timestamp")
        df_b = pd.DataFrame({"price_eur_mwh": [10.0, 12.0, 13.0, 14.0]}, index=idx_b)
        result_b = clean_prices(df_b, zone="DE_LU")
        last_pre_ts = pd.Timestamp("2025-09-30T23:00:00Z")
        assert last_pre_ts in result_b.index
        assert bool(result_b.loc[last_pre_ts, "filled"]) is True

    def test_post_transition_sparse_gap_still_detected(self) -> None:
        """When the index crosses a known DE_LU 60->15min transition AND the
        post-transition 15-min segment has its own sparse gap (e.g. missing
        00:45), the per-side mode-delta heuristic must still surface that
        gap. A blanket skip-at-transition would silently drop the 00:45 row.
        """
        idx = pd.DatetimeIndex([
            "2025-09-30T22:00:00Z",
            "2025-09-30T23:00:00Z",
            "2025-10-01T00:00:00Z",
            "2025-10-01T00:15:00Z",
            "2025-10-01T00:30:00Z",
            # 00:45 missing
            "2025-10-01T01:00:00Z",
            "2025-10-01T01:15:00Z",
            "2025-10-01T01:30:00Z",
        ], name="timestamp")
        df = pd.DataFrame(
            {"price_eur_mwh": [10.0, 11.0, 12.0, 13.0, 14.0, 16.0, 17.0, 18.0]},
            index=idx,
        )
        result = clean_prices(df, zone="DE_LU")
        gap_ts = pd.Timestamp("2025-10-01T00:45:00Z")
        assert gap_ts in result.index
        assert bool(result.loc[gap_ts, "filled"]) is True
        # Pre-transition rows must remain untouched (no fabricated 22:15).
        assert pd.Timestamp("2025-09-30T22:15:00Z") not in result.index

    def test_mode_delta_skipped_at_zone_resolution_transition(self) -> None:
        """When the index crosses a known DE_LU 60min->15min transition and
        no expected_window is given, the modal-delta shortcut would otherwise
        upsample the pre-transition hourly region to 15-min and fabricate
        intra-hour prices. The transition guard must keep pre-transition
        rows at hourly cadence.
        """
        idx = pd.DatetimeIndex([
            "2025-09-30T22:00:00Z",
            "2025-09-30T23:00:00Z",
            "2025-10-01T00:00:00Z",
            "2025-10-01T00:15:00Z",
            "2025-10-01T00:30:00Z",
            "2025-10-01T00:45:00Z",
            "2025-10-01T01:00:00Z",
        ], name="timestamp")
        df = pd.DataFrame(
            {"price_eur_mwh": [10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0]},
            index=idx,
        )
        result = clean_prices(df, zone="DE_LU")
        # No fabricated 22:15 / 22:30 / 22:45.
        assert pd.Timestamp("2025-09-30T22:15:00Z") not in result.index
        assert pd.Timestamp("2025-09-30T22:30:00Z") not in result.index
        # Original 7 rows preserved untouched.
        assert len(result) == 7
        assert not result["filled"].any()

    def test_sparse_gap_detected_without_expected_window(self) -> None:
        """When the caller does not pass expected_start/end, a sparse internal
        gap (e.g. a missing 02:00 in hourly data) must still be detected and
        marked, not silently swallowed. Without this the data-quality flag
        misses a whole hour of missing pricing.
        """
        idx = pd.DatetimeIndex([
            "2025-01-01T00:00:00Z",
            "2025-01-01T01:00:00Z",
            # 02:00 missing
            "2025-01-01T03:00:00Z",
            "2025-01-01T04:00:00Z",
        ], name="timestamp")
        df = pd.DataFrame({"price_eur_mwh": [10.0, 11.0, 13.0, 14.0]}, index=idx)

        result = clean_prices(df)
        assert len(result) == 5
        gap_ts = pd.Timestamp("2025-01-01T02:00:00Z")
        assert bool(result.loc[gap_ts, "filled"]) is True
        assert bool(result.loc[gap_ts, "imputed"]) is True
        # Short gap (1h) gets interpolated to the linear midpoint.
        assert result.loc[gap_ts, "price_eur_mwh"] == pytest.approx(12.0)

    def test_long_gaps_remain_nan(self) -> None:
        """Long gaps should remain visible instead of being flattened."""
        idx = pd.date_range("2025-01-01", periods=24, freq="h", tz="UTC")
        df = pd.DataFrame({"price_eur_mwh": [50.0 + i for i in range(24)]}, index=idx)
        df.index.name = "timestamp"
        df.loc[df.index[5:9], "price_eur_mwh"] = None

        result = clean_prices(df)

        assert result.loc[df.index[5:8], "price_eur_mwh"].isna().all()
        assert result.loc[df.index[5:8], "filled"].all()
        assert not result.loc[df.index[5:8], "imputed"].any()

    def test_edge_gaps_remain_nan(self) -> None:
        """Head/tail gaps should not be back-filled into fake boundary prices."""
        expected_start = pd.Timestamp("2025-01-01T00:00:00Z")
        expected_end = pd.Timestamp("2025-01-01T04:00:00Z")
        observed_idx = pd.date_range(
            "2025-01-01T01:00:00Z",
            "2025-01-01T03:00:00Z",
            freq="h",
            inclusive="left",
        )
        df = pd.DataFrame({"price_eur_mwh": [51.0, 52.0]}, index=observed_idx)
        df.index.name = "timestamp"

        result = clean_prices(
            df,
            zone="DE_LU",
            expected_start=expected_start,
            expected_end=expected_end,
        )

        assert pd.isna(result.loc[expected_start, "price_eur_mwh"])
        assert pd.isna(result.loc[pd.Timestamp("2025-01-01T03:00:00Z"), "price_eur_mwh"])
        assert bool(result.loc[expected_start, "filled"]) is True
        assert bool(result.loc[expected_start, "imputed"]) is False

    def test_data_quality_summary_tracks_imputed_and_missing(self) -> None:
        idx = pd.date_range("2025-01-01", periods=8, freq="h", tz="UTC")
        df = pd.DataFrame({"price_eur_mwh": [1.0, None, 3.0, None, None, None, None, 8.0]}, index=idx)
        df.index.name = "timestamp"

        result = clean_prices(df)
        quality = summarize_price_data_quality(result)

        assert quality["source_gap_intervals"] == 5
        assert quality["imputed_intervals"] == 1
        assert quality["missing_intervals"] == 4
        assert quality["max_source_gap_hours"] == 4.0

    def test_mixed_resolution_preserves_original_timestamps(self) -> None:
        """Mixed 60-min / 15-min histories should not crash cleaning."""
        idx = pd.DatetimeIndex(
            [
                "2025-10-01T00:00:00Z",
                "2025-10-01T01:00:00Z",
                "2025-10-01T02:00:00Z",
                "2025-10-01T02:15:00Z",
                "2025-10-01T02:30:00Z",
                "2025-10-01T02:45:00Z",
                "2025-10-01T03:00:00Z",
            ]
        )
        df = pd.DataFrame(
            {"price_eur_mwh": [50.0, 51.0, 52.0, 52.5, 53.0, 53.5, 54.0]},
            index=idx,
        )
        df.index.name = "timestamp"

        result = clean_prices(df, zone="DE_LU")

        assert result["price_eur_mwh"].isna().sum() == 0
        assert set(idx).issubset(set(result.index))

    def test_requested_boundaries_reveal_head_and_tail_gaps(self) -> None:
        """Cleaning should not hide missing first/last intervals of a request."""
        expected_start = pd.Timestamp("2025-01-01T00:00:00Z")
        expected_end = pd.Timestamp("2025-01-01T04:00:00Z")
        observed_idx = pd.date_range(
            "2025-01-01T01:00:00Z",
            "2025-01-01T03:00:00Z",
            freq="h",
            inclusive="left",
        )
        df = pd.DataFrame({"price_eur_mwh": [51.0, 52.0]}, index=observed_idx)
        df.index.name = "timestamp"

        result = clean_prices(
            df,
            zone="DE_LU",
            expected_start=expected_start,
            expected_end=expected_end,
        )

        assert list(result.index.astype(str)) == [
            "2025-01-01 00:00:00+00:00",
            "2025-01-01 01:00:00+00:00",
            "2025-01-01 02:00:00+00:00",
            "2025-01-01 03:00:00+00:00",
        ]
        assert bool(result.loc[expected_start, "filled"]) is True
        assert bool(result.loc[pd.Timestamp("2025-01-01T03:00:00Z"), "filled"]) is True

    def test_empty_df_passthrough(self) -> None:
        """Empty DataFrame should pass through without error."""
        df = pd.DataFrame(columns=["price_eur_mwh"])
        result = clean_prices(df)
        assert result.empty


# ── Test 6: cache roundtrip ──────────────────────────────────────────────────

class TestCacheRoundtrip:
    def test_write_and_read(self, tmp_path: Path, mock_price_df: pd.DataFrame) -> None:
        """Data written to cache should be readable back with matching values."""
        zone = "DE_LU"
        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-02", tz="UTC")

        # Patch DB_PATH and CACHE_DIR to use tmp_path
        with patch("src.data_ingestion.DB_PATH", tmp_path / "test.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path):
            write_cache(mock_price_df, zone)
            result = read_cache(zone, start, end)

        assert result is not None
        assert len(result) == 24
        pd.testing.assert_series_equal(
            mock_price_df["price_eur_mwh"].reset_index(drop=True),
            result["price_eur_mwh"].reset_index(drop=True),
            check_names=False,
        )

    def test_rejects_sparse_cache(self, tmp_path: Path) -> None:
        """Caches missing every other hour should force a refetch."""
        zone = "DE_LU"
        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-08", tz="UTC")
        idx = pd.date_range(start, end, freq="2h", inclusive="left")
        sparse_df = pd.DataFrame({"price_eur_mwh": range(len(idx))}, index=idx)
        sparse_df.index.name = "timestamp"

        with patch("src.data_ingestion.DB_PATH", tmp_path / "test.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path):
            write_cache(sparse_df, zone)
            result = read_cache(zone, start, end)

        assert result is None

    def test_rejects_cache_with_large_hole(self, tmp_path: Path) -> None:
        """Caches with a day-sized hole in the middle should be rejected."""
        zone = "DE_LU"
        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-08", tz="UTC")
        full_idx = pd.date_range(start, end, freq="h", inclusive="left")
        hole_idx = full_idx.delete(slice(72, 96))
        holey_df = pd.DataFrame(
            {"price_eur_mwh": range(len(hole_idx))},
            index=hole_idx,
        )
        holey_df.index.name = "timestamp"

        with patch("src.data_ingestion.DB_PATH", tmp_path / "test.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path):
            write_cache(holey_df, zone)
            result = read_cache(zone, start, end)

        assert result is None

    def test_accepts_complete_mixed_resolution_cache_and_rejects_gappy_15m_segment(
        self, tmp_path: Path,
    ) -> None:
        """Mixed 60m/15m caches should validate segment-by-segment across the boundary."""
        zone = "DE_LU"
        start = pd.Timestamp("2025-09-30T22:00:00Z")
        end = pd.Timestamp("2025-10-01T02:00:00Z")
        hourly_idx = pd.date_range(start, "2025-10-01T00:00:00Z", freq="h", inclusive="left")
        quarter_hour_idx = pd.date_range(
            "2025-10-01T00:00:00Z",
            end,
            freq="15min",
            inclusive="left",
        )
        full_idx = hourly_idx.append(quarter_hour_idx)
        full_df = pd.DataFrame(
            {"price_eur_mwh": range(len(full_idx))},
            index=full_idx,
        )
        full_df.index.name = "timestamp"
        gappy_df = full_df.drop(pd.Timestamp("2025-10-01T00:45:00Z"))

        with patch("src.data_ingestion.DB_PATH", tmp_path / "valid.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path / "valid-cache"):
            write_cache(full_df, zone)
            valid = read_cache(zone, start, end)

        with patch("src.data_ingestion.DB_PATH", tmp_path / "invalid.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path / "invalid-cache"):
            write_cache(gappy_df, zone)
            invalid = read_cache(zone, start, end)

        assert valid is not None
        assert len(valid) == len(full_df)
        assert invalid is None

    def test_rejects_stale_cache(self, tmp_path: Path, mock_price_df: pd.DataFrame) -> None:
        zone = "DE_LU"
        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-02", tz="UTC")
        stale_at = (
            pd.Timestamp.now(tz="UTC") - pd.Timedelta(hours=PRICE_CACHE_TTL_HOURS + 1)
        ).isoformat()

        with patch("src.data_ingestion.DB_PATH", tmp_path / "test.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path):
            write_cache(mock_price_df, zone)
            with sqlite3.connect(tmp_path / "test.db") as conn:
                conn.execute(
                    'UPDATE "da_prices_de_lu" SET fetched_at = ?',
                    (stale_at,),
                )
            result = read_cache(zone, start, end)

        assert result is None

    def test_rejects_stale_rows_within_requested_slice(
        self, tmp_path: Path, mock_price_df: pd.DataFrame,
    ) -> None:
        """Refreshing a different slice should not make old rows fresh."""
        zone = "DE_LU"
        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-02", tz="UTC")
        stale_at = (
            pd.Timestamp.now(tz="UTC") - pd.Timedelta(hours=PRICE_CACHE_TTL_HOURS + 1)
        ).isoformat()
        fresh_at = pd.Timestamp.now(tz="UTC").isoformat()

        with patch("src.data_ingestion.DB_PATH", tmp_path / "test.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path):
            write_cache(mock_price_df, zone)
            with sqlite3.connect(tmp_path / "test.db") as conn:
                conn.execute(
                    'UPDATE "da_prices_de_lu" SET fetched_at = ? WHERE timestamp < ?',
                    (stale_at, "2025-01-01T12:00:00+00:00"),
                )
                conn.execute(
                    'UPDATE "da_prices_de_lu" SET fetched_at = ? WHERE timestamp >= ?',
                    (fresh_at, "2025-01-01T12:00:00+00:00"),
                )
            result = read_cache(zone, start, end)

        assert result is None

    def test_old_schema_cache_misses(self, tmp_path: Path) -> None:
        """Caches written before fetched_at existed should be refetched."""
        zone = "DE_LU"
        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-02", tz="UTC")

        with patch("src.data_ingestion.DB_PATH", tmp_path / "old.db"):
            with sqlite3.connect(tmp_path / "old.db") as conn:
                conn.execute(
                    'CREATE TABLE "da_prices_de_lu" ('
                    "timestamp TEXT PRIMARY KEY, price_eur_mwh REAL NOT NULL, "
                    "zone TEXT NOT NULL)"
                )
                conn.execute(
                    'INSERT INTO "da_prices_de_lu" VALUES (?, ?, ?)',
                    ("2025-01-01T00:00:00+00:00", 50.0, zone),
                )
            result = read_cache(zone, start, end)

        assert result is None

    @patch("src.data_ingestion.fetch_entsoe_prices")
    def test_fetch_prices_does_not_persist_filled_rows(
        self, mock_fetch: MagicMock, tmp_path: Path,
    ) -> None:
        zone = "DE_LU"
        idx = pd.date_range("2025-01-01", periods=4, freq="h", tz="UTC")
        raw = pd.DataFrame(
            {"price_eur_mwh": [50.0, None, 70.0, 80.0]},
            index=idx,
        )
        raw.index.name = "timestamp"
        mock_fetch.return_value = raw

        with patch("src.data_ingestion.DB_PATH", tmp_path / "test.db"), \
             patch("src.data_ingestion.CACHE_DIR", tmp_path):
            analytics = fetch_prices(
                zone=zone,
                start=idx[0],
                end=idx[-1] + pd.Timedelta(hours=1),
                use_cache=False,
            )
            with sqlite3.connect(tmp_path / "test.db") as conn:
                rows = conn.execute(
                    'SELECT timestamp FROM "da_prices_de_lu" ORDER BY timestamp'
                ).fetchall()

        assert analytics["price_eur_mwh"].isna().sum() == 0
        assert bool(analytics.loc[idx[1], "filled"]) is True
        assert idx[1].isoformat() not in {row[0] for row in rows}


# ── Test 7: invalid zone raises error ────────────────────────────────────────

class TestInvalidZone:
    def test_raises_value_error(self) -> None:
        """Unknown zone code should raise ValueError."""
        with pytest.raises(ValueError, match="Unknown bidding zone"):
            fetch_prices(
                zone="INVALID_ZONE",
                start=pd.Timestamp("2025-01-01", tz="UTC"),
                end=pd.Timestamp("2025-01-02", tz="UTC"),
            )


# ── Test 8: empty API response handled ───────────────────────────────────────

class TestEmptyApiResponse:
    @patch("src.data_ingestion._call_entsoe_api")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_empty_entsoe_returns_empty_df(
        self, _mock_key: MagicMock, mock_api: MagicMock, caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Empty ENTSO-E response returns empty DataFrame with warning logged."""
        mock_api.return_value = pd.Series(
            [], dtype=float, index=pd.DatetimeIndex([], tz="Europe/Brussels"),
        )
        with caplog.at_level(logging.WARNING):
            df = fetch_entsoe_prices(
                "DE_LU",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        assert df.empty
        assert any("empty" in r.message.lower() for r in caplog.records)

    @patch("src.data_ingestion._call_elexon_api")
    def test_empty_elexon_returns_empty_df(
        self, mock_api: MagicMock, caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Empty Elexon response returns empty DataFrame with warning logged."""
        mock_api.return_value = []
        with caplog.at_level(logging.WARNING):
            df = fetch_elexon_prices(
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-01", tz="UTC"),
            )
        assert df.empty
        assert any("no data" in r.message.lower() for r in caplog.records)


# ── Test 9: config zone classification ───────────────────────────────────────

class TestConfigZoneClassification:
    def test_gb_is_elexon(self) -> None:
        assert is_elexon_zone("GB") is True

    def test_de_lu_is_not_elexon(self) -> None:
        assert is_elexon_zone("DE_LU") is False

    def test_all_entsoe_zones_are_not_elexon(self) -> None:
        for code in ENTSOE_ZONES.values():
            assert is_elexon_zone(code) is False

    def test_all_elexon_zones_are_elexon(self) -> None:
        for code in ELEXON_ZONES.values():
            assert is_elexon_zone(code) is True


class TestBuildZoneQueryWindow:
    def test_de_lu_uses_local_midnight_in_winter(self) -> None:
        start, end = build_zone_query_window("DE_LU", "2025-01-01", "2025-01-31")
        assert start == pd.Timestamp("2024-12-31T23:00:00Z")
        assert end == pd.Timestamp("2025-01-31T23:00:00Z")

    def test_gb_uses_local_midnight_in_bst(self) -> None:
        start, end = build_zone_query_window("GB", "2025-07-01", "2025-07-31")
        assert start == pd.Timestamp("2025-06-30T23:00:00Z")
        assert end == pd.Timestamp("2025-07-31T23:00:00Z")

    def test_de_lu_spring_forward_dst(self) -> None:
        start, end = build_zone_query_window("DE_LU", "2025-03-30", "2025-03-30")
        assert start == pd.Timestamp("2025-03-29T23:00:00Z")
        assert end == pd.Timestamp("2025-03-30T22:00:00Z")
        assert end - start == pd.Timedelta(hours=23)

    def test_de_lu_fall_back_dst(self) -> None:
        start, end = build_zone_query_window("DE_LU", "2025-10-26", "2025-10-26")
        assert start == pd.Timestamp("2025-10-25T22:00:00Z")
        assert end == pd.Timestamp("2025-10-26T23:00:00Z")
        assert end - start == pd.Timedelta(hours=25)


# ── Test 10: Fingrid data fetcher ─────────────────────────────────────────────

class TestFetchFingridData:
    @patch("src.data_ingestion.requests.get")
    def test_returns_correct_schema(self, mock_get: MagicMock) -> None:
        """fetch_fingrid_data returns DataFrame with timestamp and value."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {
            "data": [
                {"startTime": "2025-01-01T00:00:00Z", "value": 10.5},
                {"startTime": "2025-01-01T01:00:00Z", "value": 11.2},
                {"startTime": "2025-01-01T02:00:00Z", "value": 9.8},
            ]
        }
        mock_get.return_value = mock_resp

        start = pd.Timestamp("2025-01-01", tz="UTC")
        end = pd.Timestamp("2025-01-02", tz="UTC")
        with patch.dict("os.environ", {"FINGRID_API_KEY": "test-key"}, clear=False):
            df = fetch_fingrid_data(318, start, end)

        assert "timestamp" in df.columns
        assert "value" in df.columns
        assert len(df) == 3
        mock_get.assert_called_once()
        args, kwargs = mock_get.call_args
        assert args[0] == f"{FINGRID_BASE_URL}/datasets/318/data"
        assert kwargs["params"]["startTime"] == start.isoformat()
        assert kwargs["params"]["endTime"] == end.isoformat()
        assert "datasets" not in kwargs["params"]
        assert kwargs["headers"] == {"x-api-key": "test-key"}

    @patch("src.data_ingestion.requests.get")
    def test_empty_response(self, mock_get: MagicMock) -> None:
        """Empty Fingrid response returns empty DataFrame."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {"data": []}
        mock_get.return_value = mock_resp

        df = fetch_fingrid_data(
            318,
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert df.empty

    @patch("src.data_ingestion.requests.get")
    def test_deduplicates_page_boundary_timestamps(self, mock_get: MagicMock) -> None:
        """Duplicate timestamps from pagination boundaries should not inflate joins."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {
            "data": [
                {"startTime": "2025-01-01T00:00:00Z", "value": 10.0},
                {"startTime": "2025-01-01T00:00:00Z", "value": 14.0},
                {"startTime": "2025-01-01T01:00:00Z", "value": 20.0},
            ]
        }
        mock_get.return_value = mock_resp

        df = fetch_fingrid_data(
            318,
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )

        assert list(df["timestamp"]) == [
            pd.Timestamp("2025-01-01T00:00:00Z"),
            pd.Timestamp("2025-01-01T01:00:00Z"),
        ]
        assert df["value"].iloc[0] == 12.0

    @patch("src.data_ingestion.requests.get")
    def test_logs_missing_api_key_for_v2_endpoint(
        self, mock_get: MagicMock, caplog: pytest.LogCaptureFixture,
    ) -> None:
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {"data": []}
        mock_get.return_value = mock_resp

        with patch.dict("os.environ", {}, clear=True), caplog.at_level(logging.WARNING):
            fetch_fingrid_data(
                317,
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )

        assert "without FINGRID_API_KEY" in caplog.text
        assert mock_get.call_args.kwargs["headers"] == {}

    @patch("src.data_ingestion.time.sleep")
    @patch("src.data_ingestion.requests.get")
    def test_keeps_partial_rows_when_later_page_fails(
        self, mock_get: MagicMock, _mock_sleep: MagicMock,
    ) -> None:
        """A transient failure on page N+1 must not discard pages 1..N.
        Returning empty in that case used to silently swallow large slices.
        """
        page_1 = MagicMock()
        page_1.raise_for_status = MagicMock()
        page_1.json.return_value = {
            "data": [{"startTime": "2025-01-01T00:00:00Z", "value": 10.5}] * 20000
        }
        mock_get.side_effect = [
            page_1,
            requests.RequestException("boom"),
            requests.RequestException("boom"),
            requests.RequestException("boom"),
            requests.RequestException("boom"),
        ]

        with patch.dict("os.environ", {"FINGRID_API_KEY": "test-key"}, clear=False):
            df = fetch_fingrid_data(
                318,
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )

        assert not df.empty
        assert df["value"].iloc[0] == pytest.approx(10.5)
        assert list(df.columns) == ["timestamp", "value"]

    @patch("src.data_ingestion.time.sleep")
    @patch("src.data_ingestion.requests.get")
    def test_auth_error_propagates_as_data_source_auth_error(
        self, mock_get: MagicMock, _mock_sleep: MagicMock,
    ) -> None:
        """401/403 must bypass the retry loop and surface as DataSourceAuthError
        so the UI can prompt for FINGRID_API_KEY rather than swallow as 'no data'.
        """
        unauth = MagicMock()
        unauth.status_code = 401
        unauth.raise_for_status = MagicMock()
        mock_get.return_value = unauth

        with (
            patch.dict("os.environ", {"FINGRID_API_KEY": "bad-key"}, clear=False),
            pytest.raises(DataSourceAuthError),
        ):
            fetch_fingrid_data(
                318,
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        # No retry: a single GET attempt
        assert mock_get.call_count == 1


# ── Test 11: Fingrid FCR prices ───────────────────────────────────────────────

class TestFetchFingridFcrPrices:
    @patch("src.data_ingestion.fetch_fingrid_data")
    def test_returns_correct_columns(self, mock_fetch: MagicMock) -> None:
        """fetch_fingrid_fcr_prices returns correct column schema."""
        idx = pd.date_range("2025-01-01", periods=3, freq="h", tz="UTC")
        mock_fetch.return_value = pd.DataFrame({
            "timestamp": idx,
            "value": [10.0, 11.0, 12.0],
        })

        df = fetch_fingrid_fcr_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert "timestamp" in df.columns
        # At least one of the FCR price columns should be present
        fcr_cols = {"fcr_n_price", "fcr_d_up_price", "fcr_d_down_price"}
        assert len(fcr_cols.intersection(df.columns)) > 0
        assert [call.args[0] for call in mock_fetch.call_args_list] == [317, 318, 283]

    @patch("src.data_ingestion.fetch_fingrid_data")
    def test_empty_response(self, mock_fetch: MagicMock) -> None:
        """Empty Fingrid response returns empty DataFrame."""
        mock_fetch.return_value = pd.DataFrame(columns=["timestamp", "value"])
        df = fetch_fingrid_fcr_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert df.empty


class TestFetchFingridAfrrPrices:
    @patch("src.data_ingestion.fetch_fingrid_data")
    def test_uses_current_dataset_ids(self, mock_fetch: MagicMock) -> None:
        idx = pd.date_range("2025-01-01", periods=2, freq="h", tz="UTC")
        mock_fetch.return_value = pd.DataFrame(
            {"timestamp": idx, "value": [5.0, 6.0]}
        )

        df = fetch_fingrid_afrr_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )

        assert list(df.columns) == ["timestamp", "afrr_up_price", "afrr_down_price"]
        assert [call.args[0] for call in mock_fetch.call_args_list] == [52, 51]

    @patch("src.data_ingestion.fetch_fingrid_data")
    def test_join_does_not_expand_duplicate_timestamps(self, mock_fetch: MagicMock) -> None:
        duplicate_idx = [
            pd.Timestamp("2025-01-01T00:00:00Z"),
            pd.Timestamp("2025-01-01T00:00:00Z"),
        ]
        mock_fetch.return_value = pd.DataFrame(
            {"timestamp": duplicate_idx, "value": [5.0, 7.0]}
        )

        df = fetch_fingrid_afrr_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )

        assert len(df) == 1
        assert df["afrr_up_price"].iloc[0] == 6.0
        assert df["afrr_down_price"].iloc[0] == 6.0


class TestFetchElexonGeneration:
    @patch("src.data_ingestion.requests.get")
    def test_uses_business_time_params(self, mock_get: MagicMock) -> None:
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = [
            {
                "startTime": "2025-01-01T00:00:00Z",
                "fuelType": "WIND",
                "generation": 100.0,
            },
            {
                "startTime": "2025-01-01T00:00:00Z",
                "fuelType": "SOLAR",
                "generation": 10.0,
            },
        ]
        mock_get.return_value = mock_resp

        start = pd.Timestamp("2025-01-01T00:00:00Z")
        end = pd.Timestamp("2025-01-01T01:00:00Z")
        df = fetch_elexon_generation(start, end)

        assert not df.empty
        params = mock_get.call_args.kwargs["params"]
        assert params["from"] == start.isoformat()
        assert params["to"] == end.isoformat()
        assert "publishDateTimeFrom" not in params
        assert "publishDateTimeTo" not in params


class TestFetchGenerationData:
    @patch("src.data_ingestion._call_entsoe_generation")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_excludes_consumption_and_load_columns(
        self, _mock_key: MagicMock, mock_generation: MagicMock,
    ) -> None:
        idx = pd.date_range("2025-01-01", periods=2, freq="h", tz="Europe/Brussels")
        mock_generation.return_value = pd.DataFrame(
            {
                "Solar": [10.0, 20.0],
                "Solar consumption": [1000.0, 1000.0],
                "Wind Onshore": [30.0, 40.0],
                "Load": [500.0, 500.0],
                "Gas": [60.0, 40.0],
            },
            index=idx,
        )

        df = fetch_generation_data(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-01T02:00:00Z"),
        )

        assert list(df["solar_mw"]) == [10.0, 20.0]
        assert list(df["wind_onshore_mw"]) == [30.0, 40.0]
        assert list(df["total_generation_mw"]) == [100.0, 100.0]
        assert list(df["renewable_pct"]) == [40.0, 60.0]


class TestFetchRegelleistungResults:
    def _make_xlsx_bytes(self) -> bytes:
        """Create a minimal xlsx that mimics Regelleistung tender export."""
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["PRODUCT_NAME", "CAPACITY PRICE [EUR/MW]", "FROM", "DIRECTION"])
        ws.append(["FCR", 5.50, "00:00", "POS"])
        ws.append(["FCR", 6.20, "04:00", "POS"])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @patch("src.data_ingestion._call_regelleistung_api")
    def test_fetches_and_parses_xlsx(
        self, mock_api: MagicMock,
    ) -> None:
        mock_api.return_value = self._make_xlsx_bytes()
        result = fetch_regelleistung_results(
            "FCR",
            pd.Timestamp("2024-12-31T23:00:00Z"),
            pd.Timestamp("2025-01-01T23:00:00Z"),
        )
        assert result is not None
        assert len(result) == 2
        assert str(result["timestamp"].iloc[0]) == "2024-12-31 23:00:00+00:00"
        assert str(result["timestamp"].iloc[1]) == "2025-01-01 03:00:00+00:00"
        assert "capacity_price_eur_mw" in result.columns
        assert result["capacity_price_eur_mw"].iloc[0] == 5.50

    @patch("src.data_ingestion._call_regelleistung_api")
    def test_parses_range_style_time_block(self, mock_api: MagicMock) -> None:
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["PRODUCT_NAME", "CAPACITY PRICE [EUR/MW]", "FROM", "DIRECTION"])
        ws.append(["aFRR", 7.0, "00:00-04:00", "NEG"])
        buf = BytesIO()
        wb.save(buf)
        mock_api.return_value = buf.getvalue()

        result = fetch_regelleistung_results(
            "aFRR",
            pd.Timestamp("2024-12-31T23:00:00Z"),
            pd.Timestamp("2025-01-01T23:00:00Z"),
        )

        assert result is not None
        assert len(result) == 1
        assert str(result["timestamp"].iloc[0]) == "2024-12-31 23:00:00+00:00"
        assert result["direction"].iloc[0] == "Down"

    @patch("src.data_ingestion._call_regelleistung_api")
    def test_parses_current_fcr_export_and_normalises_block_price(
        self, mock_api: MagicMock,
    ) -> None:
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "DATE_FROM",
            "PRODUCT_TYPE",
            "PRODUCTNAME",
            "GERMANY_SETTLEMENTCAPACITY_PRICE_[EUR/MW]",
        ])
        ws.append(["2026-06-20", "FCR", "NEGPOS_04_08", 80.0])
        buf = BytesIO()
        wb.save(buf)
        mock_api.return_value = buf.getvalue()

        result = fetch_regelleistung_results(
            "FCR",
            pd.Timestamp("2026-06-19T22:00:00Z"),
            pd.Timestamp("2026-06-20T22:00:00Z"),
        )

        assert result is not None
        assert len(result) == 1
        assert str(result["timestamp"].iloc[0]) == "2026-06-20 02:00:00+00:00"
        assert result["capacity_price_eur_mw"].iloc[0] == 20.0
        assert result["direction"].iloc[0] == "Symmetric"

    @patch("src.data_ingestion._call_regelleistung_api")
    def test_parses_current_afrr_export_with_direction(
        self, mock_api: MagicMock,
    ) -> None:
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "DATE_FROM",
            "TYPE_OF_RESERVES",
            "PRODUCT",
            "GERMANY_AVERAGE_CAPACITY_PRICE_[(EUR/MW)/h]",
        ])
        ws.append(["2026-06-20", "aFRR", "NEG_20_24", 34.5])
        buf = BytesIO()
        wb.save(buf)
        mock_api.return_value = buf.getvalue()

        result = fetch_regelleistung_results(
            "aFRR",
            pd.Timestamp("2026-06-19T22:00:00Z"),
            pd.Timestamp("2026-06-20T22:00:00Z"),
        )

        assert result is not None
        assert len(result) == 1
        assert str(result["timestamp"].iloc[0]) == "2026-06-20 18:00:00+00:00"
        assert result["capacity_price_eur_mw"].iloc[0] == 34.5
        assert result["direction"].iloc[0] == "Down"

    @patch("src.data_ingestion._call_regelleistung_api")
    def test_returns_none_on_network_error(
        self, mock_api: MagicMock, caplog: pytest.LogCaptureFixture,
    ) -> None:
        mock_api.side_effect = requests.RequestException("timeout")
        with caplog.at_level(logging.WARNING):
            result = fetch_regelleistung_results(
                "FCR",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        assert result is None
        assert "Regelleistung fetch failed" in caplog.text

    @patch("src.data_ingestion._call_regelleistung_api")
    def test_returns_none_on_empty_xlsx(
        self, mock_api: MagicMock, caplog: pytest.LogCaptureFixture,
    ) -> None:
        """Empty xlsx (headers only) returns None with fallback warning."""
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["PRODUCT_NAME", "CAPACITY PRICE [EUR/MW]", "FROM", "DIRECTION"])
        buf = BytesIO()
        wb.save(buf)
        mock_api.return_value = buf.getvalue()

        with caplog.at_level(logging.WARNING):
            result = fetch_regelleistung_results(
                "FCR",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        assert result is None
        assert "manual DE_FCR" in caplog.text

    @patch("src.data_ingestion.time.sleep")
    @patch("src.data_ingestion.requests.get")
    def test_auth_error_propagates_as_data_source_auth_error(
        self, mock_get: MagicMock, _mock_sleep: MagicMock,
    ) -> None:
        """A 401/403 from Regelleistung must surface as DataSourceAuthError
        instead of being swallowed as a generic RequestException by the
        retry loop and outer except.
        """
        forbidden = MagicMock()
        forbidden.status_code = 403
        forbidden.raise_for_status = MagicMock()
        mock_get.return_value = forbidden

        with pytest.raises(DataSourceAuthError):
            fetch_regelleistung_results(
                "FCR",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        # Bypassed retry: one GET, not three
        assert mock_get.call_count == 1


# ── Test 12: Elexon system prices ─────────────────────────────────────────────

class TestFetchElexonSystemPrices:
    @patch("src.data_ingestion.requests.get")
    def test_returns_correct_schema(self, mock_get: MagicMock) -> None:
        """fetch_elexon_system_prices returns correct columns."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = [
            {
                "settlementDate": "2025-01-01",
                "settlementPeriod": 1,
                "systemBuyPrice": 55.0,
                "systemSellPrice": 45.0,
            },
            {
                "settlementDate": "2025-01-01",
                "settlementPeriod": 2,
                "systemBuyPrice": 60.0,
                "systemSellPrice": 50.0,
            },
        ]
        mock_get.return_value = mock_resp

        df = fetch_elexon_system_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-01", tz="UTC"),
        )
        assert "system_buy_price_gbp" in df.columns
        assert "system_sell_price_gbp" in df.columns
        assert "spread_eur" in df.columns
        assert len(df) == 2
        assert df["system_buy_price_eur"].iloc[0] == 55.0 * GBP_EUR_YEARLY[2025]

    @patch("src.data_ingestion.requests.get")
    def test_filters_to_requested_window(self, mock_get: MagicMock) -> None:
        """System price fetches should be trimmed back to [start, end)."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = [
            {
                "settlementDate": "2025-01-01",
                "settlementPeriod": 1,
                "systemBuyPrice": 55.0,
                "systemSellPrice": 45.0,
            },
            {
                "settlementDate": "2025-01-01",
                "settlementPeriod": 2,
                "systemBuyPrice": 56.0,
                "systemSellPrice": 46.0,
            },
            {
                "settlementDate": "2025-01-02",
                "settlementPeriod": 1,
                "systemBuyPrice": 57.0,
                "systemSellPrice": 47.0,
            },
            {
                "settlementDate": "2025-01-02",
                "settlementPeriod": 2,
                "systemBuyPrice": 58.0,
                "systemSellPrice": 48.0,
            },
        ]
        mock_get.return_value = mock_resp

        df = fetch_elexon_system_prices(
            start=pd.Timestamp("2025-01-01T00:30:00Z"),
            end=pd.Timestamp("2025-01-02T00:30:00Z"),
        )

        params = mock_get.call_args.kwargs["params"]
        assert params["from"] == "2025-01-01"
        assert params["to"] == "2025-01-03"
        assert list(df.index.astype(str)) == [
            "2025-01-01 00:30:00+00:00",
            "2025-01-02 00:00:00+00:00",
        ]

    @patch("src.data_ingestion.requests.get")
    def test_bst_settlement_period_maps_from_london_local_time(
        self, mock_get: MagicMock,
    ) -> None:
        """GB period 1 during BST is 23:00 UTC on the previous calendar day."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = [
            {
                "settlementDate": "2025-07-01",
                "settlementPeriod": 1,
                "systemBuyPrice": 55.0,
                "systemSellPrice": 45.0,
            },
            {
                "settlementDate": "2025-07-01",
                "settlementPeriod": 3,
                "systemBuyPrice": 56.0,
                "systemSellPrice": 46.0,
            },
        ]
        mock_get.return_value = mock_resp

        df = fetch_elexon_system_prices(
            start=pd.Timestamp("2025-06-30T23:00:00Z"),
            end=pd.Timestamp("2025-07-01T01:00:00Z"),
        )

        assert list(df.index.astype(str)) == [
            "2025-06-30 23:00:00+00:00",
            "2025-07-01 00:00:00+00:00",
        ]

    @patch("src.data_ingestion.requests.get")
    def test_fall_back_day_keeps_repeated_local_hour_distinct(
        self, mock_get: MagicMock,
    ) -> None:
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = [
            {
                "settlementDate": "2025-10-26",
                "settlementPeriod": 3,
                "systemBuyPrice": 55.0,
                "systemSellPrice": 45.0,
            },
            {
                "settlementDate": "2025-10-26",
                "settlementPeriod": 5,
                "systemBuyPrice": 56.0,
                "systemSellPrice": 46.0,
            },
        ]
        mock_get.return_value = mock_resp

        df = fetch_elexon_system_prices(
            start=pd.Timestamp("2025-10-25T23:00:00Z"),
            end=pd.Timestamp("2025-10-26T02:00:00Z"),
        )

        assert list(df.index.astype(str)) == [
            "2025-10-26 00:00:00+00:00",
            "2025-10-26 01:00:00+00:00",
        ]

    @patch("src.data_ingestion.requests.get")
    def test_empty_response(self, mock_get: MagicMock) -> None:
        """Empty Elexon response returns empty DataFrame."""
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = []
        mock_get.return_value = mock_resp

        df = fetch_elexon_system_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-01", tz="UTC"),
        )
        assert df.empty


# ── Test 13: ENTSO-E imbalance prices ─────────────────────────────────────────

class TestFetchEntsoeImbalancePrices:
    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_returns_none_on_failure(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        """Returns None gracefully when data is unavailable."""
        mock_client = MagicMock()
        mock_client.query_imbalance_prices.side_effect = Exception("No data")
        mock_client_cls.return_value = mock_client

        result = fetch_entsoe_imbalance_prices(
            "RO",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert result is None

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_returns_dataframe_on_success(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        """Returns DataFrame with imbalance prices on success."""
        idx = pd.date_range("2025-01-01", periods=24, freq="h", tz="Europe/Brussels")
        raw = pd.DataFrame(
            {"long": range(24), "short": range(24, 48)},
            index=idx,
        )
        mock_client = MagicMock()
        mock_client.query_imbalance_prices.return_value = raw
        mock_client_cls.return_value = mock_client

        result = fetch_entsoe_imbalance_prices(
            "RO",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert result is not None
        assert "imbalance_price_long" in result.columns
        assert "imbalance_price_short" in result.columns
        assert len(result) == 24

    @patch("src.data_ingestion.get_api_key", side_effect=OSError("missing key"))
    def test_missing_api_key_raises_auth_error(
        self, _mock_key: MagicMock,
    ) -> None:
        """Pre-fix the imbalance fetcher logged a warning and returned None
        when the API key was missing, leaving the UI to render "no data" with
        no actionable hint. Now mirror DA/IDA: raise DataSourceAuthError so
        the sidebar can prompt the user to set ENTSOE_API_KEY.
        """
        with pytest.raises(DataSourceAuthError, match="ENTSOE_API_KEY"):
            fetch_entsoe_imbalance_prices(
                "RO",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )


# ── Test 14: ENTSO-E Intraday Auction prices (IDA1/2/3) ────────────────────

class TestFetchIntradayPrices:
    def test_unsupported_zone_returns_none(self) -> None:
        """Zones outside INTRADAY_SUPPORTED_ZONES return None without an API call."""
        result = fetch_intraday_prices(
            "RO",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert result is None

    def test_invalid_sequence_raises(self) -> None:
        with pytest.raises(ValueError):
            fetch_intraday_prices(
                "DE_LU",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
                sequence=4,
            )

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_returns_correct_schema(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        idx = pd.date_range("2025-01-01", periods=24, freq="h", tz="UTC")
        raw = pd.Series([45.0 + i for i in range(24)], index=idx, name="price")
        mock_client = MagicMock()
        mock_client.query_intraday_prices.return_value = raw
        mock_client_cls.return_value = mock_client

        result = fetch_intraday_prices(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
            sequence=1,
        )
        assert result is not None
        assert "intraday_price_eur_mwh" in result.columns
        assert len(result) == 24
        # UTC-normalised index
        assert str(result.index.tz) == "UTC"

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_no_data_returns_none_not_error(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        """entsoe-py raises NoMatchingDataError (a ValueError subclass) when
        a zone simply has no IDA data — surface as None, not an exception.
        """
        mock_client = MagicMock()
        mock_client.query_intraday_prices.side_effect = ValueError("no data")
        mock_client_cls.return_value = mock_client
        result = fetch_intraday_prices(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert result is None

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_network_error_propagates(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        import requests as _rq
        mock_client = MagicMock()
        mock_client.query_intraday_prices.side_effect = _rq.Timeout("boom")
        mock_client_cls.return_value = mock_client
        with pytest.raises(DataSourceNetworkError):
            fetch_intraday_prices(
                "DE_LU",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_sequence_passed_through_to_client(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        """Regression for the IDA round selector: the sequence arg must
        reach entsoe-py unchanged so a user requesting IDA2 doesn't
        silently receive IDA1 data.
        """
        idx = pd.date_range("2025-01-01", periods=4, freq="h", tz="UTC")
        mock_client = MagicMock()
        mock_client.query_intraday_prices.return_value = pd.Series(
            [50.0] * 4, index=idx, name="price",
        )
        mock_client_cls.return_value = mock_client
        fetch_intraday_prices(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
            sequence=2,
        )
        kwargs = mock_client.query_intraday_prices.call_args.kwargs
        assert kwargs["sequence"] == 2

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_no_matching_data_returns_none(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        """entsoe-py raises NoMatchingDataError (subclass of Exception, NOT
        ValueError) when a supported zone has no IDA print in the window.
        The fetcher must catch this explicitly.
        """
        from entsoe.exceptions import NoMatchingDataError
        mock_client = MagicMock()
        mock_client.query_intraday_prices.side_effect = NoMatchingDataError("nope")
        mock_client_cls.return_value = mock_client
        result = fetch_intraday_prices(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert result is None

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_http_401_propagates_as_auth_error(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
    ) -> None:
        """ENTSO-E 401/403 during the query comes back as requests.HTTPError;
        classify as DataSourceAuthError so the sidebar can prompt for
        ENTSOE_API_KEY rather than a generic network error.
        """
        import requests as _rq
        resp = MagicMock()
        resp.status_code = 401
        mock_client = MagicMock()
        mock_client.query_intraday_prices.side_effect = _rq.HTTPError(
            "401 Client Error", response=resp,
        )
        mock_client_cls.return_value = mock_client
        with pytest.raises(DataSourceAuthError):
            fetch_intraday_prices(
                "DE_LU",
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )

    @patch("src.data_ingestion.EntsoePandasClient")
    @patch("src.data_ingestion.get_api_key", return_value="fake-key")
    def test_successful_fetch_persists_to_sqlite_cache(
        self, _mock_key: MagicMock, mock_client_cls: MagicMock,
        tmp_path, monkeypatch,
    ) -> None:
        """A successful IDA fetch must write to SQLite so a subsequent
        read_intraday_cache returns the same data without another API call.
        """
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")

        idx = pd.date_range("2025-01-01", periods=4, freq="h", tz="UTC")
        mock_client = MagicMock()
        mock_client.query_intraday_prices.return_value = pd.Series(
            [40.0, 42.0, 44.0, 46.0], index=idx, name="price",
        )
        mock_client_cls.return_value = mock_client
        fetched = di.fetch_intraday_prices(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert fetched is not None and len(fetched) == 4

        cached = di.read_intraday_cache(
            "DE_LU",
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
            sequence=1,
        )
        assert cached is not None
        assert len(cached) == 4
        assert cached["intraday_price_eur_mwh"].tolist() == [40.0, 42.0, 44.0, 46.0]


# ── Test 14b: Manual IDA CSV import ─────────────────────────────────────────

class TestParseIntradayCsv:
    def test_template_round_trips(self) -> None:
        df = parse_intraday_csv(generate_intraday_template_csv(), default_zone="DE_LU")
        assert list(df.columns) == ["zone", "sequence", "intraday_price_eur_mwh"]
        assert df.index.tz is not None and str(df.index.tz) == "UTC"
        assert df["zone"].unique().tolist() == ["DE_LU"]
        assert df["sequence"].unique().tolist() == [1]
        assert len(df) == 3

    def test_minimal_csv_uses_defaults_and_keeps_negative(self) -> None:
        content = (
            "timestamp,ida_price_eur_mwh\n"
            "2026-02-01 00:00:00,50\n"
            "2026-02-01 01:00:00,-5.5\n"
        )
        df = parse_intraday_csv(content, default_zone="NL", default_sequence=2)
        assert df["zone"].unique().tolist() == ["NL"]
        assert df["sequence"].unique().tolist() == [2]
        assert (df["intraday_price_eur_mwh"] < 0).any()  # negative prices retained

    def test_case_insensitive_headers_and_dedupe_keep_last(self) -> None:
        content = (
            "Timestamp,IDA_Price_EUR_MWh\n"
            "2026-02-01T00:00:00+00:00,10\n"
            "2026-02-01T00:00:00+00:00,99\n"
        )
        df = parse_intraday_csv(content, default_zone="FR")
        assert len(df) == 1
        assert df["intraday_price_eur_mwh"].iloc[0] == 99.0

    def test_per_row_zone_and_sequence_override_defaults(self) -> None:
        content = (
            "timestamp,ida_price_eur_mwh,sequence,zone\n"
            "2026-03-01T00:00:00+00:00,40,1,DE_LU\n"
            "2026-03-01T00:00:00+00:00,41,2,NL\n"
        )
        df = parse_intraday_csv(content, default_zone="DE_LU")
        assert set(zip(df["zone"], df["sequence"], strict=True)) == {
            ("DE_LU", 1), ("NL", 2),
        }

    def test_missing_required_columns_raises(self) -> None:
        with pytest.raises(ValueError, match="must have at least"):
            parse_intraday_csv("foo,bar\n1,2\n", default_zone="DE_LU")

    def test_missing_zone_without_default_raises(self) -> None:
        with pytest.raises(ValueError, match="without a zone"):
            parse_intraday_csv(
                "timestamp,ida_price_eur_mwh\n2026-01-01,5\n", default_zone=None,
            )

    def test_invalid_sequence_raises(self) -> None:
        with pytest.raises(ValueError, match="sequence"):
            parse_intraday_csv(
                "timestamp,ida_price_eur_mwh,sequence\n2026-01-01,5,7\n",
                default_zone="DE_LU",
            )

    def test_fractional_sequence_raises_not_truncated(self) -> None:
        # 1.5 must NOT be silently truncated to IDA1.
        with pytest.raises(ValueError, match="whole number"):
            parse_intraday_csv(
                "timestamp,ida_price_eur_mwh,sequence\n2026-01-01,5,1.5\n",
                default_zone="DE_LU",
            )

    def test_ida_label_sequence_is_parsed_not_dropped(self) -> None:
        # 'IDA2' must map to round 2, not be coerced to the default round.
        df = parse_intraday_csv(
            "timestamp,ida_price_eur_mwh,sequence,zone\n"
            "2026-01-01T00:00:00+00:00,5,IDA2,DE_LU\n"
            "2026-01-01T01:00:00+00:00,6,ida_3,NL\n",
            default_zone="DE_LU",
        )
        assert set(zip(df["zone"], df["sequence"], strict=True)) == {
            ("DE_LU", 2), ("NL", 3),
        }

    def test_non_numeric_sequence_raises_not_coerced(self) -> None:
        # Regression: 'IDA2'-style typos or garbage must not silently become
        # the default round (which would file the rows under the wrong table).
        with pytest.raises(ValueError, match="not understood"):
            parse_intraday_csv(
                "timestamp,ida_price_eur_mwh,sequence\n2026-01-01,5,bogus\n",
                default_zone="DE_LU",
            )

    @pytest.mark.filterwarnings("ignore:Could not infer format")
    def test_garbage_sequence_in_dropped_row_is_tolerated(self) -> None:
        # A bad sequence on a row that is dropped for a bad timestamp must not
        # fail the whole upload. ('notadate' triggers pandas per-element date
        # parsing, which is the intended robust-coerce behaviour here.)
        df = parse_intraday_csv(
            "timestamp,ida_price_eur_mwh,sequence\n"
            "notadate,5,bogus\n"
            "2026-01-01T00:00:00+00:00,6,1\n",
            default_zone="DE_LU",
        )
        assert len(df) == 1
        assert df["sequence"].iloc[0] == 1

    def test_unknown_zone_raises(self) -> None:
        with pytest.raises(ValueError, match="Unknown bidding zone"):
            parse_intraday_csv(
                "timestamp,ida_price_eur_mwh,zone\n2026-01-01,5,XX_FAKE\n",
                default_zone="DE_LU",
            )

    def test_persist_writes_to_same_cache_the_fetch_reads(
        self, tmp_path, monkeypatch,
    ) -> None:
        """Manual IDA rows must land in the same SQLite tables the live fetch
        reads, so the cockpit/revenue cache-first path picks them up."""
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        content = (
            "timestamp,ida_price_eur_mwh,sequence,zone\n"
            "2026-01-01T00:00:00+00:00,40,1,DE_LU\n"
            "2026-01-01T01:00:00+00:00,42,1,DE_LU\n"
            "2026-01-01T00:00:00+00:00,30,1,NL\n"
        )
        parsed = parse_intraday_csv(content, default_zone="DE_LU")
        summaries = persist_intraday_frame(parsed)
        assert {(s["zone"], s["sequence"]): s["rows"] for s in summaries} == {
            ("DE_LU", 1): 2, ("NL", 1): 1,
        }
        cached = read_intraday_cache(
            "DE_LU",
            pd.Timestamp("2026-01-01", tz="UTC"),
            pd.Timestamp("2026-01-02", tz="UTC"),
            sequence=1,
        )
        assert cached is not None
        assert cached["intraday_price_eur_mwh"].tolist() == [40.0, 42.0]

    def test_provenance_is_durable_and_relabelled_by_live_write(
        self, tmp_path, monkeypatch,
    ) -> None:
        """Manual provenance survives in SQLite (not just session state), and a
        later live write to the same (zone, sequence) relabels it."""
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")

        parsed = di.parse_intraday_csv(
            "timestamp,ida_price_eur_mwh,sequence,zone\n"
            "2026-01-01T00:00:00+00:00,40,1,DE_LU\n"
            "2026-01-01T01:00:00+00:00,42,1,DE_LU\n",
            default_zone="DE_LU",
        )
        di.persist_intraday_frame(parsed)

        sources = di.read_intraday_sources()
        assert sources[("DE_LU", 1)]["source"] == di.IDA_SOURCE_MANUAL
        assert sources[("DE_LU", 1)]["rows"] == 2
        assert sources[("DE_LU", 1)]["imported_at"]  # timestamp recorded

        # A live fetch writing the same table relabels provenance (no stale
        # "Manual CSV" after a real fetch arrives).
        live = pd.DataFrame(
            {"intraday_price_eur_mwh": [50.0]},
            index=pd.DatetimeIndex(
                [pd.Timestamp("2026-01-01T02:00:00+00:00")], name="timestamp",
            ),
        )
        di.write_intraday_cache(live, "DE_LU", 1)  # default source = ENTSO-E
        relabelled = di.read_intraday_sources()
        # Manual + live rows coexist in the table, so provenance is Mixed
        # (not the last writer's source).
        assert relabelled[("DE_LU", 1)]["source"] == di.IDA_SOURCE_MIXED
        assert relabelled[("DE_LU", 1)]["rows"] == 3  # union of both writes

    def test_legacy_table_without_source_column_is_migrated(
        self, tmp_path, monkeypatch,
    ) -> None:
        """A pre-provenance IDA table (no source column, e.g. the synthetic
        seed) is migrated and its legacy rows are labelled, not silently
        attributed to a real source."""
        from src import data_ingestion as di
        db = tmp_path / "legacy.db"
        monkeypatch.setattr(di, "DB_PATH", db)
        with sqlite3.connect(db) as conn:
            conn.execute(
                'CREATE TABLE "ida_prices_de_lu_seq1" '
                "(timestamp TEXT PRIMARY KEY, intraday_price_eur_mwh REAL NOT NULL)"
            )
            conn.execute(
                'INSERT INTO "ida_prices_de_lu_seq1" VALUES '
                '("2026-01-01T00:00:00+00:00", 45.0)'
            )
        # A manual upload onto the legacy table migrates it and marks Mixed
        # (legacy sentinel + Manual CSV).
        parsed = di.parse_intraday_csv(
            "timestamp,ida_price_eur_mwh,sequence,zone\n"
            "2026-01-01T01:00:00+00:00,40,1,DE_LU\n",
            default_zone="DE_LU",
        )
        di.persist_intraday_frame(parsed)
        assert di.read_intraday_sources()[("DE_LU", 1)]["source"] == di.IDA_SOURCE_MIXED
        # The legacy price row is still readable (migration preserved data).
        cached = di.read_intraday_cache(
            "DE_LU",
            pd.Timestamp("2026-01-01", tz="UTC"),
            pd.Timestamp("2026-01-02", tz="UTC"),
            sequence=1,
        )
        assert cached is not None and len(cached) == 2


# ── Test 15: REE ESIOS (Spain) ──────────────────────────────────────────────

class TestFetchEsiosIndicator:
    @patch("src.data_ingestion.requests.get")
    @patch("src.data_ingestion.get_esios_api_key", return_value="fake-key")
    def test_returns_correct_schema(
        self, _mock_key: MagicMock, mock_get: MagicMock,
    ) -> None:
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {
            "indicator": {
                "values": [
                    {"datetime": "2025-01-01T00:00:00+00:00", "value": 12.5, "geo_id": 8741},
                    {"datetime": "2025-01-01T01:00:00+00:00", "value": 13.0, "geo_id": 8741},
                ],
            },
        }
        mock_get.return_value = mock_resp

        df = fetch_esios_indicator(
            634,
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
            column="secondary_up",
        )
        assert list(df.columns) == ["timestamp", "secondary_up"]
        assert len(df) == 2
        assert df["secondary_up"].iloc[0] == pytest.approx(12.5)

    @patch("src.data_ingestion.requests.get")
    @patch("src.data_ingestion.get_esios_api_key", return_value="fake-key")
    def test_collapses_multi_geo_rows_by_mean(
        self, _mock_key: MagicMock, mock_get: MagicMock,
    ) -> None:
        """ESIOS sometimes returns one row per geo_id for the same timestamp;
        the fetcher collapses to one row per timestamp so downstream merges
        on timestamp do not blow up.
        """
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {
            "indicator": {
                "values": [
                    {"datetime": "2025-01-01T00:00:00+00:00", "value": 10.0, "geo_id": 1},
                    {"datetime": "2025-01-01T00:00:00+00:00", "value": 20.0, "geo_id": 2},
                ],
            },
        }
        mock_get.return_value = mock_resp
        df = fetch_esios_indicator(
            634,
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
            column="x",
        )
        assert len(df) == 1
        assert df["x"].iloc[0] == pytest.approx(15.0)  # mean of 10 and 20

    @patch("src.data_ingestion.get_esios_api_key", return_value="")
    def test_missing_key_raises_auth_error(self, _mock_key: MagicMock) -> None:
        with pytest.raises(DataSourceAuthError):
            fetch_esios_indicator(
                634,
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )

    @patch("src.data_ingestion.time.sleep")
    @patch("src.data_ingestion.requests.get")
    @patch("src.data_ingestion.get_esios_api_key", return_value="bad-key")
    def test_http_401_propagates_as_auth_error(
        self, _mock_key: MagicMock, mock_get: MagicMock, _mock_sleep: MagicMock,
    ) -> None:
        unauth = MagicMock()
        unauth.status_code = 401
        unauth.raise_for_status = MagicMock()
        mock_get.return_value = unauth
        with pytest.raises(DataSourceAuthError):
            fetch_esios_indicator(
                634,
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        # Bypassed retry: one GET, not multiple
        assert mock_get.call_count == 1

    @patch("src.data_ingestion.time.sleep")
    @patch("src.data_ingestion.get_esios_api_key", return_value="fake-key")
    def test_network_error_message_scrubbed_and_cause_suppressed(
        self,
        _mock_key: MagicMock,
        _mock_sleep: MagicMock,
    ) -> None:
        """PR5c: ESIOS uses a header API key so the URL itself does NOT
        carry credentials, but request exception strings still embed the
        upstream URL and any params. Streamlit's ``__cause__`` rendering
        would leak that URL chain. Match the ENTSO-E pattern: scrub +
        ``from None``.
        """
        with (
            patch(
                "src.data_ingestion.requests.get",
                side_effect=requests.ConnectionError(
                    "HTTPSConnectionPool: failed connecting to "
                    "https://api.esios.ree.es/indicators/634?"
                    "start_date=2025-01-01&end_date=2025-01-02&"
                    "api_key=SHOULD-BE-REDACTED"
                ),
            ),
            pytest.raises(DataSourceNetworkError) as exc_info,
        ):
            fetch_esios_indicator(
                634,
                pd.Timestamp("2025-01-01", tz="UTC"),
                pd.Timestamp("2025-01-02", tz="UTC"),
            )
        # ``from None`` suppresses the cause chain so __cause__ is None.
        assert exc_info.value.__cause__ is None
        # Even if a secret accidentally shows up in the upstream string
        # (api_key here), it must not appear in the surfaced message.
        assert "SHOULD-BE-REDACTED" not in str(exc_info.value)
        assert "api_key=***" in str(exc_info.value)

    @patch("src.data_ingestion.requests.get")
    @patch("src.data_ingestion.get_esios_api_key", return_value="fake-key")
    def test_bundle_merges_indicators_on_timestamp(
        self, _mock_key: MagicMock, mock_get: MagicMock,
    ) -> None:
        """fetch_esios_ancillary_prices should merge 4 indicator series into a
        single wide table keyed on timestamp.
        """
        def _make_payload(value: float) -> dict:
            return {
                "indicator": {
                    "values": [
                        {"datetime": "2025-01-01T00:00:00+00:00", "value": value, "geo_id": 1},
                        {"datetime": "2025-01-01T01:00:00+00:00", "value": value + 1, "geo_id": 1},
                    ],
                },
            }

        # Each indicator returns a slightly different value
        side_effects = [
            MagicMock(json=lambda v=v: _make_payload(v), status_code=200, raise_for_status=MagicMock())
            for v in (10.0, 11.0, 20.0, 21.0)
        ]
        mock_get.side_effect = side_effects
        df = fetch_esios_ancillary_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        expected_cols = {
            "timestamp",
            "secondary_up_capacity_eur_mw",
            "secondary_down_capacity_eur_mw",
            "tertiary_up_energy_eur_mwh",
            "tertiary_down_energy_eur_mwh",
        }
        assert expected_cols <= set(df.columns)
        assert len(df) == 2
        assert df["secondary_up_capacity_eur_mw"].iloc[0] == pytest.approx(10.0)

    @patch("src.data_ingestion.fetch_esios_indicator")
    def test_bundle_continues_when_one_indicator_fails(
        self, mock_fetch: MagicMock,
    ) -> None:
        """A partial failure should not erase successful indicators — the
        user is better off with N-1 series than zero.
        """
        good_df = pd.DataFrame({
            "timestamp": [pd.Timestamp("2025-01-01", tz="UTC")],
            "secondary_up_capacity_eur_mw": [12.0],
        })
        mock_fetch.side_effect = [
            good_df,
            DataSourceNetworkError("oops"),
            DataSourceNetworkError("oops"),
            DataSourceNetworkError("oops"),
        ]
        df = fetch_esios_ancillary_prices(
            pd.Timestamp("2025-01-01", tz="UTC"),
            pd.Timestamp("2025-01-02", tz="UTC"),
        )
        assert "secondary_up_capacity_eur_mw" in df.columns
        assert len(df) == 1


class TestCapacityCache:
    """Unified reserve-capacity persistence + provenance (Step 2 / 6b)."""

    @staticmethod
    def _frame(csv_text: str) -> pd.DataFrame:
        from src.ancillary import parse_capacity_import_csv
        return parse_capacity_import_csv(csv_text)

    def test_persist_and_read_back(self, tmp_path, monkeypatch) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        frame = self._frame(
            "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
            "2026-05-01T00:00:00Z,DE_LU,FCR,symmetric,12.5\n"
            "2026-05-01T00:00:00Z,DE_LU,aFRR,up,8.2\n"
        )
        summaries = di.persist_capacity_frame(frame)
        assert {(s["product"], s["direction"]) for s in summaries} == {
            ("FCR", "symmetric"), ("aFRR", "up"),
        }
        back = di.read_capacity_cache("DE_LU")
        assert len(back) == 2
        assert set(back["product_type"]) == {"FCR", "aFRR"}
        assert str(back.index.tz) == "UTC"

    def test_provenance_is_per_stream(self, tmp_path, monkeypatch) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_capacity_frame(
            self._frame(
                "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
                "2026-05-01T00:00:00Z,DE_LU,FCR,symmetric,12.5\n"
            ),
            source="Manual CSV",
        )
        di.persist_capacity_frame(
            self._frame(
                "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
                "2026-05-01T00:00:00Z,DE_LU,aFRR,up,8.2\n"
            ),
            source="TSO API",
        )
        sources = di.read_capacity_sources()
        # Independent label per (zone, product, direction).
        assert sources[("DE_LU", "FCR", "symmetric")]["source"] == "Manual CSV"
        assert sources[("DE_LU", "aFRR", "up")]["source"] == "TSO API"

    def test_keep_last_overwrite_refreshes_price_and_provenance(
        self, tmp_path, monkeypatch,
    ) -> None:
        # Red-line: re-importing the same (timestamp, product, direction) must
        # overwrite the value AND refresh provenance — no stale source on fresh
        # data.
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_capacity_frame(
            self._frame(
                "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
                "2026-05-01T00:00:00Z,DE_LU,FCR,symmetric,10.0\n"
            ),
            source="Manual CSV",
        )
        di.persist_capacity_frame(
            self._frame(
                "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
                "2026-05-01T00:00:00Z,DE_LU,FCR,symmetric,20.0\n"
            ),
            source="TSO API",
        )
        back = di.read_capacity_cache("DE_LU")
        assert len(back) == 1  # keep-last, not duplicated
        assert back["capacity_price_eur_mw"].iloc[0] == 20.0  # new value
        assert (
            di.read_capacity_sources()[("DE_LU", "FCR", "symmetric")]["source"]
            == "TSO API"  # provenance followed the overwrite
        )

    def test_mixed_sources_on_one_stream_labelled_mixed(
        self, tmp_path, monkeypatch,
    ) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_capacity_frame(
            self._frame(
                "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
                "2026-05-01T00:00:00Z,DE_LU,FCR,symmetric,10.0\n"
            ),
            source="Manual CSV",
        )
        di.persist_capacity_frame(  # different timestamp, same stream, new source
            self._frame(
                "timestamp,zone,product,direction,capacity_price_eur_mw_h\n"
                "2026-05-01T04:00:00Z,DE_LU,FCR,symmetric,11.0\n"
            ),
            source="TSO API",
        )
        sources = di.read_capacity_sources()
        assert sources[("DE_LU", "FCR", "symmetric")]["source"] == di.CAPACITY_SOURCE_MIXED
        assert sources[("DE_LU", "FCR", "symmetric")]["rows"] == 2

    def test_read_empty_returns_none(self, tmp_path, monkeypatch) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        assert di.read_capacity_cache("DE_LU") is None
        assert di.read_capacity_sources() == {}

    def test_nan_direction_persists_as_blank_not_nan_string(
        self, tmp_path, monkeypatch,
    ) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        frame = pd.DataFrame(
            {
                "zone": ["DE_LU"],
                "product_type": ["FCR"],
                "direction": [float("nan")],
                "capacity_price_eur_mw": [12.5],
            },
            index=pd.DatetimeIndex(["2026-05-01T00:00:00Z"], name="timestamp"),
        )
        di.persist_capacity_frame(frame, source="Manual CSV")
        back = di.read_capacity_cache("DE_LU")
        assert back["direction"].iloc[0] == ""
        assert ("DE_LU", "FCR", "") in di.read_capacity_sources()


class TestActivationCache:
    """Unified activation-energy persistence + provenance (Step 3 / 3b)."""

    _HDR = (
        "timestamp,zone,product,direction,activation_price_eur_mwh,"
        "system_activated_volume_mw\n"
    )

    @classmethod
    def _frame(cls, body: str) -> pd.DataFrame:
        from src.ancillary import parse_activation_import_csv
        return parse_activation_import_csv(cls._HDR + body)

    def test_persist_and_read_back(self, tmp_path, monkeypatch) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        frame = self._frame(
            "2026-05-01T00:00:00Z,DE_LU,aFRR,up,85.4,320\n"
            "2026-05-01T00:00:00Z,DE_LU,mFRR,down,5.0,90\n"
        )
        summaries = di.persist_activation_frame(frame)
        assert {(s["product"], s["direction"]) for s in summaries} == {
            ("aFRR", "up"), ("mFRR", "down"),
        }
        back = di.read_activation_cache("DE_LU")
        assert len(back) == 2
        assert set(back["product_type"]) == {"aFRR", "mFRR"}
        assert str(back.index.tz) == "UTC"

    def test_system_volume_persists_and_reads_back(self, tmp_path, monkeypatch) -> None:
        # The energy leg's quantity dimension survives the round-trip untouched.
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_activation_frame(
            self._frame("2026-05-01T00:00:00Z,DE_LU,aFRR,up,85.0,1234.5\n")
        )
        back = di.read_activation_cache("DE_LU")
        assert back["system_activated_volume_mw"].iloc[0] == 1234.5
        assert back["activation_price_eur_mwh"].iloc[0] == 85.0

    def test_provenance_is_per_stream(self, tmp_path, monkeypatch) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_activation_frame(
            self._frame("2026-05-01T00:00:00Z,DE_LU,aFRR,up,85.0,100\n"),
            source="Manual CSV",
        )
        di.persist_activation_frame(
            self._frame("2026-05-01T00:00:00Z,DE_LU,mFRR,down,5.0,50\n"),
            source="TSO API",
        )
        sources = di.read_activation_sources()
        assert sources[("DE_LU", "aFRR", "up")]["source"] == "Manual CSV"
        assert sources[("DE_LU", "mFRR", "down")]["source"] == "TSO API"

    def test_keep_last_overwrite_refreshes_value_and_provenance(
        self, tmp_path, monkeypatch,
    ) -> None:
        # Red-line: re-importing the same (timestamp, product, direction) must
        # overwrite price AND volume AND refresh provenance.
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_activation_frame(
            self._frame("2026-05-01T00:00:00Z,DE_LU,aFRR,up,80.0,100\n"),
            source="Manual CSV",
        )
        di.persist_activation_frame(
            self._frame("2026-05-01T00:00:00Z,DE_LU,aFRR,up,95.0,250\n"),
            source="TSO API",
        )
        back = di.read_activation_cache("DE_LU")
        assert len(back) == 1  # keep-last, not duplicated
        assert back["activation_price_eur_mwh"].iloc[0] == 95.0
        assert back["system_activated_volume_mw"].iloc[0] == 250.0
        assert (
            di.read_activation_sources()[("DE_LU", "aFRR", "up")]["source"]
            == "TSO API"
        )

    def test_mixed_sources_on_one_stream_labelled_mixed(
        self, tmp_path, monkeypatch,
    ) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        di.persist_activation_frame(
            self._frame("2026-05-01T00:00:00Z,DE_LU,aFRR,up,80.0,100\n"),
            source="Manual CSV",
        )
        di.persist_activation_frame(  # different timestamp, same stream, new source
            self._frame("2026-05-01T04:00:00Z,DE_LU,aFRR,up,82.0,110\n"),
            source="TSO API",
        )
        sources = di.read_activation_sources()
        assert sources[("DE_LU", "aFRR", "up")]["source"] == di.ACTIVATION_SOURCE_MIXED
        assert sources[("DE_LU", "aFRR", "up")]["rows"] == 2

    def test_read_empty_returns_none(self, tmp_path, monkeypatch) -> None:
        from src import data_ingestion as di
        monkeypatch.setattr(di, "DB_PATH", tmp_path / "bess.db")
        assert di.read_activation_cache("DE_LU") is None
        assert di.read_activation_sources() == {}
