"""Contract pins for external annual trader-benchmark reconciliation."""

from __future__ import annotations

import math
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.export import cockpit_tables_to_excel
from src.pages.forward_scenarios import (
    _BENCHMARK_HARD_CAPTION,
    _benchmark_assumptions,
    _benchmark_display_frame,
    _platform_basis_caption,
)
from src.trader_benchmark import (
    MODEL_REVENUE_COLUMN,
    QUOTE_REVENUE_COLUMN,
    benchmark_comparability_notes,
    build_forward_model_yearly,
    generate_trader_benchmark_template_csv,
    parse_trader_benchmark_csv,
    reconcile_trader_benchmark,
)


def _csv(*rows: str) -> str:
    header = (
        "zone,scenario,year,revenue_eur_per_mw_yr,asset_type,market_scope,"
        "revenue_basis,duration_hours,max_efc_per_day,source,as_of"
    )
    return "\n".join([header, *rows]) + "\n"


def _benchmark() -> pd.DataFrame:
    return parse_trader_benchmark_csv(
        _csv(
            "IT_SICI,Base,2027,100000,standalone,da-only,gross,6.7,,Desk,2026-01-01",
            "IT_SICI,Base,2028,80000,standalone,da-only,gross,6.7,,Desk,2026-01-01",
        )
    )


class TestParseTraderBenchmark:
    def test_template_is_parseable_and_audited(self) -> None:
        parsed = parse_trader_benchmark_csv(
            generate_trader_benchmark_template_csv()
        )
        assert len(parsed) == 2
        assert parsed["asset_type"].unique().tolist() == ["standalone"]
        assert parsed["market_scope"].unique().tolist() == ["all-in"]
        assert parsed["revenue_basis"].unique().tolist() == ["unknown"]

    @pytest.mark.parametrize(
        ("column", "value"),
        [
            ("asset_type", "hybrid-ish"),
            ("market_scope", "everything"),
            ("revenue_basis", "ebitda"),
        ],
    )
    def test_controlled_basis_fields_reject_unknown_labels(
        self, column: str, value: str,
    ) -> None:
        row = {
            "asset_type": "standalone",
            "market_scope": "da-only",
            "revenue_basis": "gross",
        }
        row[column] = value
        content = _csv(
            f"DE_LU,Base,2027,100,{row['asset_type']},{row['market_scope']},"
            f"{row['revenue_basis']},2,,,"
        )
        with pytest.raises(ValueError, match=column):
            parse_trader_benchmark_csv(content)

    def test_signed_revenue_is_preserved(self) -> None:
        parsed = parse_trader_benchmark_csv(
            _csv("DE_LU,Downside,2027,-5000,standalone,da-only,gross,,,,")
        )
        assert parsed.loc[0, "revenue_eur_per_mw_yr"] == -5000.0

    @pytest.mark.parametrize(
        "year",
        ["2027.5", "not-a-year", "1999", "2201"],
    )
    def test_year_must_be_an_integer_in_domain(self, year: str) -> None:
        with pytest.raises(ValueError, match="year"):
            parse_trader_benchmark_csv(
                _csv(
                    f"DE_LU,Base,{year},100,standalone,da-only,gross,,,,"
                )
            )

    def test_duplicate_scenario_year_rejected(self) -> None:
        content = _csv(
            "DE_LU,Base,2027,100,standalone,da-only,gross,,,,",
            "DE_LU,Base,2027,110,standalone,da-only,gross,,,,",
        )
        with pytest.raises(ValueError, match="duplicate"):
            parse_trader_benchmark_csv(content)

    def test_metadata_cannot_drift_inside_one_scenario(self) -> None:
        content = _csv(
            "DE_LU,Base,2027,100,standalone,da-only,gross,2,,,",
            "DE_LU,Base,2028,90,standalone,all-in,gross,2,,,",
        )
        with pytest.raises(ValueError, match=r"constant.*market_scope"):
            parse_trader_benchmark_csv(content)

    @pytest.mark.parametrize(
        ("duration", "cycles"),
        [("0", "1"), ("-1", "1"), ("2", "0"), ("bad", "1")],
    )
    def test_optional_physical_metadata_must_be_positive_when_present(
        self, duration: str, cycles: str,
    ) -> None:
        with pytest.raises(ValueError, match="positive finite"):
            parse_trader_benchmark_csv(
                _csv(
                    "DE_LU,Base,2027,100,standalone,da-only,gross,"
                    f"{duration},{cycles},,"
                )
            )

    def test_invalid_as_of_rejected(self) -> None:
        with pytest.raises(ValueError, match="as_of"):
            parse_trader_benchmark_csv(
                _csv(
                    "DE_LU,Base,2027,100,standalone,da-only,gross,,,Desk,nope"
                )
            )

    @pytest.mark.parametrize("content", ["", "# comments only\n# no CSV\n"])
    def test_empty_or_comment_only_file_has_friendly_error(
        self, content: str,
    ) -> None:
        with pytest.raises(ValueError, match="no header or data rows"):
            parse_trader_benchmark_csv(content)


class TestForwardModelYearly:
    def test_uses_existing_yearly_revenue_convention(self) -> None:
        daily = pd.DataFrame(
            {
                "date": [pd.Timestamp("2027-01-01"), pd.Timestamp("2027-01-02")],
                "spread": [10.0, 20.0],
                "lp_revenue": [100.0, 300.0],
                "n_cycles": [1.0, 2.0],
            }
        )
        result = build_forward_model_yearly(
            {"DE_LU": daily},
            power_mw=2.0,
            duration_hours=4.0,
            efficiency=0.88,
            capture_rate=0.5,
        )
        row = result.iloc[0]
        # Mean LP revenue 200/day x 50% capture x 365.25 / 2 MW.
        assert row[MODEL_REVENUE_COLUMN] == pytest.approx(18262.5)
        assert row["avg_cycles_per_day"] == 1.5
        assert row["n_days"] == 2
        assert bool(row["is_partial_year"])

    def test_empty_mapping_returns_typed_empty(self) -> None:
        result = build_forward_model_yearly(
            {},
            power_mw=1.0,
            duration_hours=2.0,
            efficiency=0.88,
            capture_rate=0.7,
        )
        assert result.empty
        assert MODEL_REVENUE_COLUMN in result.columns


class TestReconcileTraderBenchmark:
    def test_overlap_math_and_endpoint_cagr(self) -> None:
        model = pd.DataFrame(
            {
                "zone": ["IT_SICI", "IT_SICI"],
                "year": [2027, 2028],
                "n_days": [365, 366],
                "coverage_pct": [100.0, 100.0],
                "is_partial_year": [False, False],
                "avg_cycles_per_day": [1.0, 1.0],
                MODEL_REVENUE_COLUMN: [125000.0, 100000.0],
            }
        )
        comparison, summary = reconcile_trader_benchmark(
            _benchmark(), model, zone="IT_SICI", scenario="Base"
        )
        assert comparison[QUOTE_REVENUE_COLUMN].tolist() == [100000.0, 80000.0]
        assert comparison["model_minus_benchmark_eur_per_mw_yr"].tolist() == [
            25000.0,
            20000.0,
        ]
        assert comparison["benchmark_to_model_ratio"].tolist() == [0.8, 0.8]
        assert summary["n_overlap_years"] == 2
        assert summary["benchmark_to_model_ratio"] == 0.8
        assert summary["benchmark_endpoint_cagr"] == pytest.approx(-0.2)

    def test_non_overlapping_years_remain_visible_without_fake_ratio(self) -> None:
        model = pd.DataFrame(
            {
                "zone": ["IT_SICI"],
                "year": [2030],
                MODEL_REVENUE_COLUMN: [90000.0],
                "coverage_pct": [100.0],
                "is_partial_year": [False],
                "avg_cycles_per_day": [1.0],
            }
        )
        comparison, summary = reconcile_trader_benchmark(
            _benchmark(), model, zone="IT_SICI", scenario="Base"
        )
        assert len(comparison) == 2
        assert comparison[MODEL_REVENUE_COLUMN].isna().all()
        assert summary["n_overlap_years"] == 0
        assert math.isnan(float(summary["benchmark_to_model_ratio"]))

    def test_nonpositive_model_never_creates_capture_like_ratio(self) -> None:
        model = pd.DataFrame(
            {
                "zone": ["IT_SICI"],
                "year": [2027],
                MODEL_REVENUE_COLUMN: [0.0],
                "coverage_pct": [100.0],
                "is_partial_year": [False],
                "avg_cycles_per_day": [0.0],
            }
        )
        comparison, summary = reconcile_trader_benchmark(
            _benchmark(), model, zone="IT_SICI", scenario="Base"
        )
        assert math.isnan(float(comparison.loc[0, "benchmark_to_model_ratio"]))
        assert math.isnan(float(summary["benchmark_to_model_ratio"]))


class TestComparabilityNotes:
    def test_matching_standalone_da_gross_case_has_no_metadata_warning(self) -> None:
        notes = benchmark_comparability_notes(
            _benchmark(), zone="IT_SICI", scenario="Base", model_duration_hours=6.7
        )
        assert notes == []

    def test_colocation_scope_basis_duration_and_cycle_cap_are_all_disclosed(
        self,
    ) -> None:
        benchmark = parse_trader_benchmark_csv(
            _csv(
                "IT_SICI,Hybrid,2027,100000,co-located,all-in,net-of-fees,"
                "4,2,Desk,2026-01-01"
            )
        )
        notes = benchmark_comparability_notes(
            benchmark,
            zone="IT_SICI",
            scenario="Hybrid",
            model_duration_hours=6.7,
        )
        assert len(notes) == 5
        assert any("co-location" in note for note in notes)
        assert any("all-in" in note for note in notes)
        assert any("net-of-fees" in note for note in notes)
        assert any("4h" in note and "6.7h" in note for note in notes)
        assert any("cycle cap" in note for note in notes)


class TestBenchmarkPresentationContract:
    def test_locked_caption_is_verbatim(self) -> None:
        assert _BENCHMARK_HARD_CAPTION == (
            "External benchmark reconciliation only: the uploaded annual "
            "revenue curve is user-supplied and is never used as a price "
            "curve, contracted floor, capture rate, solver input, or "
            "bankable forecast."
        )

    def test_platform_basis_caption_discloses_live_capture_haircut(self) -> None:
        caption = _platform_basis_caption(0.7)
        assert caption == (
            "Platform curve basis: forward-synthetic DA-only dispatch revenue "
            "x 70.0% sidebar capture haircut; before wear, fees, tax, and "
            "financing. The benchmark/model ratio embeds that haircut."
        )
        assert "85.0%" in _platform_basis_caption(0.85)

    def test_display_ratio_is_percentage_points_without_mutating_export(self) -> None:
        raw = pd.DataFrame(
            {
                "year": [2027],
                "benchmark_to_model_ratio": [0.8],
                "coverage_pct": [100.0],
            }
        )
        display = _benchmark_display_frame(raw)
        assert display.loc[0, "benchmark_to_model_ratio"] == 80.0
        assert display.loc[0, "coverage_pct"] == 100.0
        assert raw.loc[0, "benchmark_to_model_ratio"] == 0.8

    def test_export_assumptions_keep_quote_and_model_bases_separate(self) -> None:
        selected = _benchmark()
        assumptions = _benchmark_assumptions(
            selected,
            power_mw=45.0,
            duration_hours=6.7,
            efficiency=0.88,
            capture_rate=0.7,
        ).set_index("parameter")
        assert assumptions.loc["External benchmark source", "value"] == "Desk"
        assert assumptions.loc["External benchmark as-of", "value"] == "2026-01-01"
        assert "DA-only" in assumptions.loc["Platform comparison basis", "value"]
        assert assumptions.loc["Platform BESS case", "value"].startswith(
            "45 MW / 6.7h"
        )
        ratio_semantics = assumptions.loc[
            "Benchmark/model ratio semantics", "value"
        ]
        assert "NOT a capture-rate estimate" in ratio_semantics

    def test_excel_roundtrip_keeps_raw_ratio_and_provenance(self) -> None:
        selected = _benchmark()
        model = pd.DataFrame(
            {
                "zone": ["IT_SICI", "IT_SICI"],
                "year": [2027, 2028],
                MODEL_REVENUE_COLUMN: [125000.0, 100000.0],
                "coverage_pct": [100.0, 100.0],
                "is_partial_year": [False, False],
                "avg_cycles_per_day": [1.2, 1.1],
            }
        )
        comparison, _ = reconcile_trader_benchmark(
            selected,
            model,
            zone="IT_SICI",
            scenario="Base",
        )
        assumptions = _benchmark_assumptions(
            selected,
            power_mw=45.0,
            duration_hours=6.7,
            efficiency=0.88,
            capture_rate=0.7,
        )
        data = cockpit_tables_to_excel(
            {
                "External benchmark": selected,
                "Benchmark reconciliation": comparison,
                "Platform annual curve": model,
            },
            assumptions=assumptions,
        )

        workbook = load_workbook(BytesIO(data), data_only=True)
        assert workbook.sheetnames == [
            "External benchmark",
            "Benchmark reconciliation",
            "Platform annual curve",
            "Assumptions",
        ]
        reconciliation = workbook["Benchmark reconciliation"]
        headers = [cell.value for cell in reconciliation[1]]
        ratio_col = headers.index("benchmark_to_model_ratio") + 1
        assert reconciliation.cell(row=2, column=ratio_col).value == 0.8

        audit = workbook["Assumptions"]
        audit_rows = {
            audit.cell(row=row, column=1).value: audit.cell(row=row, column=2).value
            for row in range(2, audit.max_row + 1)
        }
        assert audit_rows["External benchmark source"] == "Desk"
        assert "NOT a capture-rate estimate" in audit_rows[
            "Benchmark/model ratio semantics"
        ]
