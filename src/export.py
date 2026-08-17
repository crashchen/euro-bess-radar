"""Excel and PDF report export for eu-bess-pulse."""

from __future__ import annotations

import base64
import datetime as _dt
import logging
import math
import tempfile
from collections.abc import Mapping
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics import (
    build_price_heatmap,
    calculate_negative_price_hours,
    filter_to_complete_local_days,
)
from src.config import CACHE_DIR
from src.data_ingestion import summarize_price_data_quality
from src.project_case import (
    BOOTSTRAP_ALGORITHM_V1,
    CASHFLOW_RECONCILIATION_ABS_TOL_EUR_V1,
    CASHFLOW_RECONCILIATION_REL_TOL_V1,
    CASHFLOW_RECONCILIATION_VERSION_V1,
    CONTRACT_ASSET_SCOPE_V1,
    CONTRACT_CASHFLOW_TABLE_STATISTIC_V1,
    CONTRACT_QUOTE_BASIS_V1,
    CONTRACT_SETTLEMENT_ALGORITHM_V1,
    CONTRACT_SETTLEMENT_FREQUENCY_V1,
    EXPECTED_GRID_REGISTRY_VERSION,
    NULL_CASHFLOW_TABLE_STATISTIC_V1,
    PC_A_CALCULATOR_VERSION,
    PC_D2_CALCULATOR_VERSION,
    PROJECT_CASE_SCHEMA_VERSION,
    CapacityMaintenanceBasis,
    CashflowRow,
    CashflowTable,
    ContractCurrencyBasisMode,
    ContractQuoteStatus,
    ContractSettlementBasis,
    NpvDistribution,
    NpvOutcome,
    ProjectionKind,
    RunResult,
    bootstrap_annual_sums,
    fingerprint_hex,
)
from src.project_case.enums import (
    CONTRACT_PRODUCT_DISCLOSURE_V1,
    LIFECYCLE_UNKNOWN_MESSAGE,
    LIFECYCLE_UNKNOWN_STATUS,
    MAX_PROJECT_LIFE_YEARS,
    MAX_SIMULATIONS,
    MIN_SIMULATIONS,
)
from src.project_case.schema import (
    _replay_typed_project_case_payload,
    _replay_typed_strategy_payload,
    _validate_strategy_wire_payload,
)

logger = logging.getLogger(__name__)


# ── Styling helpers ──────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_PRICE_FMT = "#,##0.00"
_PCT_FMT = "0.0%"
_PROJECT_CASE_NPV_SHEET = "Project Case NPVs"
_SCREENING_CASHFLOW_SHEET = "Screening Cash Flow"
_LIFECYCLE_CASHFLOW_SHEET = "Lifecycle Cash Flow"
_CONTRACT_SETTLEMENT_SHEET = "Contract Settlement"
_PROJECT_CASE_ASSUMPTIONS_SHEET = "Assumptions & Provenance"
_CONTRACT_SETTLEMENT_BASIS = (
    ContractSettlementBasis.ANNUAL_PRE_LIFECYCLE_STRATEGY_CASH_FLOOR_V1.value
)
_CONTRACT_CURRENCY_BASIS = (
    ContractCurrencyBasisMode.USER_ASSERTED_REAL_BASE_YEAR_EUR_CURVE.value
)
_CONTRACT_QUOTE_STATUSES = frozenset(status.value for status in ContractQuoteStatus)
# One locked literal shared with the Streamlit disclosure surface (§8).
_CONTRACT_DISCLOSURE = CONTRACT_PRODUCT_DISCLOSURE_V1
_PROJECT_CASE_STATIC_RED_LINES = {
    "cash_npv_includes_shadow_wear": False,
    "vom_rededucted": False,
    "mw_rescaled": False,
    "wear_net_floor_comparator_included": False,
    "pre_tax_unlevered": True,
    "tax_included": False,
    "debt_included": False,
    "financing_fees_included": False,
}
_PROJECT_CASE_DYNAMIC_RED_LINES = frozenset(
    {"contract_settlement_included", "contract_settlement_basis"}
)
_PROJECT_CASE_PROVENANCE_KEYS = frozenset(
    {
        "calculator_version",
        "project_case_input_fingerprint",
        "strategy_run_fingerprint",
        "project_case",
        "strategy_run_result",
        "bootstrap",
        "projection",
        "valuation",
        "capacity_maintenance_basis",
        "contract_settlement",
        "cashflow_table_statistic",
        "cashflow_reconciliation",
        "red_line_assertions",
    }
)
_PROJECT_CASE_PAYLOAD_KEYS = frozenset(
    {
        "asset_case",
        "lifecycle_case",
        "market_case",
        "valuation_case",
        "bootstrap_case",
        "contract_case",
    }
)
_MARKET_CASE_PAYLOAD_KEYS = frozenset(
    {"strategy_run_fingerprint", "projection"}
)
_PROJECTION_PAYLOAD_KEYS = frozenset(
    {
        "projection_kind",
        "annual_decay_rate",
        "decay_floor_share",
        "multipliers",
        "source",
        "as_of",
    }
)
_BOOTSTRAP_CASE_KEYS = frozenset(
    {"seed", "n_simulations", "bootstrap_algorithm_version"}
)
_BOOTSTRAP_PROVENANCE_KEYS = frozenset(
    {
        "algorithm_version",
        "seed",
        "n_simulations",
        "days_per_annual_draw",
        "percentile_method",
        "prob_positive_rule",
    }
)
_PROJECTION_PROVENANCE_KEYS = frozenset(
    {"projection_kind", "resolved_annual_multipliers"}
)
_ASSET_CASE_KEYS = frozenset(
    {
        "power_mw",
        "duration_hours",
        "round_trip_efficiency",
        "installed_capex_eur",
        "fixed_om_eur_per_mw_yr",
    }
)
_LIFECYCLE_CASE_KEYS = frozenset(
    {
        "project_life_years",
        "capacity_maintenance_basis",
        "capacity_maintenance_source",
        "capacity_maintenance_as_of",
        "augmentation_events",
        "eol_residual_value_eur",
        "decommissioning_cost_eur",
    }
)
_AUGMENTATION_EVENT_KEYS = frozenset(
    {"year", "cost_eur", "capacity_restored_frac", "residual_value_eur"}
)
_VALUATION_CASE_KEYS = frozenset({"discount_rate", "base_year"})
_VALUATION_PROVENANCE_KEYS = frozenset(
    {"discount_rate", "base_year", "currency_convention", "cash_timing"}
)
_STRATEGY_RUN_RESULT_PAYLOAD_KEYS = frozenset(
    {
        "strategy_kind",
        "daily_realised_cash_series",
        "cash_basis",
        "power_mw",
        "duration_hours",
        "round_trip_efficiency",
        "zone",
        "sample_window",
        "currency_basis",
        "forecast_audits",
        "reserve_product",
        "reserve_source",
        "availability",
        "reserve_coverage_audit",
        "coverage_audit",
        "adapter_provenance",
        "embedded_vom_cost_eur_mwh",
        "source_data_content_hash",
        "calculator_version",
    }
)
_CONTRACT_CASE_KEYS = frozenset({"settlement_basis", "settlement_terms"})
_CONTRACT_TERMS_KEYS = frozenset(
    {
        "contract_start_project_year",
        "floor_rate_real_eur_per_modeled_mw_year_by_contract_year",
        "floor_entitlement_factor_by_contract_year",
        "quote_basis",
        "settlement_frequency",
        "asset_scope",
        "currency_basis",
        "quote_status",
        "source",
        "source_as_of_date",
        "source_document_sha256",
    }
)
_CONTRACT_CURRENCY_KEYS = frozenset({"mode", "target_base_year"})
_CONTRACT_SETTLEMENT_KEYS = frozenset(
    {
        "basis",
        "algorithm_version",
        "resolved_floor_by_project_year",
        "representative_interpolation",
    }
)
_RESOLVED_FLOOR_KEYS = frozenset({"year", "effective_floor_eur"})
_INTERPOLATION_KEYS = frozenset(
    {
        "lower_sorted_rank",
        "upper_sorted_rank",
        "lower_original_draw_index",
        "upper_original_draw_index",
        "interpolation_weight",
    }
)
_CASHFLOW_RECONCILIATION_KEYS = frozenset(
    {"version", "relative_tolerance", "absolute_tolerance_eur"}
)
_EXCEL_CELL_TEXT_LIMIT = 32_767
_AUDIT_TEXT_PLAIN = "plain-utf8"
_AUDIT_TEXT_BASE64 = "base64-utf8-surrogatepass"

# Strings whose first non-whitespace character is one of these are
# interpreted by Excel / LibreOffice Calc / Google Sheets as a formula. A
# malicious uploader can smuggle ``=HYPERLINK("//evil/...", "Click")`` or
# ``=cmd|'/C calc'!A1`` through any free-text field (e.g. ancillary CSV
# ``product``, forward CSV ``contract``) and have it execute when the
# recipient opens the export. Prefix a single quote so the cell is forced
# to text mode. Gemini-3.1 P0 flagged that checking ``value[0]`` alone
# misses payloads with a leading space (Calc still evaluates them).
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n", "|")


def _safe_cell_value(value):
    """Return ``value`` with formula triggers neutralised for spreadsheet writes."""
    if isinstance(value, str) and value:
        stripped = value.lstrip()
        if stripped and stripped[0] in _FORMULA_TRIGGERS:
            return "'" + value
    return value


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=_safe_cell_value(header))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_column_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _write_kv_pair(ws, row: int, key: str, value, fmt: str | None = None) -> int:
    """Write a key-value pair to the summary sheet."""
    ws.cell(row=row, column=1, value=_safe_cell_value(key)).font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=_safe_cell_value(value))
    if fmt:
        cell.number_format = fmt
    return row + 1


# ── Sheet builders ───────────────────────────────────────────────────────────

def _build_summary_sheet(
    ws, zone: str, price_df: pd.DataFrame,
    percentiles: dict[str, float],
    revenue_estimate: dict[str, float],
    negative_stats: dict[str, float],
    tz: str | None = None,
) -> None:
    """Populate the Summary sheet with key-value metrics."""
    ws.title = "Summary"
    row = 1
    ws.cell(row=row, column=1, value="BESS Pulse — Market Report").font = Font(
        bold=True, size=14
    )
    row = 3

    dates = price_df.index.tz_convert(tz) if tz else price_df.index
    row = _write_kv_pair(ws, row, "Zone", zone)
    row = _write_kv_pair(ws, row, "Timezone", tz or "UTC")
    row = _write_kv_pair(ws, row, "Date Range Start", str(dates.min().date()))
    row = _write_kv_pair(ws, row, "Date Range End", str(dates.max().date()))
    total_days = (dates.max().date() - dates.min().date()).days + 1
    row = _write_kv_pair(ws, row, "Total Days", total_days)
    quality = summarize_price_data_quality(price_df)
    if quality["source_gap_intervals"] > 0:
        row = _write_kv_pair(
            ws, row, "Source Gap Intervals", quality["source_gap_intervals"],
        )
        row = _write_kv_pair(
            ws, row, "Short-Gap Imputed Intervals", quality["imputed_intervals"],
        )
        row = _write_kv_pair(
            ws, row, "Unresolved Missing Intervals", quality["missing_intervals"],
        )
        row = _write_kv_pair(
            ws, row, "Unresolved Missing %", quality["missing_ratio"], _PCT_FMT,
        )
        row = _write_kv_pair(
            ws, row, "Max Source Gap (hours)", quality["max_source_gap_hours"],
        )
    row += 1

    row = _write_kv_pair(ws, row, "Avg Price (EUR/MWh)", round(price_df["price_eur_mwh"].mean(), 2), _PRICE_FMT)
    row = _write_kv_pair(ws, row, "Median Price (EUR/MWh)", round(price_df["price_eur_mwh"].median(), 2), _PRICE_FMT)
    row += 1

    row = _write_kv_pair(ws, row, "50th-percentile Spread", round(percentiles["p50"], 2), _PRICE_FMT)
    row = _write_kv_pair(ws, row, "75th-percentile Spread", round(percentiles["p75"], 2), _PRICE_FMT)
    row = _write_kv_pair(ws, row, "90th-percentile Spread", round(percentiles["p90"], 2), _PRICE_FMT)
    row = _write_kv_pair(ws, row, "Mean Spread", round(percentiles["mean"], 2), _PRICE_FMT)
    row += 1

    row = _write_kv_pair(ws, row, "Est. Annual Revenue (EUR/MW)", revenue_estimate["annual_revenue_eur_per_mw"], _PRICE_FMT)
    if "total_eur" in revenue_estimate:
        row = _write_kv_pair(ws, row, "Headline Annual Revenue (EUR)", revenue_estimate["total_eur"], _PRICE_FMT)
    if "headline_total_mode" in revenue_estimate:
        row = _write_kv_pair(ws, row, "Headline Total Mode", revenue_estimate["headline_total_mode"])
    if "gross_additive_total_eur" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Gross Additive Total (Reference, EUR)",
            revenue_estimate["gross_additive_total_eur"],
            _PRICE_FMT,
        )
    if revenue_estimate.get("capacity_stack_warning"):
        row = _write_kv_pair(
            ws, row,
            "Capacity Stack Warning",
            revenue_estimate["capacity_stack_warning"],
        )
    if "joint_cooptimized_total_eur" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Joint MILP Co-optimized Total (EUR)",
            revenue_estimate["joint_cooptimized_total_eur"],
            _PRICE_FMT,
        )
        row = _write_kv_pair(
            ws, row,
            "Joint MILP Avg Reserve Commitment",
            revenue_estimate.get("joint_cooptimized_avg_reserve_fraction", 0.0),
            _PCT_FMT,
        )
    if "source_revenues" in revenue_estimate:
        for source, value in revenue_estimate["source_revenues"].items():
            row = _write_kv_pair(ws, row, f"{source} Revenue (EUR)", value, _PRICE_FMT)
    if "power_mw" in revenue_estimate:
        row = _write_kv_pair(ws, row, "BESS Power (MW)", revenue_estimate["power_mw"])
    if "duration_hours" in revenue_estimate:
        row = _write_kv_pair(ws, row, "BESS Duration (h)", revenue_estimate["duration_hours"])
    if "roundtrip_efficiency" in revenue_estimate:
        row = _write_kv_pair(ws, row, "Round-Trip Efficiency", revenue_estimate["roundtrip_efficiency"], _PCT_FMT)
    row = _write_kv_pair(ws, row, "Modeled Cycles per Day", revenue_estimate["cycles_per_day_assumption"])
    row = _write_kv_pair(ws, row, "Capture Rate Assumption", revenue_estimate["capture_rate_assumption"])
    if "capture_basis" in revenue_estimate:
        row = _write_kv_pair(
            ws, row, "Capture Basis", revenue_estimate["capture_basis"],
        )
    if "annual_degradation_cost_eur" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Annual Shadow Wear Proxy (EUR)",
            revenue_estimate["annual_degradation_cost_eur"],
            _PRICE_FMT,
        )
    if "net_revenue_eur" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Economic Margin after Shadow Wear (EUR)",
            revenue_estimate["net_revenue_eur"],
            _PRICE_FMT,
        )
    if "degradation_pct" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Shadow Wear % of Gross Revenue",
            revenue_estimate["degradation_pct"] / 100,
            _PCT_FMT,
        )
    if "effective_life_years" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Effective Battery Lifetime (years)",
            revenue_estimate["effective_life_years"],
        )
    if "lifetime_limiting_factor" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Lifetime Limiting Factor",
            revenue_estimate["lifetime_limiting_factor"],
        )
    if "annual_throughput_mwh" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Annual Throughput (MWh)",
            revenue_estimate["annual_throughput_mwh"],
            _PRICE_FMT,
        )
    if "lcos_eur_mwh" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Two-leg Throughput Cost (EUR/MWh)",
            revenue_estimate["lcos_eur_mwh"],
            _PRICE_FMT,
        )
    if "net_payback_years" in revenue_estimate:
        net_payback = revenue_estimate["net_payback_years"]
        row = _write_kv_pair(
            ws, row,
            "Economic Payback Proxy (years)",
            net_payback if math.isfinite(net_payback) else "N/A",
        )
    if "cash_npv_includes_shadow_wear" in revenue_estimate:
        row = _write_kv_pair(
            ws, row,
            "Cash NPV Includes Shadow Wear",
            revenue_estimate["cash_npv_includes_shadow_wear"],
        )
    row += 1

    row = _write_kv_pair(ws, row, "Negative Price Hours", negative_stats["negative_hours"])
    row = _write_kv_pair(
        ws, row, "Negative Price Intervals", negative_stats["negative_intervals"],
    )
    row = _write_kv_pair(
        ws, row, "Negative Price % of Intervals", negative_stats["pct_negative"] / 100,
        _PCT_FMT,
    )
    if negative_stats["negative_intervals"] > 0:
        row = _write_kv_pair(ws, row, "Avg Negative Price", negative_stats["avg_negative_price"], _PRICE_FMT)
        row = _write_kv_pair(ws, row, "Most Negative Price", negative_stats["most_negative_price"], _PRICE_FMT)

    _auto_column_width(ws)


def _build_table_sheet(ws, title: str, df: pd.DataFrame) -> None:
    """Write a DataFrame as a formatted table sheet."""
    ws.title = title
    headers = list(df.columns)
    _write_header_row(ws, 1, headers)

    for r_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if value is pd.NA:
                # openpyxl rejects pandas' NA scalar ("Cannot convert <NA>
                # to Excel"); nullable-dtype gaps stay blank cells, matching
                # an empty CSV field. float NaN / NaT keep their branches.
                continue
            if isinstance(value, float):
                cell.value = round(value, 2)
                cell.number_format = _PRICE_FMT
            elif isinstance(value, pd.Timestamp):
                cell.value = str(value)
            else:
                # User-supplied cells (e.g. ancillary CSV ``product``,
                # forward CSV ``contract``) flow through this branch as
                # plain strings. Neutralise formula-trigger leading chars.
                cell.value = _safe_cell_value(value)

    _auto_column_width(ws)


def _build_heatmap_sheet(ws, title: str, heatmap: pd.DataFrame) -> None:
    """Write a heatmap matrix to a sheet."""
    ws.title = title
    # Header row: "Hour" + month columns
    headers = ["Hour", *list(heatmap.columns)]
    _write_header_row(ws, 1, headers)

    for r_idx, (hour, row_data) in enumerate(heatmap.iterrows(), 2):
        ws.cell(row=r_idx, column=1, value=hour)
        for c_idx, value in enumerate(row_data, 2):
            cell = ws.cell(row=r_idx, column=c_idx)
            if pd.notna(value):
                cell.value = round(float(value), 2)
                cell.number_format = _PRICE_FMT

    _auto_column_width(ws)


# ── Project Case (PC-C/PC-D) sheet builders ──────────────────────────────────

def _require_exact_mapping(
    value: Any,
    expected_keys: frozenset[str],
    label: str,
) -> Mapping[str, Any]:
    """Return one exact audit map or fail closed on aliases/extensions."""
    if not isinstance(value, Mapping):
        raise ValueError(f"{label} must be a mapping")
    actual = set(value)
    if actual != expected_keys:
        missing = sorted(expected_keys - actual)
        extra = sorted(actual - expected_keys)
        raise ValueError(
            f"{label} must match the exact v1.1 schema "
            f"(missing={missing}, extra={extra})"
        )
    return value


def _audit_finite_number(value: Any, label: str) -> float:
    """Return one finite audit number without accepting booleans."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} must be numeric")
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{label} must be finite")
    return number


def _audit_f64(value: Any, label: str) -> float:
    """Return one wire-authoritative PC-CBOR-F64 real (never int/bool)."""
    if type(value) is not float:
        raise ValueError(f"{label} must be encoded as a float64 real")
    if not math.isfinite(value):
        raise ValueError(f"{label} must be finite")
    return value


def _audit_strict_int(value: Any, label: str) -> int:
    """Return one public-schema integer without bool/float coercion."""
    if isinstance(value, bool) or not isinstance(value, int):
        raise ValueError(f"{label} must be an integer")
    return int(value)


def _validate_contract_case_payload(
    contract_case: Any,
    project_case: Mapping[str, Any],
) -> Mapping[str, Any] | None:
    """Validate the one PC-D contract union directly from fingerprinted payload."""
    if contract_case is None:
        return None
    contract = _require_exact_mapping(
        contract_case,
        _CONTRACT_CASE_KEYS,
        "RunResult provenance project_case.contract_case",
    )
    if contract.get("settlement_basis") != _CONTRACT_SETTLEMENT_BASIS:
        raise ValueError("Project Case export rejects unknown contract settlement basis")
    terms = _require_exact_mapping(
        contract.get("settlement_terms"),
        _CONTRACT_TERMS_KEYS,
        "RunResult provenance contract settlement_terms",
    )
    if terms.get("quote_basis") != CONTRACT_QUOTE_BASIS_V1:
        raise ValueError("Project Case export requires the modeled-project-MW quote basis")
    if terms.get("settlement_frequency") != CONTRACT_SETTLEMENT_FREQUENCY_V1:
        raise ValueError("Project Case export requires annual project-year-end settlement")
    if terms.get("asset_scope") != CONTRACT_ASSET_SCOPE_V1:
        raise ValueError("Project Case export rejects partial/qualified-MW contract scope")

    currency = _require_exact_mapping(
        terms.get("currency_basis"),
        _CONTRACT_CURRENCY_KEYS,
        "RunResult provenance contract currency_basis",
    )
    if currency.get("mode") != _CONTRACT_CURRENCY_BASIS:
        raise ValueError("Project Case export rejects an unsupported contract currency basis")
    valuation_case = _require_exact_mapping(
        project_case.get("valuation_case"),
        frozenset({"discount_rate", "base_year"}),
        "RunResult provenance project_case.valuation_case",
    )
    if currency.get("target_base_year") != valuation_case.get("base_year"):
        raise ValueError("Contract currency target base year must match ValuationCase")

    rates = terms.get(
        "floor_rate_real_eur_per_modeled_mw_year_by_contract_year"
    )
    factors = terms.get("floor_entitlement_factor_by_contract_year")
    if not isinstance(rates, (tuple, list)) or not rates:
        raise ValueError("Contract floor-rate curve must be a non-empty array")
    if not isinstance(factors, (tuple, list)) or len(factors) != len(rates):
        raise ValueError("Contract entitlement-factor curve must match floor-rate length")
    for index, rate in enumerate(rates):
        if _audit_f64(rate, f"contract floor rate[{index}]") < 0.0:
            raise ValueError("Contract floor rates must be >= 0")
    for index, factor in enumerate(factors):
        value = _audit_f64(factor, f"contract entitlement factor[{index}]")
        if not 0.0 <= value <= 1.0:
            raise ValueError("Contract entitlement factors must be in [0, 1]")

    start = _audit_strict_int(
        terms.get("contract_start_project_year"),
        "contract_start_project_year",
    )
    lifecycle = _require_exact_mapping(
        project_case.get("lifecycle_case"),
        frozenset(
            {
                "project_life_years",
                "capacity_maintenance_basis",
                "capacity_maintenance_source",
                "capacity_maintenance_as_of",
                "augmentation_events",
                "eol_residual_value_eur",
                "decommissioning_cost_eur",
            }
        ),
        "RunResult provenance project_case.lifecycle_case",
    )
    life = _audit_strict_int(lifecycle.get("project_life_years"), "project_life_years")
    if start < 1 or start + len(rates) - 1 > life:
        raise ValueError("Contract coverage must fit entirely inside project life")

    status = terms.get("quote_status")
    if status not in _CONTRACT_QUOTE_STATUSES:
        raise ValueError("Project Case export rejects unknown contract quote status")
    source = terms.get("source")
    if not isinstance(source, str) or not source.strip():
        raise ValueError("Contract source must be non-empty")
    as_of = terms.get("source_as_of_date")
    if not isinstance(as_of, str):
        raise ValueError("Contract source_as_of_date must be an ISO date")
    try:
        parsed_as_of = _dt.date.fromisoformat(as_of)
    except ValueError as exc:
        raise ValueError("Contract source_as_of_date must be a real ISO date") from exc
    if parsed_as_of.isoformat() != as_of:
        raise ValueError("Contract source_as_of_date must be canonical YYYY-MM-DD")
    document_hash = terms.get("source_document_sha256")
    if status == "USER_SCENARIO":
        if document_hash is not None:
            raise ValueError("USER_SCENARIO source_document_sha256 must be null")
    elif not (
        isinstance(document_hash, str)
        and len(document_hash) == 64
        and all(char in "0123456789abcdef" for char in document_hash)
    ):
        raise ValueError("User-asserted quote source_document_sha256 must be lowercase hex64")
    return contract


def _validate_contract_settlement_provenance(
    provenance: Mapping[str, Any],
    project_case: Mapping[str, Any],
    contract_case: Mapping[str, Any] | None,
) -> Mapping[str, Any]:
    """Cross-check PC-D settlement audit against the fingerprinted contract."""
    settlement = _require_exact_mapping(
        provenance.get("contract_settlement"),
        _CONTRACT_SETTLEMENT_KEYS,
        "RunResult provenance contract_settlement",
    )
    floors = settlement.get("resolved_floor_by_project_year")
    if not isinstance(floors, (tuple, list)):
        raise ValueError("contract_settlement.resolved_floor_by_project_year must be an array")

    if contract_case is None:
        if (
            settlement.get("basis") is not None
            or settlement.get("algorithm_version") is not None
            or settlement.get("representative_interpolation") is not None
            or len(floors) != 0
        ):
            raise ValueError("Null ContractCase requires the exact empty settlement audit")
        if (
            provenance.get("cashflow_table_statistic")
            != NULL_CASHFLOW_TABLE_STATISTIC_V1
        ):
            raise ValueError("Null ContractCase requires the merchant P50 statistic")
        return settlement

    if settlement.get("basis") != _CONTRACT_SETTLEMENT_BASIS:
        raise ValueError("Contract settlement audit basis does not match ContractCase")
    if settlement.get("algorithm_version") != CONTRACT_SETTLEMENT_ALGORITHM_V1:
        raise ValueError("Project Case export requires the current contract settlement algorithm")
    if (
        provenance.get("cashflow_table_statistic")
        != CONTRACT_CASHFLOW_TABLE_STATISTIC_V1
    ):
        raise ValueError("Non-null ContractCase requires the rank-interpolated P50 statistic")

    terms = contract_case["settlement_terms"]
    start = int(terms["contract_start_project_year"])
    rates = terms["floor_rate_real_eur_per_modeled_mw_year_by_contract_year"]
    factors = terms["floor_entitlement_factor_by_contract_year"]
    asset = _require_exact_mapping(
        project_case.get("asset_case"),
        frozenset(
            {
                "power_mw",
                "duration_hours",
                "round_trip_efficiency",
                "installed_capex_eur",
                "fixed_om_eur_per_mw_yr",
            }
        ),
        "RunResult provenance project_case.asset_case",
    )
    power_mw = _audit_f64(asset.get("power_mw"), "AssetCase.power_mw")
    expected_by_year = {
        start + index: float(rate) * power_mw * float(factors[index])
        for index, rate in enumerate(rates)
    }
    actual_by_year: dict[int, float] = {}
    previous_year = 0
    for index, record in enumerate(floors):
        item = _require_exact_mapping(
            record,
            _RESOLVED_FLOOR_KEYS,
            f"contract_settlement.resolved_floor_by_project_year[{index}]",
        )
        year = _audit_strict_int(item.get("year"), "resolved floor year")
        value = _audit_f64(item.get("effective_floor_eur"), "effective floor")
        if year <= previous_year or year in actual_by_year:
            raise ValueError("Resolved contract floors must have unique increasing years")
        previous_year = year
        actual_by_year[year] = value
    if set(actual_by_year) != set(expected_by_year):
        raise ValueError("Resolved contract-floor year set does not equal contract coverage")
    for year, expected in expected_by_year.items():
        if actual_by_year[year].hex() != float(expected).hex():
            raise ValueError(
                "Resolved contract floor does not match rate x MW x entitlement"
            )

    interpolation = _require_exact_mapping(
        settlement.get("representative_interpolation"),
        _INTERPOLATION_KEYS,
        "contract_settlement.representative_interpolation",
    )
    bootstrap = provenance.get("bootstrap")
    if not isinstance(bootstrap, Mapping):
        raise ValueError("RunResult provenance bootstrap must be a mapping")
    n_simulations = _audit_strict_int(bootstrap.get("n_simulations"), "n_simulations")
    if n_simulations <= 0:
        raise ValueError("n_simulations must be positive")
    ranks = (
        _audit_strict_int(interpolation.get("lower_sorted_rank"), "lower_sorted_rank"),
        _audit_strict_int(interpolation.get("upper_sorted_rank"), "upper_sorted_rank"),
        _audit_strict_int(
            interpolation.get("lower_original_draw_index"),
            "lower_original_draw_index",
        ),
        _audit_strict_int(
            interpolation.get("upper_original_draw_index"),
            "upper_original_draw_index",
        ),
    )
    if any(index < 0 or index >= n_simulations for index in ranks):
        raise ValueError("Contract representative interpolation index is out of range")
    h = 0.5 * (n_simulations - 1)
    expected_lower_rank = math.floor(h)
    expected_upper_rank = math.ceil(h)
    expected_weight = h - expected_lower_rank
    if ranks[:2] != (expected_lower_rank, expected_upper_rank):
        raise ValueError("Contract representative ranks do not match linear P50")
    weight = _audit_f64(
        interpolation.get("interpolation_weight"),
        "contract interpolation weight",
    )
    if weight != expected_weight:
        raise ValueError("Contract representative weight does not match linear P50")
    return settlement


def _project_case_red_lines(result: RunResult) -> Mapping[str, Any]:
    """Return PC-D2 assertions, failing closed on any cash-basis contradiction."""
    if not isinstance(result, RunResult):
        raise TypeError("project_case_result must be a RunResult")
    if result.schema_version != PROJECT_CASE_SCHEMA_VERSION:
        raise ValueError("Project Case export requires RunResult schema project-case-v1.1")
    provenance = result.provenance
    if set(provenance) != _PROJECT_CASE_PROVENANCE_KEYS:
        missing = sorted(_PROJECT_CASE_PROVENANCE_KEYS - set(provenance))
        extra = sorted(set(provenance) - _PROJECT_CASE_PROVENANCE_KEYS)
        raise ValueError(
            "RunResult provenance must match the exact current PC-D2 schema "
            f"(missing={missing}, extra={extra})"
        )

    assertions = provenance.get("red_line_assertions")
    if not isinstance(assertions, Mapping):
        raise ValueError("RunResult provenance has no red_line_assertions mapping")
    expected_assertion_keys = (
        set(_PROJECT_CASE_STATIC_RED_LINES) | set(_PROJECT_CASE_DYNAMIC_RED_LINES)
    )
    if set(assertions) != expected_assertion_keys:
        missing = sorted(expected_assertion_keys - set(assertions))
        extra = sorted(set(assertions) - expected_assertion_keys)
        raise ValueError(
            "RunResult red_line_assertions must match the exact current PC-D2 schema "
            f"(missing={missing}, extra={extra})"
        )
    for key, required in _PROJECT_CASE_STATIC_RED_LINES.items():
        if assertions.get(key) is not required:
            raise ValueError(
                f"Project Case export requires red_line_assertions.{key}={required!r}"
            )
    if provenance.get("calculator_version") != PC_D2_CALCULATOR_VERSION:
        raise ValueError("Project Case export requires the current PC-D2 calculator")

    strategy = provenance.get("strategy_run_result")
    if not isinstance(strategy, Mapping):
        raise ValueError("RunResult provenance has no strategy_run_result mapping")
    if set(strategy) != _STRATEGY_RUN_RESULT_PAYLOAD_KEYS:
        raise ValueError(
            "RunResult provenance strategy_run_result must match the exact PC-A payload schema"
        )
    _validate_strategy_wire_payload(strategy)
    replayed_strategy = _replay_typed_strategy_payload(strategy)
    computed_strategy_fingerprint = fingerprint_hex("StrategyRunResult", strategy)
    if provenance.get("strategy_run_fingerprint") != computed_strategy_fingerprint:
        raise ValueError(
            "RunResult provenance StrategyRunResult fingerprint does not match its payload"
        )

    project_case = provenance.get("project_case")
    if not isinstance(project_case, Mapping):
        raise ValueError("RunResult provenance has no project_case mapping")
    if set(project_case) != _PROJECT_CASE_PAYLOAD_KEYS:
        raise ValueError(
            "RunResult provenance project_case must match the exact ProjectCase payload schema"
        )
    _replay_typed_project_case_payload(project_case, replayed_strategy)
    market_case = project_case.get("market_case")
    if not isinstance(market_case, Mapping) or set(market_case) != _MARKET_CASE_PAYLOAD_KEYS:
        raise ValueError(
            "RunResult provenance project_case.market_case must match the exact MarketCase payload schema"
        )
    if market_case.get("strategy_run_fingerprint") != computed_strategy_fingerprint:
        raise ValueError(
            "RunResult provenance ProjectCase does not link to its StrategyRunResult payload"
        )
    asset_case = _require_exact_mapping(
        project_case.get("asset_case"),
        _ASSET_CASE_KEYS,
        "RunResult provenance project_case.asset_case",
    )
    for field in ("power_mw", "duration_hours", "round_trip_efficiency"):
        if strategy.get(field) != asset_case.get(field):
            raise ValueError(
                f"StrategyRunResult.{field} does not match ProjectCase AssetCase"
            )
    strategy_currency = strategy.get("currency_basis")
    valuation_case = _require_exact_mapping(
        project_case.get("valuation_case"),
        _VALUATION_CASE_KEYS,
        "RunResult provenance project_case.valuation_case",
    )
    if (
        not isinstance(strategy_currency, Mapping)
        or strategy_currency.get("target_base_year")
        != valuation_case.get("base_year")
    ):
        raise ValueError(
            "StrategyRunResult currency base year does not match ValuationCase"
        )
    computed_project_case_fingerprint = fingerprint_hex("ProjectCase", project_case)
    if (
        provenance.get("project_case_input_fingerprint")
        != computed_project_case_fingerprint
        or result.input_fingerprint != computed_project_case_fingerprint
    ):
        raise ValueError(
            "RunResult provenance ProjectCase fingerprint does not match its payload/result"
        )

    contract_case = _validate_contract_case_payload(
        project_case.get("contract_case"), project_case
    )
    _validate_contract_settlement_provenance(
        provenance, project_case, contract_case
    )
    expected_contract_included = contract_case is not None
    if assertions.get("contract_settlement_included") is not expected_contract_included:
        raise ValueError(
            "red_line_assertions.contract_settlement_included does not match ContractCase"
        )
    expected_basis = (
        _CONTRACT_SETTLEMENT_BASIS if expected_contract_included else None
    )
    if assertions.get("contract_settlement_basis") != expected_basis:
        raise ValueError(
            "red_line_assertions.contract_settlement_basis does not match ContractCase"
        )

    reconciliation = _require_exact_mapping(
        provenance.get("cashflow_reconciliation"),
        _CASHFLOW_RECONCILIATION_KEYS,
        "RunResult provenance cashflow_reconciliation",
    )
    if (
        reconciliation.get("version") != CASHFLOW_RECONCILIATION_VERSION_V1
        or reconciliation.get("relative_tolerance")
        != CASHFLOW_RECONCILIATION_REL_TOL_V1
        or reconciliation.get("absolute_tolerance_eur")
        != CASHFLOW_RECONCILIATION_ABS_TOL_EUR_V1
    ):
        raise ValueError("Project Case export requires the locked P50 reconciliation basis")

    if strategy.get("calculator_version") != PC_A_CALCULATOR_VERSION:
        raise ValueError("Project Case export requires the current PC-A producer")
    adapter = strategy.get("adapter_provenance")
    if not isinstance(adapter, Mapping) or (
        adapter.get("expected_grid_registry_version")
        != EXPECTED_GRID_REGISTRY_VERSION
    ):
        raise ValueError("Project Case export requires the current market-grid registry")

    return assertions


def _audit_close(actual: Any, expected: float, label: str) -> None:
    """Require one exported number to equal its independently replayed value."""
    value = _audit_f64(actual, label)
    if value.hex() != float(expected).hex():
        raise ValueError(f"{label} does not match fingerprinted ProjectCase inputs")


def _replayed_project_vectors(
    result: RunResult,
) -> tuple[
    Mapping[str, Any],
    Mapping[str, Any],
    int,
    np.ndarray,
    np.ndarray,
    np.ndarray,
    np.ndarray,
    np.ndarray,
]:
    """Replay ProjectCase's bootstrap, projection, discount, and lifecycle vectors.

    This intentionally consumes the fingerprinted payload rather than the typed
    table.  Excel is a public cash-output boundary, so an unsafe/forged
    ``RunResult`` must not be able to preserve P50 while changing the displayed
    cash decomposition.
    """
    provenance = result.provenance
    project = provenance["project_case"]
    strategy = provenance["strategy_run_result"]
    asset = _require_exact_mapping(
        project.get("asset_case"), _ASSET_CASE_KEYS, "project_case.asset_case"
    )
    lifecycle = _require_exact_mapping(
        project.get("lifecycle_case"),
        _LIFECYCLE_CASE_KEYS,
        "project_case.lifecycle_case",
    )
    life = _audit_strict_int(lifecycle.get("project_life_years"), "project_life_years")
    if not 1 <= life <= MAX_PROJECT_LIFE_YEARS:
        raise ValueError("project_life_years is outside the Project Case domain")

    project_bootstrap = _require_exact_mapping(
        project.get("bootstrap_case"),
        _BOOTSTRAP_CASE_KEYS,
        "project_case.bootstrap_case",
    )
    seed = _audit_strict_int(project_bootstrap.get("seed"), "bootstrap seed")
    n_simulations = _audit_strict_int(
        project_bootstrap.get("n_simulations"), "bootstrap n_simulations"
    )
    if seed < 0:
        raise ValueError("bootstrap seed must be non-negative")
    if not MIN_SIMULATIONS <= n_simulations <= MAX_SIMULATIONS:
        raise ValueError("bootstrap n_simulations is outside the Project Case domain")
    if project_bootstrap.get("bootstrap_algorithm_version") != BOOTSTRAP_ALGORITHM_V1:
        raise ValueError("Project Case export requires the locked bootstrap algorithm")
    bootstrap = _require_exact_mapping(
        provenance.get("bootstrap"),
        _BOOTSTRAP_PROVENANCE_KEYS,
        "RunResult provenance bootstrap",
    )
    expected_bootstrap = {
        "algorithm_version": BOOTSTRAP_ALGORITHM_V1,
        "seed": seed,
        "n_simulations": n_simulations,
        "days_per_annual_draw": 365,
        "percentile_method": "linear",
        "prob_positive_rule": "strict_gt_zero_all_draws",
    }
    if any(bootstrap.get(key) != value for key, value in expected_bootstrap.items()):
        raise ValueError("RunResult bootstrap provenance does not match ProjectCase")

    series = strategy.get("daily_realised_cash_series")
    if not isinstance(series, (tuple, list)) or not series:
        raise ValueError("StrategyRunResult daily cash series must be non-empty")
    daily_values: list[float] = []
    for index, item in enumerate(series):
        if not isinstance(item, (tuple, list)) or len(item) != 2:
            raise ValueError(f"daily cash series[{index}] must be [date, cash]")
        daily_values.append(
            _audit_f64(item[1], f"daily cash series[{index}].cash")
        )
    annual_draws = bootstrap_annual_sums(
        np.asarray(daily_values, dtype=np.float64),
        seed=seed,
        n_simulations=n_simulations,
    )

    market = _require_exact_mapping(
        project.get("market_case"),
        _MARKET_CASE_PAYLOAD_KEYS,
        "project_case.market_case",
    )
    projection = _require_exact_mapping(
        market.get("projection"),
        _PROJECTION_PAYLOAD_KEYS,
        "project_case.market_case.projection",
    )
    kind = projection.get("projection_kind")
    if kind == ProjectionKind.FlatRealProjection.value:
        if any(
            projection.get(key) is not None
            for key in (
                "annual_decay_rate",
                "decay_floor_share",
                "multipliers",
                "source",
                "as_of",
            )
        ):
            raise ValueError("FlatRealProjection payload contains inapplicable fields")
        multipliers = np.ones(life, dtype=np.float64)
    elif kind == ProjectionKind.DAOnlySpreadDecay.value:
        decay = _audit_f64(
            projection.get("annual_decay_rate"), "annual_decay_rate"
        )
        floor = _audit_f64(
            projection.get("decay_floor_share"), "decay_floor_share"
        )
        if not 0.0 <= decay < 1.0 or not 0.0 <= floor <= 1.0:
            raise ValueError("DAOnlySpreadDecay parameters are outside their domain")
        if any(
            projection.get(key) is not None
            for key in ("multipliers", "source", "as_of")
        ):
            raise ValueError("DAOnlySpreadDecay payload contains inapplicable fields")
        multipliers = np.asarray(
            [max((1.0 - decay) ** year, floor) for year in range(life)],
            dtype=np.float64,
        )
    elif kind == ProjectionKind.ExplicitAnnualMultiplierCurve.value:
        raw = projection.get("multipliers")
        if not isinstance(raw, (tuple, list)) or len(raw) != life:
            raise ValueError("ExplicitAnnualMultiplierCurve must cover project life")
        multipliers = np.asarray(
            [
                _audit_f64(value, f"projection multiplier[{index}]")
                for index, value in enumerate(raw)
            ],
            dtype=np.float64,
        )
        if (
            projection.get("annual_decay_rate") is not None
            or projection.get("decay_floor_share") is not None
            or not isinstance(projection.get("source"), str)
            or not projection.get("source").strip()
            or not isinstance(projection.get("as_of"), str)
        ):
            raise ValueError("ExplicitAnnualMultiplierCurve payload is incomplete")
    else:
        raise ValueError("Project Case export rejects an unknown projection kind")
    if (
        multipliers.shape != (life,)
        or not np.isfinite(multipliers).all()
        or np.any(multipliers < 0.0)
        or multipliers[0] != 1.0
    ):
        raise ValueError("Resolved projection multipliers are outside their domain")
    projection_audit = _require_exact_mapping(
        provenance.get("projection"),
        _PROJECTION_PROVENANCE_KEYS,
        "RunResult provenance projection",
    )
    resolved = projection_audit.get("resolved_annual_multipliers")
    if (
        projection_audit.get("projection_kind") != kind
        or not isinstance(resolved, (tuple, list))
        or len(resolved) != life
    ):
        raise ValueError("RunResult projection provenance does not match ProjectCase")
    for year, (actual, expected) in enumerate(
        zip(resolved, multipliers, strict=True), start=1
    ):
        _audit_close(actual, float(expected), f"projection multiplier year {year}")

    valuation = _require_exact_mapping(
        project.get("valuation_case"),
        _VALUATION_CASE_KEYS,
        "project_case.valuation_case",
    )
    rate = _audit_f64(valuation.get("discount_rate"), "discount_rate")
    if rate <= -1.0:
        raise ValueError("discount_rate must be greater than -1")
    base_year = _audit_strict_int(valuation.get("base_year"), "valuation base_year")
    valuation_audit = _require_exact_mapping(
        provenance.get("valuation"),
        _VALUATION_PROVENANCE_KEYS,
        "RunResult provenance valuation",
    )
    if (
        valuation_audit.get("discount_rate") != valuation.get("discount_rate")
        or valuation_audit.get("base_year") != base_year
        or valuation_audit.get("currency_convention") != "real_base_year_eur"
        or valuation_audit.get("cash_timing") != "end_of_year"
    ):
        raise ValueError("RunResult valuation provenance does not match ProjectCase")
    with np.errstate(over="ignore", under="ignore", divide="ignore", invalid="ignore"):
        discount_factors = np.power(
            1.0 + rate,
            -np.arange(1, life + 1, dtype=np.float64),
            dtype=np.float64,
        )
    if not np.isfinite(discount_factors).all() or np.any(discount_factors <= 0.0):
        raise ValueError("Project Case discount factors are not finite and positive")

    asset_reals = {
        key: _audit_f64(asset.get(key), f"AssetCase.{key}")
        for key in _ASSET_CASE_KEYS
    }
    power_mw = asset_reals["power_mw"]
    fixed_om_rate = asset_reals["fixed_om_eur_per_mw_yr"]
    if (
        power_mw <= 0.0
        or asset_reals["duration_hours"] <= 0.0
        or not 0.0 < asset_reals["round_trip_efficiency"] <= 1.0
        or asset_reals["installed_capex_eur"] < 0.0
        or fixed_om_rate < 0.0
    ):
        raise ValueError("AssetCase real fields are outside their domains")
    maintenance_unknown = (
        lifecycle.get("capacity_maintenance_basis")
        == CapacityMaintenanceBasis.UNKNOWN.value
    )
    fixed_om = 0.0 if maintenance_unknown else power_mw * fixed_om_rate
    if not math.isfinite(fixed_om):
        raise ValueError("annual fixed O&M is not finite")
    augmentation = np.zeros(life, dtype=np.float64)
    events = lifecycle.get("augmentation_events")
    if not isinstance(events, (tuple, list)):
        raise ValueError("LifecycleCase augmentation_events must be an array")
    event_cash_by_year: list[list[float]] = [[] for _ in range(life)]
    for index, raw_event in enumerate(events):
        event = _require_exact_mapping(
            raw_event,
            _AUGMENTATION_EVENT_KEYS,
            f"LifecycleCase augmentation_events[{index}]",
        )
        year = _audit_strict_int(event.get("year"), "augmentation event year")
        if not 1 <= year <= life:
            raise ValueError("augmentation event year is outside project life")
        cost = _audit_f64(event.get("cost_eur"), "augmentation event cost")
        restored = _audit_f64(
            event.get("capacity_restored_frac"), "augmentation capacity restored"
        )
        residual = _audit_f64(
            event.get("residual_value_eur"), "augmentation event residual"
        )
        if cost < 0.0 or residual < 0.0 or not 0.0 <= restored <= 1.0:
            raise ValueError("augmentation event real fields are outside their domains")
        if not maintenance_unknown:
            event_cash_by_year[year - 1].append(cost - residual)
    if not maintenance_unknown:
        try:
            augmentation = np.asarray(
                [math.fsum(values) for values in event_cash_by_year],
                dtype=np.float64,
            )
        except OverflowError as exc:
            raise ValueError("annual augmentation cash overflowed") from exc
    if not np.isfinite(augmentation).all():
        raise ValueError("annual augmentation cash is not finite")
    terminal = np.zeros(life, dtype=np.float64)
    eol_residual = _audit_f64(
        lifecycle.get("eol_residual_value_eur"), "eol residual value"
    )
    decommissioning = _audit_f64(
        lifecycle.get("decommissioning_cost_eur"), "decommissioning cost"
    )
    if eol_residual < 0.0 or decommissioning < 0.0:
        raise ValueError("Lifecycle terminal inputs must be non-negative")
    if not maintenance_unknown:
        terminal[-1] = eol_residual - decommissioning
    if not math.isfinite(float(terminal[-1])):
        raise ValueError("terminal cash is not finite")
    return (
        project,
        lifecycle,
        life,
        annual_draws,
        multipliers,
        discount_factors,
        np.full(life, fixed_om, dtype=np.float64),
        augmentation,
        terminal,
    )


def _replayed_contract_floor_vectors(
    result: RunResult,
    *,
    life: int,
) -> tuple[np.ndarray, np.ndarray, tuple[float | None, ...]]:
    """Resolve F=q*MW*a from fingerprinted ContractCase, never its audit copy."""
    project = result.provenance["project_case"]
    covered = np.zeros(life, dtype=np.bool_)
    floors = np.zeros(life, dtype=np.float64)
    optional: list[float | None] = [None] * life
    contract = project["contract_case"]
    if contract is None:
        return covered, floors, tuple(optional)
    terms = contract["settlement_terms"]
    start = int(terms["contract_start_project_year"])
    power_mw = float(project["asset_case"]["power_mw"])
    rates = terms["floor_rate_real_eur_per_modeled_mw_year_by_contract_year"]
    factors = terms["floor_entitlement_factor_by_contract_year"]
    for offset, (rate, factor) in enumerate(zip(rates, factors, strict=True)):
        index = start - 1 + offset
        value = float(rate) * power_mw
        value *= float(factor)
        if not math.isfinite(value):
            raise ValueError("Replayed effective contract floor is not finite")
        covered[index] = True
        floors[index] = value
        optional[index] = value
    return covered, floors, tuple(optional)


def _replayed_representative_revenue(
    result: RunResult,
    *,
    annual_draws: np.ndarray,
    multipliers: np.ndarray,
    life: int,
) -> tuple[np.ndarray, tuple[float | None, ...], np.ndarray, np.ndarray]:
    """Replay the exact null/contract representative revenue decomposition."""
    project = result.provenance["project_case"]
    contract = project["contract_case"]
    if contract is None:
        annual_p50 = float(np.percentile(annual_draws, 50.0, method="linear"))
        merchant = annual_p50 * multipliers
        return (
            merchant,
            tuple(None for _ in range(life)),
            np.zeros(life, dtype=np.float64),
            merchant.copy(),
        )

    settlement = result.provenance["contract_settlement"]
    covered, floor_array, floor_values = _replayed_contract_floor_vectors(
        result, life=life
    )
    interpolation = settlement["representative_interpolation"]
    order = np.lexsort(
        (np.arange(annual_draws.size, dtype=np.int64), annual_draws)
    )
    lower_rank = int(interpolation["lower_sorted_rank"])
    upper_rank = int(interpolation["upper_sorted_rank"])
    lower_original = int(interpolation["lower_original_draw_index"])
    upper_original = int(interpolation["upper_original_draw_index"])
    if (
        lower_original != int(order[lower_rank])
        or upper_original != int(order[upper_rank])
    ):
        raise ValueError(
            "Contract representative original draw indices do not match bootstrap order"
        )
    weight = float(interpolation["interpolation_weight"])
    lower_merchant = annual_draws[lower_original] * multipliers
    upper_merchant = annual_draws[upper_original] * multipliers
    lower_settled = lower_merchant.copy()
    upper_settled = upper_merchant.copy()
    for offset in range(life):
        if bool(covered[offset]):
            lower_settled[offset] = max(lower_settled[offset], floor_array[offset])
            upper_settled[offset] = max(upper_settled[offset], floor_array[offset])
    merchant = (1.0 - weight) * lower_merchant + weight * upper_merchant
    settled = (1.0 - weight) * lower_settled + weight * upper_settled
    top_up = settled - merchant
    if not (
        np.isfinite(merchant).all()
        and np.isfinite(settled).all()
        and np.isfinite(top_up).all()
    ):
        raise ValueError("Replayed representative contract cash is not finite")
    return merchant, floor_values, top_up, settled


def _require_available_outcome(outcome: Any, label: str) -> None:
    """Validate an available NPV outcome even if RunResult construction was bypassed."""
    if not isinstance(outcome, NpvOutcome):
        raise ValueError(f"{label} NPV outcome must be an NpvOutcome")
    if (
        outcome.available is not True
        or outcome.status != "ok"
        or outcome.message is not None
        or outcome.distribution is None
    ):
        raise ValueError(f"{label} NPV outcome must be the exact available state")


def _replayed_npv_draws(
    result: RunResult,
    *,
    annual_draws: np.ndarray,
    multipliers: np.ndarray,
    discount_factors: np.ndarray,
    fixed_om: np.ndarray,
    augmentation: np.ndarray,
    terminal: np.ndarray,
    maintenance_unknown: bool,
) -> tuple[np.ndarray, np.ndarray | None]:
    """Replay every screening/lifecycle NPV draw from fingerprinted inputs."""
    project = result.provenance["project_case"]
    life = int(project["lifecycle_case"]["project_life_years"])
    capex = float(project["asset_case"]["installed_capex_eur"])
    screening = np.full(annual_draws.shape, -capex, dtype=np.float64)
    if project["contract_case"] is None:
        with np.errstate(over="ignore", invalid="ignore"):
            for multiplier, discount_factor in zip(
                multipliers, discount_factors, strict=True
            ):
                screening += annual_draws * multiplier * discount_factor
    else:
        covered, floors, _optional = _replayed_contract_floor_vectors(
            result, life=life
        )
        with np.errstate(over="ignore", invalid="ignore"):
            merchant = annual_draws[:, np.newaxis] * multipliers[np.newaxis, :]
        if not np.isfinite(merchant).all():
            raise ValueError("Replayed merchant cash matrix is not finite")
        settled = merchant.copy()
        if np.any(covered):
            with np.errstate(over="ignore", invalid="ignore"):
                settled[:, covered] = np.maximum(
                    merchant[:, covered], floors[np.newaxis, covered]
                )
        with np.errstate(over="ignore", invalid="ignore"):
            for index, discount_factor in enumerate(discount_factors):
                screening += settled[:, index] * float(discount_factor)
    if not np.isfinite(screening).all():
        raise ValueError("Replayed screening NPV draws are not finite")
    if maintenance_unknown:
        return screening, None

    with np.errstate(over="ignore", invalid="ignore"):
        annual_adjustment = -fixed_om - augmentation + terminal
        discounted_adjustment = annual_adjustment * discount_factors
    if not (
        np.isfinite(annual_adjustment).all()
        and np.isfinite(discounted_adjustment).all()
    ):
        raise ValueError("Replayed lifecycle adjustment is not finite")
    try:
        adjustment = math.fsum(float(value) for value in discounted_adjustment)
    except OverflowError as exc:
        raise ValueError("Replayed lifecycle adjustment overflowed") from exc
    with np.errstate(over="ignore", invalid="ignore"):
        lifecycle = screening + adjustment
    if not np.isfinite(lifecycle).all():
        raise ValueError("Replayed lifecycle NPV draws are not finite")
    return screening, lifecycle


def _validate_replayed_outcome(
    outcome: Any,
    *,
    label: str,
    draws: np.ndarray,
) -> None:
    """Bind all four reported distribution statistics to replayed NPV draws."""
    _require_available_outcome(outcome, label)
    distribution = outcome.distribution
    if not isinstance(distribution, NpvDistribution):
        raise ValueError(f"{label} NPV outcome has no typed distribution")
    percentiles = np.percentile(
        draws,
        [10.0, 50.0, 90.0],
        method="linear",
    )
    probability = float(np.mean(draws > 0.0))
    for field, actual, expected in (
        ("p10", distribution.p10, percentiles[0]),
        ("p50", distribution.p50, percentiles[1]),
        ("p90", distribution.p90, percentiles[2]),
        ("prob_positive", distribution.prob_positive, probability),
    ):
        _audit_close(actual, float(expected), f"{label} NPV {field}")


def _validate_replayed_cashflow_table(
    table: Any,
    *,
    basis: str,
    merchant: np.ndarray,
    floors: tuple[float | None, ...],
    top_up: np.ndarray,
    settled: np.ndarray,
    discount_factors: np.ndarray,
    fixed_om: np.ndarray,
    augmentation: np.ndarray,
    terminal: np.ndarray,
) -> None:
    """Bind every displayed row component to the fingerprinted input replay."""
    if not isinstance(table, CashflowTable) or table.basis != basis:
        raise ValueError(f"{basis} cash-flow table has the wrong type or basis")
    life = len(floors)
    if len(table.rows) != life:
        raise ValueError(f"{basis} cash-flow table must cover exactly years 1..{life}")
    for index, row in enumerate(table.rows):
        year = index + 1
        if not isinstance(row, CashflowRow) or row.year != year:
            raise ValueError(f"{basis} cash-flow table must cover exactly years 1..{life}")
        expected_opex = 0.0 if basis == "screening" else float(fixed_om[index])
        expected_augmentation = (
            0.0 if basis == "screening" else float(augmentation[index])
        )
        expected_terminal = 0.0 if basis == "screening" else float(terminal[index])
        expected_net = (
            float(settled[index])
            - expected_opex
            - expected_augmentation
            + expected_terminal
        )
        expected_discounted = expected_net * float(discount_factors[index])
        expected_floor = floors[index]
        if expected_floor is None:
            if row.effective_contract_floor_eur is not None:
                raise ValueError(
                    f"{basis} cash-flow year {year} has a floor outside contract coverage"
                )
        elif row.effective_contract_floor_eur is None:
            raise ValueError(
                f"{basis} cash-flow year {year} omits the fingerprinted contract floor"
            )
        else:
            _audit_close(
                row.effective_contract_floor_eur,
                expected_floor,
                f"{basis} cash-flow year {year} effective floor",
            )
        for field, actual, expected in (
            ("merchant cash", row.merchant_revenue_eur, merchant[index]),
            ("contract top-up", row.contract_top_up_eur, top_up[index]),
            ("settled cash", row.revenue_eur, settled[index]),
            ("fixed O&M", row.opex_eur, expected_opex),
            ("augmentation", row.augmentation_eur, expected_augmentation),
            ("terminal cash", row.terminal_eur, expected_terminal),
            ("net cash", row.net_eur, expected_net),
            ("discount factor", row.discount_factor, discount_factors[index]),
            ("discounted net cash", row.discounted_net_eur, expected_discounted),
        ):
            _audit_close(
                actual,
                float(expected),
                f"{basis} cash-flow year {year} {field}",
            )


def _validate_project_case_cashflow_replay(result: RunResult) -> None:
    """Independently validate RunResult state and both representative tables."""
    (
        _project,
        lifecycle,
        life,
        annual_draws,
        multipliers,
        discount_factors,
        fixed_om,
        augmentation,
        terminal,
    ) = _replayed_project_vectors(result)
    merchant, floors, top_up, settled = _replayed_representative_revenue(
        result,
        annual_draws=annual_draws,
        multipliers=multipliers,
        life=life,
    )

    basis = lifecycle.get("capacity_maintenance_basis")
    if result.provenance.get("capacity_maintenance_basis") != basis:
        raise ValueError(
            "RunResult capacity_maintenance_basis does not match fingerprinted LifecycleCase"
        )
    maintenance_unknown = basis == CapacityMaintenanceBasis.UNKNOWN.value
    screening_draws, lifecycle_draws = _replayed_npv_draws(
        result,
        annual_draws=annual_draws,
        multipliers=multipliers,
        discount_factors=discount_factors,
        fixed_om=fixed_om,
        augmentation=augmentation,
        terminal=terminal,
        maintenance_unknown=maintenance_unknown,
    )
    _validate_replayed_outcome(
        result.no_lifecycle_cost_screening_npv,
        label="screening",
        draws=screening_draws,
    )
    _validate_replayed_cashflow_table(
        result.screening_cashflow_table,
        basis="screening",
        merchant=merchant,
        floors=floors,
        top_up=top_up,
        settled=settled,
        discount_factors=discount_factors,
        fixed_om=fixed_om,
        augmentation=augmentation,
        terminal=terminal,
    )

    if maintenance_unknown:
        outcome = result.lifecycle_cash_npv
        if not isinstance(outcome, NpvOutcome) or (
            outcome.available is not False
            or outcome.status != LIFECYCLE_UNKNOWN_STATUS
            or outcome.message != LIFECYCLE_UNKNOWN_MESSAGE
            or outcome.distribution is not None
            or result.lifecycle_cashflow_table is not None
        ):
            raise ValueError(
                "UNKNOWN capacity-maintenance basis requires the exact unavailable "
                "lifecycle outcome and a null lifecycle table"
            )
        return
    active_bases = {
        CapacityMaintenanceBasis.SCHEDULED_NAMEPLATE_MAINTENANCE.value,
        CapacityMaintenanceBasis.NO_AUGMENTATION_REQUIRED_ASSERTED.value,
    }
    if basis not in active_bases:
        raise ValueError("Project Case export rejects an unknown maintenance basis")
    if lifecycle_draws is None:  # pragma: no cover - guarded by basis above
        raise ValueError("Active maintenance basis has no replayed lifecycle draws")
    _validate_replayed_outcome(
        result.lifecycle_cash_npv,
        label="lifecycle",
        draws=lifecycle_draws,
    )
    _validate_replayed_cashflow_table(
        result.lifecycle_cashflow_table,
        basis="lifecycle",
        merchant=merchant,
        floors=floors,
        top_up=top_up,
        settled=settled,
        discount_factors=discount_factors,
        fixed_om=fixed_om,
        augmentation=augmentation,
        terminal=terminal,
    )


def _strategy_coverage_summary(result: RunResult) -> dict[str, Any]:
    """Return typed producer/coverage facts for the visible NPV summary."""
    strategy = result.provenance["strategy_run_result"]
    adapter = strategy["adapter_provenance"]
    audit = strategy.get("coverage_audit")
    if not isinstance(audit, Mapping):
        raise ValueError("RunResult provenance has no coverage_audit mapping")
    date_sets: dict[str, tuple[Any, ...]] = {}
    for key in (
        "observed_dates",
        "valid_dates",
        "missing_dates",
        "solver_failed_dates",
    ):
        value = audit.get(key)
        if not isinstance(value, (tuple, list)):
            raise ValueError(f"RunResult coverage_audit.{key} must be an array")
        date_sets[key] = tuple(value)
    return {
        "Strategy Kind": strategy.get("strategy_kind"),
        "Producer Adapter": adapter.get("producer_adapter_id"),
        "Observed Dates": len(date_sets["observed_dates"]),
        "Valid Dates": len(date_sets["valid_dates"]),
        "Missing Dates": len(date_sets["missing_dates"]),
        "Solver-failed Dates": len(date_sets["solver_failed_dates"]),
    }


def _outcome_value(outcome: NpvOutcome, field: str) -> Any:
    """Read one displayed NPV value using only the typed availability envelope."""
    if field == "available":
        return outcome.available
    if field == "status":
        return outcome.status
    if field == "message":
        return outcome.message
    if not outcome.available:
        return None
    distribution = outcome.distribution
    # NpvOutcome's state matrix guarantees this, but keep the presentation
    # boundary fail-closed if a future schema version weakens that invariant.
    if distribution is None:
        raise ValueError("available NpvOutcome has no distribution")
    return getattr(distribution, field)


def _build_project_case_npv_sheet(ws, result: RunResult) -> None:
    """Write the two explicitly labelled Project Case NPV outcomes."""
    ws.title = _PROJECT_CASE_NPV_SHEET
    screening = result.no_lifecycle_cost_screening_npv
    lifecycle = result.lifecycle_cash_npv
    headers = [
        "Metric",
        "No-lifecycle-cost screening NPV",
        "Pre-tax unlevered lifecycle cash NPV",
    ]
    _write_header_row(ws, 1, headers)
    rows = [
        ("Available", _outcome_value(screening, "available"), _outcome_value(lifecycle, "available")),
        ("Status", _outcome_value(screening, "status"), _outcome_value(lifecycle, "status")),
        ("Message", _outcome_value(screening, "message"), _outcome_value(lifecycle, "message")),
        (
            "P10 (Downside, EUR)",
            _outcome_value(screening, "p10"),
            _outcome_value(lifecycle, "p10"),
        ),
        (
            "P50 (Median, EUR)",
            _outcome_value(screening, "p50"),
            _outcome_value(lifecycle, "p50"),
        ),
        (
            "P90 (Upside, EUR)",
            _outcome_value(screening, "p90"),
            _outcome_value(lifecycle, "p90"),
        ),
        (
            "P(NPV > 0)",
            _outcome_value(screening, "prob_positive"),
            _outcome_value(lifecycle, "prob_positive"),
        ),
    ]
    for row_index, values in enumerate(rows, 2):
        for column_index, value in enumerate(values, 1):
            cell = ws.cell(
                row=row_index,
                column=column_index,
                value=_safe_cell_value(value),
            )
            if row_index in {5, 6, 7} and column_index > 1 and value is not None:
                cell.number_format = "#,##0.00"
            elif row_index == 8 and column_index > 1 and value is not None:
                cell.number_format = _PCT_FMT

    row = len(rows) + 3
    row = _write_kv_pair(ws, row, "Schema Version", result.schema_version)
    row = _write_kv_pair(ws, row, "Project Case Input Fingerprint", result.input_fingerprint)
    row = _write_kv_pair(
        ws,
        row,
        "Cash-flow Table Statistic",
        result.provenance.get("cashflow_table_statistic"),
    )
    assertions = _project_case_red_lines(result)
    row = _write_kv_pair(
        ws,
        row,
        "Typed Contract Settlement Included",
        assertions["contract_settlement_included"],
    )
    row = _write_kv_pair(
        ws,
        row,
        "Contract Settlement Basis",
        assertions["contract_settlement_basis"],
    )
    row = _write_kv_pair(
        ws,
        row,
        "Wear-net Floor Comparator Included",
        assertions["wear_net_floor_comparator_included"],
    )
    row = _write_kv_pair(ws, row, "Valuation Basis", "Pre-tax unlevered; real base-year EUR")
    contract_included = assertions["contract_settlement_included"]
    row = _write_kv_pair(
        ws,
        row,
        "Representative Table Basis",
        (
            "Rank-interpolated P50 NPV reconciliation path; neither an actual "
            "scenario nor a per-year median"
            if contract_included
            else "Linear P50 annual bootstrap draw; not an expected-value table"
        ),
    )
    row = _write_kv_pair(
        ws,
        row,
        "P50 Reconciliation",
        "NPV P50 = -year-0 CapEx + sum(discounted net cash flow)",
    )
    if contract_included:
        row = _write_kv_pair(ws, row, "Contract Product Boundary", _CONTRACT_DISCLOSURE)
    row += 1
    for label, value in _strategy_coverage_summary(result).items():
        row = _write_kv_pair(ws, row, label, value)
    _auto_column_width(ws)


def _cashflow_frame(table: CashflowTable) -> pd.DataFrame:
    """Map every typed CashflowRow field to one stable Excel column."""
    return pd.DataFrame(
        [
            {
                "Project Year": row.year,
                "Merchant Strategy Cash (EUR)": row.merchant_revenue_eur,
                "Effective Contract Floor (EUR)": row.effective_contract_floor_eur,
                "Contract Top-Up (EUR)": row.contract_top_up_eur,
                "Settled Strategy Cash (EUR)": row.revenue_eur,
                "Fixed O&M (EUR)": row.opex_eur,
                "Augmentation Net Cost (EUR)": row.augmentation_eur,
                "Terminal Cash (EUR)": row.terminal_eur,
                "Net Cash Flow (EUR)": row.net_eur,
                "Discount Factor": row.discount_factor,
                "Discounted Net Cash Flow (EUR)": row.discounted_net_eur,
            }
            for row in table.rows
        ]
    )


def _build_project_case_cashflow_sheet(
    ws,
    title: str,
    table: CashflowTable,
    *,
    installed_capex_eur: float,
    reported_p50_eur: float,
) -> None:
    """Write one representative table plus an explicit year-0 reconciliation."""
    frame = _cashflow_frame(table)
    ws.title = title
    if not math.isfinite(installed_capex_eur) or installed_capex_eur < 0.0:
        raise ValueError("Project Case installed_capex_eur must be finite and >= 0")
    if not math.isfinite(reported_p50_eur):
        raise ValueError("Project Case reported P50 NPV must be finite")
    discounted_project_cash_eur = math.fsum(
        row.discounted_net_eur for row in table.rows
    )
    reconciled_p50_eur = -installed_capex_eur + discounted_project_cash_eur
    if not math.isclose(
        reconciled_p50_eur,
        reported_p50_eur,
        rel_tol=CASHFLOW_RECONCILIATION_REL_TOL_V1,
        abs_tol=CASHFLOW_RECONCILIATION_ABS_TOL_EUR_V1,
    ):
        raise ValueError(
            f"{title} does not reconcile: -installed CapEx + discounted "
            "project-year cash flow differs from reported P50 NPV"
        )

    _write_header_row(ws, 1, ["Representative P50 reconciliation", "Value (EUR)"])
    reconciliation_rows = (
        ("Installed CapEx at year 0", installed_capex_eur),
        ("Discounted project-year cash flow", discounted_project_cash_eur),
        ("Reconciled P50 NPV", reconciled_p50_eur),
        ("Reported P50 NPV", reported_p50_eur),
        ("Reconciliation difference", reconciled_p50_eur - reported_p50_eur),
    )
    for row_index, (label, value) in enumerate(reconciliation_rows, start=2):
        ws.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row_index, column=2, value=value)
        cell.number_format = "#,##0.00"

    table_header_row = len(reconciliation_rows) + 3
    _write_header_row(ws, table_header_row, list(frame.columns))
    discount_column = list(frame.columns).index("Discount Factor") + 1
    for row_index, values in enumerate(
        frame.itertuples(index=False), table_header_row + 1
    ):
        for column_index, value in enumerate(values, 1):
            # Do not round the stored value: the representative rows must
            # reconcile to the typed P50 NPV when a user sums them in Excel.
            cell = ws.cell(row=row_index, column=column_index, value=value)
            if column_index == discount_column:
                cell.number_format = "0.000000"
            elif column_index > 1:
                cell.number_format = "#,##0.00"
    _auto_column_width(ws)


def _build_contract_settlement_sheet(ws, result: RunResult) -> None:
    """Write the minimum human-readable PC-D2 contract disclosure surface."""
    ws.title = _CONTRACT_SETTLEMENT_SHEET
    provenance = result.provenance
    project_case = provenance["project_case"]
    contract_case = project_case["contract_case"]
    if not isinstance(contract_case, Mapping):
        raise ValueError("Contract Settlement sheet requires a non-null ContractCase")
    terms = contract_case["settlement_terms"]
    settlement = provenance["contract_settlement"]
    interpolation = settlement["representative_interpolation"]
    asset_case = project_case["asset_case"]

    ws.cell(row=1, column=1, value="Project Case — Contract Settlement").font = Font(
        bold=True, size=14
    )
    row = 3
    row = _write_kv_pair(ws, row, "Product Boundary", _CONTRACT_DISCLOSURE)
    row = _write_kv_pair(ws, row, "Settlement Basis", contract_case["settlement_basis"])
    row = _write_kv_pair(ws, row, "Settlement Algorithm", settlement["algorithm_version"])
    row = _write_kv_pair(ws, row, "Quote Status", terms["quote_status"])
    row = _write_kv_pair(ws, row, "Source", terms["source"])
    row = _write_kv_pair(ws, row, "Source As-of Date", terms["source_as_of_date"])
    row = _write_kv_pair(ws, row, "Source Document SHA-256", terms["source_document_sha256"])
    row = _write_kv_pair(ws, row, "Modelled Whole-project Power (MW)", asset_case["power_mw"])
    row = _write_kv_pair(ws, row, "Quote Basis", terms["quote_basis"])
    row = _write_kv_pair(ws, row, "Asset Scope", terms["asset_scope"])
    row = _write_kv_pair(ws, row, "Settlement Frequency", terms["settlement_frequency"])
    row = _write_kv_pair(
        ws,
        row,
        "Contract Start Project Year",
        terms["contract_start_project_year"],
    )
    rates = terms["floor_rate_real_eur_per_modeled_mw_year_by_contract_year"]
    factors = terms["floor_entitlement_factor_by_contract_year"]
    row = _write_kv_pair(ws, row, "Contract Tenor (years)", len(rates))
    currency_basis = terms["currency_basis"]
    row = _write_kv_pair(ws, row, "Currency Basis", currency_basis["mode"])
    row = _write_kv_pair(ws, row, "Real-EUR Base Year", currency_basis["target_base_year"])
    row = _write_kv_pair(
        ws,
        row,
        "Representative Table Basis",
        "Rank-interpolated P50 path; neither an actual scenario nor a per-year median",
    )
    for key, label in (
        ("lower_sorted_rank", "Lower Sorted Rank (zero-based)"),
        ("upper_sorted_rank", "Upper Sorted Rank (zero-based)"),
        ("lower_original_draw_index", "Lower Original Draw Index (zero-based)"),
        ("upper_original_draw_index", "Upper Original Draw Index (zero-based)"),
        ("interpolation_weight", "Interpolation Weight"),
    ):
        row = _write_kv_pair(ws, row, label, interpolation[key])
    reconciliation = provenance["cashflow_reconciliation"]
    row = _write_kv_pair(ws, row, "Reconciliation Version", reconciliation["version"])
    row = _write_kv_pair(
        ws,
        row,
        "Reconciliation Tolerance",
        (
            f"rel={reconciliation['relative_tolerance']!r}; "
            f"abs EUR={reconciliation['absolute_tolerance_eur']!r}"
        ),
    )

    table_header_row = row + 1
    headers = [
        "Contract Year",
        "Project Year",
        "Real Floor Rate (EUR/modelled MW-year)",
        "Floor Entitlement Factor",
        "Effective Whole-project Floor (EUR)",
    ]
    _write_header_row(ws, table_header_row, headers)
    floor_by_year = {
        item["year"]: item["effective_floor_eur"]
        for item in settlement["resolved_floor_by_project_year"]
    }
    start = terms["contract_start_project_year"]
    for contract_year, (rate, factor) in enumerate(
        zip(rates, factors, strict=True), start=1
    ):
        project_year = start + contract_year - 1
        values = (
            contract_year,
            project_year,
            rate,
            factor,
            floor_by_year[project_year],
        )
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(
                row=table_header_row + contract_year,
                column=column_index,
                value=value,
            )
            if column_index in {3, 5}:
                cell.number_format = "#,##0.00"
            elif column_index == 4:
                cell.number_format = _PCT_FMT
    _auto_column_width(ws)


def _flatten_audit_tree(value: Any, path: str) -> list[tuple[str, Any, str]]:
    """Flatten the immutable provenance tree without dropping null/empty values."""
    if isinstance(value, Mapping):
        if not value:
            return [(path, "{}", "object")]
        rows: list[tuple[str, Any, str]] = []
        for key in sorted(value):
            child = f"{path}.{key}" if path else str(key)
            rows.extend(_flatten_audit_tree(value[key], child))
        return rows
    if isinstance(value, (tuple, list)):
        if not value:
            return [(path, "[]", "array")]
        rows = []
        for index, item in enumerate(value):
            rows.extend(_flatten_audit_tree(item, f"{path}[{index}]"))
        return rows
    if value is None:
        return [(path, None, "null")]
    value_type = "boolean" if isinstance(value, bool) else type(value).__name__
    return [(path, value, value_type)]


def _audit_raw_text(value: Any) -> str:
    """Return a lossless, type-accompanied scalar representation for Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        return repr(value)
    return str(value)


def _is_xml_10_text(value: str) -> bool:
    """Return whether every character can be represented in an XML 1.0 cell."""
    for char in value:
        codepoint = ord(char)
        if not (
            codepoint in (0x09, 0x0A, 0x0D)
            or 0x20 <= codepoint <= 0xD7FF
            or 0xE000 <= codepoint <= 0xFFFD
            or 0x10000 <= codepoint <= 0x10FFFF
        ):
            return False
    return True


def _excel_audit_text(value: str) -> tuple[str, str, int, int]:
    """Return reversible XML-safe text plus its reconstruction metadata.

    Safe values remain human-readable.  A value containing any XML 1.0-illegal
    character is encoded as base64 over UTF-8 with ``surrogatepass`` so even an
    isolated surrogate has a deterministic reversible representation.  Encoding
    happens before cell-limit chunking.
    """
    raw_bytes = value.encode("utf-8", errors="surrogatepass")
    if _is_xml_10_text(value):
        return value, _AUDIT_TEXT_PLAIN, len(value), len(raw_bytes)
    encoded = base64.b64encode(raw_bytes).decode("ascii")
    return encoded, _AUDIT_TEXT_BASE64, len(value), len(raw_bytes)


def _ordered_text_chunks(value: str) -> tuple[str, ...]:
    """Split text at Excel's cell limit without dropping an empty value."""
    if not value:
        return ("",)
    return tuple(
        value[start : start + _EXCEL_CELL_TEXT_LIMIT]
        for start in range(0, len(value), _EXCEL_CELL_TEXT_LIMIT)
    )


def _write_explicit_text_cell(cell, value: str) -> None:
    """Store exact raw text while preventing spreadsheet formula evaluation."""
    cell.value = value
    cell.data_type = "s"
    cell.quotePrefix = True


def _build_project_case_assumptions_sheet(ws, result: RunResult) -> None:
    """Write the complete RunResult provenance as a path-addressable audit trail."""
    ws.title = _PROJECT_CASE_ASSUMPTIONS_SHEET
    audit_root = {
        "run_result": {
            "schema_version": result.schema_version,
            "input_fingerprint": result.input_fingerprint,
        },
        "provenance": result.provenance,
    }
    rows = []
    for key in sorted(audit_root):
        rows.extend(_flatten_audit_tree(audit_root[key], key))
    _write_header_row(
        ws,
        1,
        [
            "Path",
            "Value (encoded text)",
            "Value Type",
            "Chunk Index",
            "Chunk Count",
            "Text Encoding",
            "Original Character Count",
            "Original UTF-8 Byte Count",
        ],
    )
    output_row = 2
    for path, value, value_type in rows:
        encoded, encoding, character_count, byte_count = _excel_audit_text(
            _audit_raw_text(value)
        )
        chunks = _ordered_text_chunks(encoded)
        for chunk_index, chunk in enumerate(chunks, start=1):
            _write_explicit_text_cell(ws.cell(row=output_row, column=1), path)
            _write_explicit_text_cell(ws.cell(row=output_row, column=2), chunk)
            _write_explicit_text_cell(ws.cell(row=output_row, column=3), value_type)
            ws.cell(row=output_row, column=4, value=chunk_index)
            ws.cell(row=output_row, column=5, value=len(chunks))
            _write_explicit_text_cell(ws.cell(row=output_row, column=6), encoding)
            ws.cell(row=output_row, column=7, value=character_count)
            ws.cell(row=output_row, column=8, value=byte_count)
            output_row += 1
    _auto_column_width(ws)


def _installed_capex_from_provenance(result: RunResult) -> float:
    """Read the typed input CapEx carried inside PC-B's immutable provenance."""
    try:
        value = result.provenance["project_case"]["asset_case"][
            "installed_capex_eur"
        ]
    except (KeyError, TypeError) as exc:
        raise ValueError(
            "RunResult provenance has no project_case.asset_case.installed_capex_eur"
        ) from exc
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError("RunResult installed_capex_eur provenance must be numeric")
    capex = float(value)
    if not math.isfinite(capex) or capex < 0.0:
        raise ValueError("RunResult installed_capex_eur provenance must be finite and >= 0")
    return capex


def _write_project_case_sheets(
    wb: Workbook,
    result: RunResult,
    *,
    use_active_sheet: bool = False,
) -> None:
    """Append the stable PC-C worksheet set to ``wb``."""
    _project_case_red_lines(result)
    _validate_project_case_cashflow_replay(result)
    npv_ws = wb.active if use_active_sheet else wb.create_sheet()
    _build_project_case_npv_sheet(npv_ws, result)

    # RunResult's state matrix guarantees screening availability/table presence.
    screening_table = result.screening_cashflow_table
    if screening_table is None:
        raise ValueError("RunResult has no screening_cashflow_table")
    installed_capex_eur = _installed_capex_from_provenance(result)
    screening_distribution = result.no_lifecycle_cost_screening_npv.distribution
    if screening_distribution is None:
        raise ValueError("available screening NPV has no distribution")
    _build_project_case_cashflow_sheet(
        wb.create_sheet(),
        _SCREENING_CASHFLOW_SHEET,
        screening_table,
        installed_capex_eur=installed_capex_eur,
        reported_p50_eur=screening_distribution.p50,
    )

    # Branch only on the typed NpvOutcome availability flag. UNKNOWN therefore
    # gets blank lifecycle NPV numbers above and no misleading empty/zero sheet.
    if result.lifecycle_cash_npv.available:
        lifecycle_table = result.lifecycle_cashflow_table
        if lifecycle_table is None:
            raise ValueError("available lifecycle NPV has no lifecycle_cashflow_table")
        lifecycle_distribution = result.lifecycle_cash_npv.distribution
        if lifecycle_distribution is None:
            raise ValueError("available lifecycle NPV has no distribution")
        _build_project_case_cashflow_sheet(
            wb.create_sheet(),
            _LIFECYCLE_CASHFLOW_SHEET,
            lifecycle_table,
            installed_capex_eur=installed_capex_eur,
            reported_p50_eur=lifecycle_distribution.p50,
        )
    if result.provenance["project_case"]["contract_case"] is not None:
        _build_contract_settlement_sheet(wb.create_sheet(), result)
    _build_project_case_assumptions_sheet(wb.create_sheet(), result)


# ── Public API ───────────────────────────────────────────────────────────────

def _write_excel_workbook(
    writer: pd.ExcelWriter,
    zone: str,
    price_df: pd.DataFrame,
    daily_spreads: pd.DataFrame,
    monthly_spreads: pd.DataFrame,
    percentiles: dict[str, float],
    revenue_estimate: dict[str, float],
    negative_stats: dict[str, float],
    tz: str | None = None,
    project_case_result: RunResult | None = None,
) -> None:
    """Populate the workbook opened via pandas ExcelWriter."""
    wb = writer.book
    summary_ws = wb.active if wb.active is not None else wb.create_sheet()

    # Recompute negative-price stats and the heatmap from the same complete-
    # local-day subset that calculate_daily_spreads uses, so a third-party
    # caller passing raw price_df + raw negative_stats can't end up with an
    # Excel where the summary counts hours from days the spread tables drop.
    # An empty subset (every day was excluded) deliberately yields zero
    # negatives — the fallback to caller stats would re-leak the inconsistency.
    complete_df = filter_to_complete_local_days(price_df, tz=tz)
    consistent_neg_stats = calculate_negative_price_hours(complete_df)

    _build_summary_sheet(
        summary_ws,
        zone,
        price_df,
        percentiles,
        revenue_estimate,
        consistent_neg_stats,
        tz=tz,
    )
    _build_table_sheet(wb.create_sheet(), "Daily Spreads", daily_spreads)
    _build_table_sheet(wb.create_sheet(), "Monthly Summary", monthly_spreads)

    hourly = price_df[["price_eur_mwh"]].reset_index()
    hourly["timestamp"] = hourly["timestamp"].astype(str)
    _build_table_sheet(wb.create_sheet(), "Hourly Prices", hourly)

    heatmap = build_price_heatmap(complete_df, tz=tz)
    _build_heatmap_sheet(wb.create_sheet(), "Price Heatmap", heatmap)
    if project_case_result is not None:
        _write_project_case_sheets(wb, project_case_result)

def export_to_excel(
    zone: str,
    price_df: pd.DataFrame,
    daily_spreads: pd.DataFrame,
    monthly_spreads: pd.DataFrame,
    percentiles: dict[str, float],
    revenue_estimate: dict[str, float],
    negative_stats: dict[str, float],
    output_path: Path | None = None,
    tz: str | None = None,
    project_case_result: RunResult | None = None,
) -> Path:
    """Export all analytics to a formatted Excel workbook.

    Args:
        zone: Bidding zone code.
        price_df: Cleaned price DataFrame.
        daily_spreads: Daily spread DataFrame.
        monthly_spreads: Monthly aggregated spreads.
        percentiles: Spread percentile dict.
        revenue_estimate: Revenue estimate dict.
        negative_stats: Negative price stats dict.
        output_path: Optional output path. Auto-generated if None.
        tz: IANA timezone for local-time date display and heatmap.
        project_case_result: Optional typed Project Case result. When provided,
            append the PC-C NPV, cash-flow and assumptions/provenance sheets.

    Returns:
        Path to the created .xlsx file.
    """
    if output_path is None:
        dates = price_df.index.tz_convert(tz) if tz else price_df.index
        start_str = dates.min().strftime("%Y%m%d")
        end_str = dates.max().strftime("%Y%m%d")
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = CACHE_DIR / f"{zone}_{start_str}_{end_str}_report.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_excel_workbook(
            writer,
            zone,
            price_df,
            daily_spreads,
            monthly_spreads,
            percentiles,
            revenue_estimate,
            negative_stats,
            tz=tz,
            project_case_result=project_case_result,
        )
    return output_path


def export_to_bytes(
    zone: str,
    price_df: pd.DataFrame,
    daily_spreads: pd.DataFrame,
    monthly_spreads: pd.DataFrame,
    percentiles: dict[str, float],
    revenue_estimate: dict[str, float],
    negative_stats: dict[str, float],
    tz: str | None = None,
    project_case_result: RunResult | None = None,
) -> bytes:
    """Export to in-memory bytes for Streamlit download button.

    Args:
        Same as export_to_excel, plus tz for local-time display.

    Returns:
        Bytes content of the .xlsx file.
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_excel_workbook(
            writer,
            zone,
            price_df,
            daily_spreads,
            monthly_spreads,
            percentiles,
            revenue_estimate,
            negative_stats,
            tz=tz,
            project_case_result=project_case_result,
        )
    return buf.getvalue()


def project_case_to_excel(result: RunResult) -> bytes:
    """Export one typed Project Case RunResult as a standalone audit workbook.

    The workbook contains only values carried by ``RunResult`` and its immutable
    provenance. The wear-net contracted-floor comparator is intentionally absent;
    only the fingerprinted PC-D strategy-cash settlement may enter Project Case
    cash. An unavailable lifecycle result keeps its numeric cells blank and omits
    the lifecycle cash-flow sheet rather than presenting it as EUR 0.
    """
    wb = Workbook()
    _write_project_case_sheets(wb, result, use_active_sheet=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF export ───────────────────────────────────────────────────────────────


def _render_figure_to_image(
    fig: Any,
    width: int = 1200,
    height: int = 500,
) -> bytes:
    """Convert a Plotly figure to PNG bytes via kaleido."""
    return fig.to_image(format="png", width=width, height=height)


def _build_pdf_report(
    zone: str,
    price_df: pd.DataFrame,
    percentiles: dict[str, float],
    revenue_estimate: dict[str, float],
    negative_stats: dict[str, float],
    tz: str | None = None,
    figures: dict[str, Any] | None = None,
) -> bytes:
    """Lay out a multi-page PDF report using fpdf2."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    # ── Page 1: Summary ─────────────────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 14, "BESS Pulse - Market Report", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 11)
    dates = price_df.index.tz_convert(tz) if tz else price_df.index
    start_str = str(dates.min().date())
    end_str = str(dates.max().date())
    total_days = (dates.max().date() - dates.min().date()).days + 1
    quality = summarize_price_data_quality(price_df)
    # Recompute neg-stats from the same complete-local-day subset that the
    # spread tables use so the PDF summary stays internally consistent even
    # when the caller passed unfiltered stats. An empty subset deliberately
    # zeros the stats — the previous fallback to caller-supplied numbers
    # could re-leak the inconsistency it was meant to prevent.
    complete_df = filter_to_complete_local_days(price_df, tz=tz)
    negative_stats = calculate_negative_price_hours(complete_df)

    rows: list[tuple[str, str]] = [
        ("Zone", zone),
        ("Timezone", tz or "UTC"),
        ("Date Range", f"{start_str}  to  {end_str}"),
        ("Total Days", str(total_days)),
    ]
    if quality["source_gap_intervals"] > 0:
        rows.extend([
            ("Source Gap Intervals", str(quality["source_gap_intervals"])),
            ("Short-Gap Imputed Intervals", str(quality["imputed_intervals"])),
            ("Unresolved Missing Intervals", str(quality["missing_intervals"])),
            ("Unresolved Missing %", f"{quality['missing_ratio']:.1%}"),
            ("Max Source Gap (hours)", str(quality["max_source_gap_hours"])),
        ])
    rows.extend([
        ("", ""),
        ("Avg Price (EUR/MWh)", f"{price_df['price_eur_mwh'].mean():.2f}"),
        ("Median Price (EUR/MWh)", f"{price_df['price_eur_mwh'].median():.2f}"),
        ("", ""),
        ("50th-percentile Spread", f"{percentiles['p50']:.2f}"),
        ("75th-percentile Spread", f"{percentiles['p75']:.2f}"),
        ("90th-percentile Spread", f"{percentiles['p90']:.2f}"),
        ("Mean Spread", f"{percentiles['mean']:.2f}"),
        ("", ""),
        ("Est. Annual Revenue (EUR/MW)",
         f"{revenue_estimate['annual_revenue_eur_per_mw']:,.0f}"),
    ])

    if "total_eur" in revenue_estimate:
        rows.append(("Headline Annual Revenue (EUR)",
                      f"{revenue_estimate['total_eur']:,.0f}"))
    if "headline_total_mode" in revenue_estimate:
        rows.append(("Headline Total Mode",
                      str(revenue_estimate["headline_total_mode"])))
    if "gross_additive_total_eur" in revenue_estimate:
        rows.append(("Gross Additive Total (Reference, EUR)",
                      f"{revenue_estimate['gross_additive_total_eur']:,.0f}"))
    if revenue_estimate.get("capacity_stack_warning"):
        rows.append(("Note", str(revenue_estimate["capacity_stack_warning"])))
    if "joint_cooptimized_total_eur" in revenue_estimate:
        rows.append(("Joint MILP Co-optimized Total (EUR)",
                     f"{revenue_estimate['joint_cooptimized_total_eur']:,.0f}"))
        rows.append(("Joint MILP Avg Reserve Commitment",
                     f"{revenue_estimate.get('joint_cooptimized_avg_reserve_fraction', 0.0):.0%}"))
    if "source_revenues" in revenue_estimate:
        rows.append(("", ""))
        for source, value in revenue_estimate["source_revenues"].items():
            rows.append((f"{source} Revenue (EUR)", f"{value:,.0f}"))

    rows.append(("", ""))
    rows.append(("Modeled Cycles per Day",
                  str(revenue_estimate.get("cycles_per_day_assumption", ""))))
    rows.append(("Capture Rate Assumption",
                  str(revenue_estimate.get("capture_rate_assumption", ""))))
    if "capture_basis" in revenue_estimate:
        rows.append(("Capture Basis", str(revenue_estimate["capture_basis"])))
    if "roundtrip_efficiency" in revenue_estimate:
        rows.append(("Round-Trip Efficiency",
                      f"{revenue_estimate['roundtrip_efficiency']:.0%}"))
    if "annual_degradation_cost_eur" in revenue_estimate:
        rows.append(("Annual Shadow Wear Proxy (EUR)",
                     f"{revenue_estimate['annual_degradation_cost_eur']:,.0f}"))
    if "net_revenue_eur" in revenue_estimate:
        rows.append(("Economic Margin after Shadow Wear (EUR)",
                     f"{revenue_estimate['net_revenue_eur']:,.0f}"))
    if "degradation_pct" in revenue_estimate:
        rows.append(("Shadow Wear % of Gross Revenue",
                     f"{revenue_estimate['degradation_pct']:.1f}%"))
    if "effective_life_years" in revenue_estimate:
        rows.append(("Effective Battery Lifetime (years)",
                     f"{revenue_estimate['effective_life_years']:.1f}"))
    if "lifetime_limiting_factor" in revenue_estimate:
        rows.append(("Lifetime Limiting Factor",
                     str(revenue_estimate["lifetime_limiting_factor"])))
    if "annual_throughput_mwh" in revenue_estimate:
        rows.append(("Annual Throughput (MWh)",
                     f"{revenue_estimate['annual_throughput_mwh']:,.0f}"))
    if "lcos_eur_mwh" in revenue_estimate:
        rows.append(("Two-leg Throughput Cost (EUR/MWh)",
                     f"{revenue_estimate['lcos_eur_mwh']:,.0f}"))
    if "net_payback_years" in revenue_estimate:
        net_payback = revenue_estimate["net_payback_years"]
        rows.append(("Economic Payback Proxy (years)",
                     f"{net_payback:.1f}" if math.isfinite(net_payback) else "N/A"))
    if "cash_npv_includes_shadow_wear" in revenue_estimate:
        rows.append(("Cash NPV Includes Shadow Wear",
                     str(revenue_estimate["cash_npv_includes_shadow_wear"])))
    rows.append(("", ""))
    rows.append(("Negative Price Hours",
                  str(negative_stats.get("negative_hours", 0))))
    rows.append(("Negative Price Intervals",
                  str(negative_stats.get("negative_intervals", 0))))
    rows.append(("Negative Price % of Intervals",
                  f"{negative_stats.get('pct_negative', 0):.1f}%"))

    col_w = 100
    for label, value in rows:
        if not label and not value:
            pdf.ln(3)
            continue
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(col_w, 7, label)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 7, value, new_x="LMARGIN", new_y="NEXT")

    # ── Chart pages ─────────────────────────────────────────────────────
    if figures:
        chart_order = [
            ("Price Time Series", "price_ts"),
            ("Price Heatmap", "price_heatmap"),
            ("Daily Spread", "spread_ts"),
            ("Revenue Breakdown", "revenue_bar"),
            ("Revenue Waterfall", "revenue_waterfall"),
            ("Spread Heatmap", "spread_heatmap"),
            ("Monthly Seasonality", "monthly_seasonality"),
        ]
        for title, key in chart_order:
            fig = figures.get(key)
            if fig is None:
                continue
            try:
                img_bytes = _render_figure_to_image(fig)
            except Exception:
                logger.warning("Failed to render chart '%s' for PDF", key)
                continue

            pdf.add_page()
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp.write(img_bytes)
                tmp.flush()
                page_w = pdf.w - pdf.l_margin - pdf.r_margin
                page_h = pdf.h - pdf.get_y() - 15
                img_w = page_w
                img_h = img_w * 500 / 1200
                if img_h > page_h:
                    img_h = page_h
                    img_w = img_h * 1200 / 500
                pdf.image(tmp.name, w=img_w, h=img_h)
                Path(tmp.name).unlink(missing_ok=True)

    return bytes(pdf.output())


def export_to_pdf_bytes(
    zone: str,
    price_df: pd.DataFrame,
    daily_spreads: pd.DataFrame,
    monthly_spreads: pd.DataFrame,
    percentiles: dict[str, float],
    revenue_estimate: dict[str, float],
    negative_stats: dict[str, float],
    tz: str | None = None,
    figures: dict[str, Any] | None = None,
) -> bytes:
    """Export analytics to a PDF report as in-memory bytes.

    Args:
        zone: Bidding zone code.
        price_df: Cleaned price DataFrame.
        daily_spreads: Daily spread DataFrame.
        monthly_spreads: Monthly aggregated spreads.
        percentiles: Spread percentile dict.
        revenue_estimate: Revenue estimate dict.
        negative_stats: Negative price stats dict.
        tz: IANA timezone for local-time date display.
        figures: Optional dict mapping chart keys to Plotly figures.
            Recognized keys: price_ts, price_heatmap, spread_ts,
            revenue_bar, revenue_waterfall, spread_heatmap,
            monthly_seasonality.

    Returns:
        Bytes content of the PDF file.
    """
    return _build_pdf_report(
        zone=zone,
        price_df=price_df,
        percentiles=percentiles,
        revenue_estimate=revenue_estimate,
        negative_stats=negative_stats,
        tz=tz,
        figures=figures,
    )


# ── Comparison export ────────────────────────────────────────────────────────

_COMPARISON_COLUMNS = {
    "zone": ("Zone", None),
    "avg_price": ("Avg Price (EUR/MWh)", _PRICE_FMT),
    "std_price": ("Std Dev", _PRICE_FMT),
    "avg_spread": ("Avg Spread (EUR/MWh)", _PRICE_FMT),
    "p50_spread": ("50th-percentile Spread", _PRICE_FMT),
    "p90_spread": ("90th-percentile Spread", _PRICE_FMT),
    "negative_pct": ("Neg Price %", _PCT_FMT),
    "estimated_annual_revenue_per_mw": ("Revenue (EUR/MW/yr)", "#,##0"),
    "dispatch_method": ("Dispatch Method", None),
    "avg_cycles_per_day": ("Avg Cycles/Day", "0.00"),
    "net_revenue_per_mw": ("Economic Margin after Shadow Wear (EUR/MW/yr)", "#,##0"),
    "lcos_eur_mwh": ("Two-leg Throughput Cost (EUR/MWh)", "#,##0.0"),
    "payback_years": ("Economic Payback Proxy (years)", "0.0"),
    "effective_life_years": ("Effective Life (years)", "0.0"),
    "limiting_factor": ("Limiting Factor", None),
}


def export_comparison_to_bytes(comparison_df: pd.DataFrame) -> bytes:
    """Export zone comparison DataFrame to an Excel workbook.

    Args:
        comparison_df: Output of ``compare_zones()``.

    Returns:
        Excel file as bytes.
    """
    # Sanitise any string cells before handing to to_excel so a future
    # callers that pass uploaded data (today everything is internal) cannot
    # smuggle a leading-=/+/-/@ formula through this path. Codex flagged
    # this as defence-in-depth — no live exploit, but cheap to harden.
    safe_df = comparison_df.copy()
    for col in safe_df.select_dtypes(include=["object", "string"]).columns:
        safe_df[col] = safe_df[col].map(_safe_cell_value)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        safe_df.to_excel(writer, sheet_name="Zone Comparison", index=False)
        ws = writer.sheets["Zone Comparison"]

        # Apply styled headers and number formats. Analytics stores negative_pct
        # as percentage points for UI readability; Excel percent cells need ratios.
        for col_idx, col_name in enumerate(safe_df.columns, 1):
            label, fmt = _COMPARISON_COLUMNS.get(col_name, (col_name, None))
            cell = ws.cell(row=1, column=col_idx, value=_safe_cell_value(label))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            if fmt:
                for row_idx in range(2, len(safe_df) + 2):
                    data_cell = ws.cell(row=row_idx, column=col_idx)
                    if col_name == "negative_pct" and data_cell.value is not None:
                        data_cell.value = data_cell.value / 100
                    data_cell.number_format = fmt

        _auto_column_width(ws)

    return buf.getvalue()


def cockpit_tables_to_excel(
    tables: dict[str, pd.DataFrame],
    *,
    assumptions: pd.DataFrame | None = None,
) -> bytes:
    """Bundle Simulation Cockpit result tables into a self-documenting xlsx.

    Each ``{sheet_name: DataFrame}`` entry becomes a formatted sheet (reusing
    the report's table formatter, so string cells are formula-safe); a
    non-empty ``assumptions`` table is appended as an 'Assumptions' sheet so
    the exported file records the haircuts behind the numbers. Empty / None
    tables are skipped; returns empty bytes when there is nothing to write.

    Args:
        tables: Ordered mapping of Excel sheet name to result DataFrame.
        assumptions: Optional model-assumption audit table.

    Returns:
        Workbook bytes for a Streamlit download button (b"" when empty).
    """
    sheets = [
        (name, df) for name, df in tables.items()
        if df is not None and not df.empty
    ]
    if assumptions is not None and not assumptions.empty:
        sheets.append(("Assumptions", assumptions))
    if not sheets:
        return b""

    wb = Workbook()
    for i, (name, df) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        # Excel caps sheet names at 31 chars.
        _build_table_sheet(ws, name[:31], df)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
